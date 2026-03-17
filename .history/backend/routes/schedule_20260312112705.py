"""
排课调度API - 含节假日检查、调整、合班功能
"""
from flask import Blueprint, jsonify, request
from datetime import date, datetime, timedelta
from models import db, ClassSchedule, Class, MonthlyPlan, TeacherCourseCombo, Topic, ScheduleConstraint
import requests
import json as _json

schedule_bp = Blueprint('schedule', __name__)


def _month_range(year, month):
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)
    return start_date, end_date


def _guess_suggestion(reason):
    text = reason or ''
    if '节假日' in text:
        return '建议顺延到下一周周六/周日'
    if '请假' in text or '不可用' in text:
        return '建议改用同课题备选讲师，或顺延一周'
    if '班主任' in text:
        return '建议调整到班主任可到场的下一周末'
    if '撞课' in text:
        return '建议同周内错开班级，或调整到下一周'
    if '日期被排除' in text:
        return '建议改到最近可用周末'
    return '建议人工调整时间或师资后再发布'


def _build_publish_checklist(year, month):
    start_date, end_date = _month_range(year, month)
    month_schedules = ClassSchedule.query.filter(
        ClassSchedule.scheduled_date >= start_date,
        ClassSchedule.scheduled_date < end_date
    ).all()

    unresolved = []
    pending = []
    resolved = []

    for s in month_schedules:
        if s.status == 'conflict':
            reason = s.notes or '冲突原因未记录'
            same_day_count = ClassSchedule.query.filter(
                ClassSchedule.scheduled_date == s.scheduled_date
            ).count()
            unresolved.append({
                'schedule_id': s.id,
                'class_name': s.class_.name if s.class_ else None,
                'topic_name': s.topic.name if s.topic else None,
                'date': s.scheduled_date.isoformat() if s.scheduled_date else None,
                'reason': reason,
                'suggestion': _guess_suggestion(reason),
                'impact_scope': f'影响日期 {s.scheduled_date.isoformat()}，同日排课 {max(same_day_count - 1, 0)} 条'
            })
        elif s.status in ['scheduled']:
            # 只有真正需要关注的才放入 pending（缺讲师、缺日期等）
            needs_attention = False
            attention_reason = ''
            if not s.combo_id:
                needs_attention = True
                attention_reason = '未分配讲师（教-课组合）'
            elif not s.scheduled_date:
                needs_attention = True
                attention_reason = '未设定上课日期'

            if needs_attention:
                pending.append({
                    'schedule_id': s.id,
                    'class_name': s.class_.name if s.class_ else None,
                    'topic_name': s.topic.name if s.topic else None,
                    'date': s.scheduled_date.isoformat() if s.scheduled_date else None,
                    'reason': attention_reason,
                    'suggestion': '建议发布前补全信息',
                    'impact_scope': '仅影响当前班级'
                })
            else:
                resolved.append({
                    'schedule_id': s.id,
                    'class_name': s.class_.name if s.class_ else None,
                    'topic_name': s.topic.name if s.topic else None,
                    'date': s.scheduled_date.isoformat() if s.scheduled_date else None
                })
        else:
            resolved.append({
                'schedule_id': s.id,
                'class_name': s.class_.name if s.class_ else None,
                'topic_name': s.topic.name if s.topic else None,
                'date': s.scheduled_date.isoformat() if s.scheduled_date else None
            })

    # 主动扫描：检查同日同讲师/同班主任未标记的冲突
    seen_conflict_ids = set(item['schedule_id'] for item in unresolved)
    from collections import defaultdict
    by_date = defaultdict(list)
    for s in month_schedules:
        if s.id not in seen_conflict_ids and s.status not in ('completed', 'cancelled'):
            by_date[s.scheduled_date].append(s)

    for dt, day_schedules in by_date.items():
        if len(day_schedules) < 2:
            continue
        # 检查周六讲师冲突
        teacher_day1 = defaultdict(list)
        for s in day_schedules:
            if s.combo and s.combo.teacher_id:
                teacher_day1[s.combo.teacher_id].append(s)
        for tid, slist in teacher_day1.items():
            if len(slist) >= 2:
                names = ', '.join(s.class_.name for s in slist if s.class_)
                teacher_name = slist[0].combo.teacher.name if slist[0].combo and slist[0].combo.teacher else '未知'
                pending.append({
                    'schedule_id': slist[0].id,
                    'class_name': names,
                    'topic_name': None,
                    'date': dt.isoformat() if dt else None,
                    'reason': f'周六讲师 {teacher_name} 同日为 {len(slist)} 个班级授课',
                    'suggestion': '建议错开日期或更换讲师',
                    'impact_scope': f'涉及 {len(slist)} 个班级'
                })
        # 检查周日讲师冲突
        teacher_day2 = defaultdict(list)
        for s in day_schedules:
            if s.combo_2 and s.combo_2.teacher_id:
                teacher_day2[s.combo_2.teacher_id].append(s)
        for tid, slist in teacher_day2.items():
            if len(slist) >= 2:
                names = ', '.join(s.class_.name for s in slist if s.class_)
                teacher_name = slist[0].combo_2.teacher.name if slist[0].combo_2 and slist[0].combo_2.teacher else '未知'
                pending.append({
                    'schedule_id': slist[0].id,
                    'class_name': names,
                    'topic_name': None,
                    'date': dt.isoformat() if dt else None,
                    'reason': f'周日讲师 {teacher_name} 同日为 {len(slist)} 个班级授课',
                    'suggestion': '建议错开日期或更换讲师',
                    'impact_scope': f'涉及 {len(slist)} 个班级'
                })
        # 检查班主任冲突
        homeroom_map = defaultdict(list)
        for s in day_schedules:
            if s.class_ and s.class_.homeroom_id:
                homeroom_map[s.class_.homeroom_id].append(s)
        for hid, slist in homeroom_map.items():
            if len(slist) >= 2:
                names = ', '.join(s.class_.name for s in slist if s.class_)
                hr_name = slist[0].class_.homeroom.name if slist[0].class_ and slist[0].class_.homeroom else '未知'
                pending.append({
                    'schedule_id': slist[0].id,
                    'class_name': names,
                    'topic_name': None,
                    'date': dt.isoformat() if dt else None,
                    'reason': f'班主任 {hr_name} 同日负责 {len(slist)} 个班级',
                    'suggestion': '如为合班属正常情况，否则建议错开日期',
                    'impact_scope': f'涉及 {len(slist)} 个班级'
                })

    return {
        'resolved': resolved,
        'pending': pending,
        'unresolved': unresolved
    }

# ==================== 节假日相关 ====================

import os
import json

# 加载本地节假日缓存（按年份动态查找 holidays_{year}.json）
_local_holiday_data = {}
_holiday_files_loaded = set()
_backend_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def _load_holiday_file(year):
    """按年份加载对应的节假日缓存文件"""
    if year in _holiday_files_loaded:
        return
    _holiday_files_loaded.add(year)
    local_path = os.path.join(_backend_dir, f'holidays_{year}.json')
    try:
        with open(local_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            _local_holiday_data.update(data)
            print(f"Loaded {len(data)} holiday records from: {local_path}")
    except FileNotFoundError:
        print(f"Info: No holiday file for {year}: {local_path}")
    except Exception as e:
        print(f"Warning: Could not load holiday file {local_path}: {e}")

# 启动时预加载当前年份
from datetime import datetime as _dt
_load_holiday_file(_dt.now().year)

_holiday_cache = {}

def is_holiday(check_date):
    """
    检查某日期是否为节假日
    优先使用本地 holidays_{year}.json（按年份自动加载）
    其次使用 timor.tech API
    最后兜底逻辑：默认为工作日（排课日）
    """
    if isinstance(check_date, str):
        check_date = date.fromisoformat(check_date)
    
    # 动态加载该年份的节假日缓存
    _load_holiday_file(check_date.year)
    
    date_str = check_date.isoformat()
    mm_dd = date_str[5:] # MM-DD
    
    # 1. 检查内存缓存
    if date_str in _holiday_cache:
        return _holiday_cache[date_str]
    
    # 2. 检查本地文件缓存 (优先)
    # 尝试 YYYY-MM-DD
    if date_str in _local_holiday_data:
        record = _local_holiday_data[date_str]
        # treat official holidays and any record that represents a make-up/adjusted workday as restricted
        is_makeup = ('after' in record) or ('补班' in (record.get('name') or '')) or record.get('workday', False)
        is_hol = bool(record.get('holiday', False) or is_makeup)
        _holiday_cache[date_str] = is_hol
        return is_hol

    # 尝试 MM-DD (holidays_2026.json 格式)
    if mm_dd in _local_holiday_data:
        record = _local_holiday_data[mm_dd]
        is_makeup = ('after' in record) or ('补班' in (record.get('name') or '')) or record.get('workday', False)
        is_hol = bool(record.get('holiday', False) or is_makeup)
        # 注意：这里我们存入缓存的是完整日期
        _holiday_cache[date_str] = is_hol
        return is_hol

    # 3. 检查API (作为补充)
    try:
        resp = requests.get(f"https://timor.tech/api/holiday/info/{date_str}", timeout=2) # 缩短超时
        data = resp.json()
        
        if data.get('code') == 0:
            holiday_info = data.get('holiday')
            is_hol = holiday_info.get('holiday', False) if holiday_info else False
            _holiday_cache[date_str] = is_hol
            return is_hol
    except Exception as e:
        # print(f"节假日API调用失败: {e}") # 减少日志干扰
        pass
    
    # 4. 兜底逻辑
    # 商学院排课主要在周六日，所以默认这些天不是节假日（除非API已明确说是）
    # 只有API或本地文件明确说是holiday: true，才是节假日
    # 否则默认都是工作日（可以排课）
    is_hol = False
    _holiday_cache[date_str] = is_hol
    return is_hol

def find_next_available_saturday(start_date):
    """找到下一个周六"""
    if isinstance(start_date, str):
        start_date = date.fromisoformat(start_date)
    
    # 找到下一个周六 (weekday 5 = Saturday)
    days_until_saturday = (5 - start_date.weekday()) % 7
    if days_until_saturday == 0 and start_date.weekday() != 5:
        days_until_saturday = 7
    
    return start_date + timedelta(days=days_until_saturday)


# ==================== 月度排课查询 ====================

@schedule_bp.route('/month/<int:year>/<int:month>', methods=['GET'])
def get_month_schedule(year, month):
    """获取指定月份的所有排课"""
    # 自动同步状态（纠正错误的completed标记）
    from .classes import sync_class_statuses
    sync_class_statuses()
    
    # 计算月份的起止日期
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)
    
    schedules = ClassSchedule.query.filter(
        ClassSchedule.scheduled_date >= start_date,
        ClassSchedule.scheduled_date < end_date
    ).order_by(ClassSchedule.scheduled_date).all()
    
    # 生成该月所有周的结构
    weeks_data = []
    
    # 找到该月第一天所在的周一
    # weekday(): 0=Mon, 6=Sun
    curr_week_start = start_date - timedelta(days=start_date.weekday())
    
    while curr_week_start < end_date:
        week_end = curr_week_start + timedelta(days=6)
        week_key = curr_week_start.isoformat()
        
        # 查找该周的课程，并附加间隔信息
        week_schedules = []
        for s in schedules:
            if not (curr_week_start <= s.scheduled_date <= week_end):
                continue
            s_dict = s.to_dict()
            # 查询该班级上一次课的日期（在当前课之前的最晚一次）
            prev = ClassSchedule.query.filter(
                ClassSchedule.class_id == s.class_id,
                ClassSchedule.scheduled_date < s.scheduled_date,
                ClassSchedule.status.in_(['completed', 'confirmed', 'scheduled', 'planning'])
            ).order_by(ClassSchedule.scheduled_date.desc()).first()
            if prev:
                s_dict['last_date'] = prev.scheduled_date.isoformat()
                s_dict['interval_days'] = (s.scheduled_date - prev.scheduled_date).days
            else:
                s_dict['last_date'] = None
                s_dict['interval_days'] = None
            week_schedules.append(s_dict)
        
        weeks_data.append({
            'week_start': week_key,
            'week_end': week_end.isoformat(),
            'schedules': week_schedules
        })
        
        curr_week_start += timedelta(weeks=1)
    
    # week_start > end_date 时停止，但可能最后一周跨月，只要 week_start < end_date 就会包含
    
    # 获取月度计划状态
    plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
    
    return jsonify({
        'year': year,
        'month': month,
        'plan_status': plan.status if plan else 'draft',
        'published_at': plan.published_at.isoformat() if plan and plan.published_at else None,
        'last_saved_at': plan.updated_at.isoformat() if plan and plan.updated_at else None,
        'weeks': weeks_data
    })


# ==================== 排课调整 ====================

@schedule_bp.route('/adjust', methods=['POST'])
def adjust_schedule():
    """调整单节课的日期或教-课组合"""
    data = request.get_json()
    schedule_id = data.get('schedule_id')
    force = data.get('force', False)
    
    schedule = ClassSchedule.query.get_or_404(schedule_id)
    
    # 状态保护：已完成/已取消的课程不允许调整
    if schedule.status in ('completed', 'cancelled'):
        return jsonify({'error': f'该课程已{schedule.status}，无法调整'}), 400
    
    # 强制模式：跳过所有冲突检查，直接应用修改并标记为冲突
    if force:
        conflict_notes = []
        if 'new_date' in data:
            schedule.scheduled_date = date.fromisoformat(data['new_date'])
        if 'combo_id' in data:
            old_combo_id = schedule.combo_id
            schedule.combo_id = data['combo_id']
            new_combo = TeacherCourseCombo.query.get(data['combo_id']) if data['combo_id'] else None
            if new_combo:
                conflict_notes.append(f'周六讲师强制改为 {new_combo.teacher.name}')
        if 'combo_id_2' in data:
            schedule.combo_id_2 = data['combo_id_2']
            new_combo2 = TeacherCourseCombo.query.get(data['combo_id_2']) if data['combo_id_2'] else None
            if new_combo2:
                conflict_notes.append(f'周日讲师强制改为 {new_combo2.teacher.name}')
        schedule.status = 'conflict'
        schedule.notes = '手动强制调整: ' + '; '.join(conflict_notes) if conflict_notes else '手动强制调整'
        db.session.commit()
        return jsonify(schedule.to_dict())
    
    # 调整日期
    if 'new_date' in data:
        new_date = date.fromisoformat(data['new_date'])
        
        # 检查是否是周六
        if new_date.weekday() != 5:
            return jsonify({'error': '排课日期必须是周六'}), 400
        
        # 检查节假日
        if is_holiday(new_date):
            return jsonify({
                'warning': '该日期是节假日',
                'holiday': True,
                'proceed': False
            }), 200
        
        # 检查班主任冲突（周六及周日）
        class_obj = schedule.class_
        if class_obj and class_obj.homeroom_id:
            # check saturday
            homeroom_conflicts = ClassSchedule.query.join(Class).filter(
                Class.homeroom_id == class_obj.homeroom_id,
                ClassSchedule.scheduled_date == new_date,
                ClassSchedule.id != schedule_id
            ).all()
            if homeroom_conflicts:
                return jsonify({
                    'error': '班主任在该日期已有其他课程',
                    'conflict_type': 'homeroom',
                    'conflicts': [c.to_dict() for c in homeroom_conflicts]
                }), 409
            # check sunday (since sunday is part of same record, treat as same new_date)
            sun = new_date + timedelta(days=1)
            homeroom_conflicts2 = ClassSchedule.query.join(Class).filter(
                Class.homeroom_id == class_obj.homeroom_id,
                ClassSchedule.scheduled_date == new_date,
                ClassSchedule.id != schedule_id
            ).all()
            # above query same as saturday because sunday is stored in same record; no extra check needed
        
        # 检查讲师冲突（如果已分配教-课组合，考虑两天）
        # 讲师冲突检查：combo_id=周六讲师, combo_id_2=周日讲师，不同天应分别检查
        if schedule.combo_id and schedule.combo:
            teacher_id = schedule.combo.teacher_id
            from models import TeacherCourseCombo
            # 周六讲师只与其他记录的周六讲师(combo_id)比较
            teacher_conflicts = ClassSchedule.query.join(
                TeacherCourseCombo, ClassSchedule.combo_id == TeacherCourseCombo.id
            ).filter(
                TeacherCourseCombo.teacher_id == teacher_id,
                ClassSchedule.scheduled_date == new_date,
                ClassSchedule.id != schedule_id
            ).all()

            if teacher_conflicts:
                return jsonify({
                    'warning': f'讲师 {schedule.combo.teacher.name} 在该周六已有其他课程',
                    'conflict_type': 'teacher',
                    'conflicts': [c.to_dict() for c in teacher_conflicts],
                    'proceed': True
                }), 200
        if schedule.combo_id_2 and schedule.combo_2:
            teacher2_id = schedule.combo_2.teacher_id
            from models import TeacherCourseCombo
            # 周日讲师只与其他记录的周日讲师(combo_id_2)比较
            teacher_conflicts2 = ClassSchedule.query.filter(
                ClassSchedule.combo_id_2.isnot(None),
                ClassSchedule.scheduled_date == new_date,
                ClassSchedule.id != schedule_id
            ).all()
            teacher_conflicts2 = [s for s in teacher_conflicts2 if s.combo_2 and s.combo_2.teacher_id == teacher2_id]
            if teacher_conflicts2:
                return jsonify({
                    'warning': f'讲师 {schedule.combo_2.teacher.name} 在该周日已有其他课程',
                    'conflict_type': 'teacher',
                    'conflicts': [c.to_dict() for c in teacher_conflicts2],
                    'proceed': True
                }), 200
        
        # optional: 如果必要，还可检查新星期天是否是节假日
        sun_date = new_date + timedelta(days=1)
        if is_holiday(sun_date):
            return jsonify({
                'warning': '新周日日期为节假日',
                'holiday': True,
                'proceed': False
            }), 200
        
        schedule.scheduled_date = new_date
    
    # 调整教-课组合（周六）
    if 'combo_id' in data:
        new_combo_id = data['combo_id']
        # 检查新周六讲师是否有冲突（只与其他记录的周六讲师比较）
        if new_combo_id:
            from models import TeacherCourseCombo
            new_combo = TeacherCourseCombo.query.get(new_combo_id)
            if new_combo:
                teacher_conflicts = ClassSchedule.query.join(
                    TeacherCourseCombo, ClassSchedule.combo_id == TeacherCourseCombo.id
                ).filter(
                    TeacherCourseCombo.teacher_id == new_combo.teacher_id,
                    ClassSchedule.scheduled_date == schedule.scheduled_date,
                    ClassSchedule.id != schedule_id
                ).all()
                
                if teacher_conflicts:
                    return jsonify({
                        'warning': f'讲师 {new_combo.teacher.name} 在该周六已有其他课程',
                        'conflict_type': 'teacher',
                        'conflicts': [c.to_dict() for c in teacher_conflicts],
                        'proceed': True
                    }), 200
        
        schedule.combo_id = new_combo_id
    # 调整教-课组合（周日）
    if 'combo_id_2' in data:
        new_combo2_id = data['combo_id_2']
        # 检查新周日讲师是否有冲突（只与其他记录的周日讲师比较）
        if new_combo2_id:
            from models import TeacherCourseCombo
            new_combo2 = TeacherCourseCombo.query.get(new_combo2_id)
            if new_combo2:
                teacher_conflicts2 = ClassSchedule.query.filter(
                    ClassSchedule.combo_id_2.isnot(None),
                    ClassSchedule.scheduled_date == schedule.scheduled_date,
                    ClassSchedule.id != schedule_id
                ).all()
                teacher_conflicts2 = [s for s in teacher_conflicts2 if s.combo_2 and s.combo_2.teacher_id == new_combo2.teacher_id]
                if teacher_conflicts2:
                    return jsonify({
                        'warning': f'讲师 {new_combo2.teacher.name} 在该周日已有其他课程',
                        'conflict_type': 'teacher',
                        'conflicts': [c.to_dict() for c in teacher_conflicts2],
                        'proceed': True
                    }), 200
        schedule.combo_id_2 = new_combo2_id
    
    # 更新备注
    if 'notes' in data:
        schedule.notes = data['notes']
    
    # 成功调整后，清除冲突标记（如果之前是冲突状态）
    if schedule.status == 'conflict':
        schedule.status = 'scheduled'
        schedule.conflict_type = None
        if schedule.notes and ('撞课' in schedule.notes or '手动' in schedule.notes or '冲突' in schedule.notes):
            schedule.notes = None
    
    db.session.commit()
    return jsonify(schedule.to_dict())


@schedule_bp.route('/move-week', methods=['POST'])
def move_to_week():
    """将课程移动到上一周或下一周（含冲突/节假日检测）"""
    data = request.get_json()
    schedule_id = data.get('schedule_id')
    direction = data.get('direction', 'next')  # 'next' or 'prev'
    force = data.get('force', False)  # 前端确认后强制移动
    
    schedule = ClassSchedule.query.get_or_404(schedule_id)
    
    # 状态保护
    if schedule.status in ('completed', 'cancelled'):
        return jsonify({'error': f'该课程已{schedule.status}，无法移动'}), 400
    
    if direction == 'next':
        new_date = schedule.scheduled_date + timedelta(weeks=1)
    else:
        new_date = schedule.scheduled_date - timedelta(weeks=1)
    
    # 确保是周六
    new_date = find_next_available_saturday(new_date - timedelta(days=1))
    
    # 检查目标日期的冲突
    warnings = []
    
    # 检查节假日
    if is_holiday(new_date):
        warnings.append(f'{new_date.isoformat()} 为节假日')
    
    # 检查班主任冲突
    if schedule.class_ and schedule.class_.homeroom_id:
        homeroom_conflict = ClassSchedule.query.join(Class).filter(
            Class.homeroom_id == schedule.class_.homeroom_id,
            ClassSchedule.scheduled_date == new_date,
            ClassSchedule.class_id != schedule.class_id
        ).first()
        if homeroom_conflict:
            warnings.append(f'班主任与 {homeroom_conflict.class_.name} 撞课')
    
    # 检查讲师冲突（周六）
    if schedule.combo:
        teacher_conflict = ClassSchedule.query.filter(
            ClassSchedule.scheduled_date == new_date,
            ClassSchedule.combo_id.isnot(None),
            ClassSchedule.id != schedule.id
        ).all()
        for tc in teacher_conflict:
            if tc.combo and tc.combo.teacher_id == schedule.combo.teacher_id:
                warnings.append(f'周六讲师 {schedule.combo.teacher.name} 与 {tc.class_.name} 撞课')
                break
    
    # 检查讲师冲突（周日）
    if schedule.combo_2:
        teacher_conflict_2 = ClassSchedule.query.filter(
            ClassSchedule.scheduled_date == new_date,
            ClassSchedule.combo_id_2.isnot(None),
            ClassSchedule.id != schedule.id
        ).all()
        for tc in teacher_conflict_2:
            if tc.combo_2 and tc.combo_2.teacher_id == schedule.combo_2.teacher_id:
                warnings.append(f'周日讲师 {schedule.combo_2.teacher.name} 与 {tc.class_.name} 撞课')
                break
    
    # 如果有警告且未强制，返回警告让前端确认
    if warnings and not force:
        return jsonify({
            'confirm_required': True,
            'new_date': new_date.isoformat(),
            'warnings': warnings,
            'message': '移动目标日期存在以下问题，是否仍要移动？'
        })
    
    schedule.scheduled_date = new_date
    # 如果强制移动到有冲突的位置，更新状态
    if warnings:
        schedule.status = 'conflict'
        schedule.conflict_type = 'teacher' if '讲师' in str(warnings) else 'homeroom'
        schedule.notes = '手动移动: ' + '; '.join(warnings)
    else:
        # 无冲突移动，清除之前的冲突标记
        if schedule.status == 'conflict':
            schedule.status = 'scheduled'
            schedule.conflict_type = None
            schedule.notes = None
    db.session.commit()
    
    return jsonify(schedule.to_dict())


# ==================== 合班功能 ====================

@schedule_bp.route('/merge-info', methods=['POST'])
def merge_info():
    """获取合班可选项（日期、讲师组合、班主任）供前端弹窗展示"""
    data = request.get_json()
    schedule_ids = data.get('schedule_ids', [])
    
    schedules = ClassSchedule.query.filter(ClassSchedule.id.in_(schedule_ids)).all()
    if len(schedules) < 2:
        return jsonify({'error': '至少需要两条记录'}), 400
    
    # 收集可选日期
    dates = list(set(s.scheduled_date.isoformat() for s in schedules if s.scheduled_date))
    dates.sort()
    
    # 收集可选周六组合
    combos_day1 = []
    seen_combo1 = set()
    for s in schedules:
        if s.combo and s.combo_id not in seen_combo1:
            seen_combo1.add(s.combo_id)
            combos_day1.append({
                'combo_id': s.combo_id,
                'label': f'{s.combo.teacher.name} - {s.combo.course.name}' if s.combo.teacher and s.combo.course else f'组合#{s.combo_id}',
                'from_class': s.class_.name if s.class_ else ''
            })
    
    # 收集可选周日组合
    combos_day2 = []
    seen_combo2 = set()
    for s in schedules:
        if s.combo_2 and s.combo_id_2 not in seen_combo2:
            seen_combo2.add(s.combo_id_2)
            combos_day2.append({
                'combo_id': s.combo_id_2,
                'label': f'{s.combo_2.teacher.name} - {s.combo_2.course.name}' if s.combo_2.teacher and s.combo_2.course else f'组合#{s.combo_id_2}',
                'from_class': s.class_.name if s.class_ else ''
            })
    
    # 收集可选班主任
    homerooms = []
    seen_hr = set()
    for s in schedules:
        if s.class_ and s.class_.homeroom and s.class_.homeroom_id not in seen_hr:
            seen_hr.add(s.class_.homeroom_id)
            homerooms.append({
                'homeroom_id': s.class_.homeroom_id,
                'name': s.class_.homeroom.name,
                'from_class': s.class_.name
            })
    
    return jsonify({
        'dates': dates,
        'combos_day1': combos_day1,
        'combos_day2': combos_day2,
        'homerooms': homerooms,
        'schedules': [s.to_dict() for s in schedules]
    })


@schedule_bp.route('/merge', methods=['POST'])
def merge_classes():
    """合班操作 - 多个班级同一课题合并上课，统一讲师和日期"""
    import json
    data = request.get_json()
    schedule_ids = data.get('schedule_ids', [])
    merged_date = data.get('merged_date')
    merged_combo_id = data.get('merged_combo_id')
    merged_combo_id_2 = data.get('merged_combo_id_2')
    lead_homeroom = data.get('lead_homeroom_name', '')
    
    if len(schedule_ids) < 2:
        return jsonify({'error': '合班至少需要两个课程'}), 400
    
    schedules = ClassSchedule.query.filter(ClassSchedule.id.in_(schedule_ids)).all()
    
    if len(schedules) != len(schedule_ids):
        return jsonify({'error': '部分课程不存在'}), 404
    
    # 验证是否为同一课题
    topic_ids = set(s.topic_id for s in schedules)
    if len(topic_ids) > 1:
        return jsonify({'error': '只能合并相同课题的课程'}), 400
    
    # 使用第一个作为主课表
    main_schedule = schedules[0]
    class_names = [s.class_.name for s in schedules if s.class_]
    
    # 统一日期
    if merged_date:
        unified_date = date.fromisoformat(merged_date)
    else:
        unified_date = main_schedule.scheduled_date
    
    # 合班前保存每条记录的快照（拆分时用于恢复）
    for s in schedules:
        snapshot = {
            'scheduled_date': s.scheduled_date.isoformat() if s.scheduled_date else None,
            'combo_id': s.combo_id,
            'combo_id_2': s.combo_id_2,
            'status': s.status,
            'conflict_type': s.conflict_type,
            'notes': s.notes
        }
        s.merge_snapshot = json.dumps(snapshot, ensure_ascii=False)
    
    # 更新所有记录的统一信息
    for s in schedules:
        s.scheduled_date = unified_date
        s.status = 'merged'
        if merged_combo_id:
            s.combo_id = int(merged_combo_id)
        if merged_combo_id_2:
            s.combo_id_2 = int(merged_combo_id_2)
    
    # 设置合班关联
    for s in schedules[1:]:
        s.merged_with = main_schedule.id
        s.notes = f'合班至 {main_schedule.class_.name}' + (f'（带班: {lead_homeroom}）' if lead_homeroom else '')
    
    main_schedule.notes = f'合班主记录（含 {", ".join(class_names[1:])}）' + (f'（带班: {lead_homeroom}）' if lead_homeroom else '')
    
    db.session.commit()
    
    return jsonify({
        'main_schedule': main_schedule.to_dict(),
        'merged_schedules': [s.to_dict() for s in schedules[1:]]
    })


@schedule_bp.route('/unmerge/<int:schedule_id>', methods=['POST'])
def unmerge_class(schedule_id):
    """取消合班 - 级联清除所有相关合班记录，并从快照恢复原始状态"""
    import json
    schedule = ClassSchedule.query.get_or_404(schedule_id)
    
    def _restore_from_snapshot(record):
        """从 merge_snapshot 恢复记录的原始状态"""
        if not record.merge_snapshot:
            return
        try:
            snap = json.loads(record.merge_snapshot)
            if snap.get('scheduled_date'):
                record.scheduled_date = date.fromisoformat(snap['scheduled_date'])
            if 'combo_id' in snap:
                record.combo_id = snap['combo_id']
            if 'combo_id_2' in snap:
                record.combo_id_2 = snap['combo_id_2']
            record.status = snap.get('status', 'scheduled')
            record.conflict_type = snap.get('conflict_type')
            record.notes = snap.get('notes')
        except (json.JSONDecodeError, ValueError):
            pass
        record.merge_snapshot = None
        record.merged_with = None
    
    # 情况1: 如果是主记录（被其他记录引用），同时恢复所有次记录
    secondary_records = ClassSchedule.query.filter_by(merged_with=schedule_id).all()
    for sec in secondary_records:
        _restore_from_snapshot(sec)
    
    # 情况2: 如果是次记录（引用其他主记录），也检查主记录
    if schedule.merged_with:
        main_record = ClassSchedule.query.get(schedule.merged_with)
        if main_record:
            # 检查主记录下还有没有其他次记录
            remaining = ClassSchedule.query.filter(
                ClassSchedule.merged_with == schedule.merged_with,
                ClassSchedule.id != schedule_id
            ).count()
            if remaining == 0:
                # 没有其他次记录了，主记录也恢复
                _restore_from_snapshot(main_record)
    
    # 恢复当前记录
    _restore_from_snapshot(schedule)
    
    db.session.commit()
    return jsonify({'message': '拆分成功，已恢复原始状态', 'schedule': schedule.to_dict()})


# ==================== 月度计划发布 ====================

@schedule_bp.route('/publish-checklist', methods=['GET'])
def publish_checklist():
    """发布前冲突处置清单"""
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    if not year or not month:
        return jsonify({'error': 'Missing year/month'}), 400

    checklist = _build_publish_checklist(year, month)
    return jsonify({
        'year': year,
        'month': month,
        'checklist': checklist
    })


@schedule_bp.route('/publish', methods=['POST'])
def publish_month():
    """发布月度计划"""
    data = request.get_json()
    year = data.get('year')
    month = data.get('month')
    force_publish = bool(data.get('force_publish', False))
    force_note = (data.get('force_note') or '').strip()

    checklist = _build_publish_checklist(year, month)
    unresolved = checklist.get('unresolved', [])
    if unresolved and not force_publish:
        return jsonify({
            'error': '存在无法自动解决的冲突，默认阻止发布',
            'code': 'UNRESOLVED_CONFLICTS',
            'checklist': checklist
        }), 409

    if force_publish and unresolved and not force_note:
        return jsonify({
            'error': '强制发布必须填写备注',
            'code': 'FORCE_NOTE_REQUIRED'
        }), 400

    # find or create monthly plan
    plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
    if not plan:
        plan = MonthlyPlan(year=year, month=month)
        db.session.add(plan)

    plan.status = 'published'
    plan.published_at = datetime.now()

    # update month schedules
    start_date, end_date = _month_range(year, month)

    ClassSchedule.query.filter(
        ClassSchedule.scheduled_date >= start_date,
        ClassSchedule.scheduled_date < end_date,
        ClassSchedule.status.in_(['scheduled', 'planning', 'merged'])
    ).update({'status': 'confirmed'}, synchronize_session=False)

    if force_publish and unresolved:
        for item in unresolved:
            sid = item.get('schedule_id')
            schedule = ClassSchedule.query.get(sid)
            if schedule:
                schedule.notes = f"{schedule.notes or ''}\n[强制发布备注] {force_note}".strip()
                schedule.status = 'confirmed'

    db.session.commit()

    return jsonify({
        'message': f'{year}年{month}月计划已发布',
        'plan': plan.to_dict(),
        'forced': force_publish,
        'forced_conflict_count': len(unresolved) if force_publish else 0
    })


@schedule_bp.route('/unpublish', methods=['POST'])
def unpublish_month():
    """取消发布月度计划，回退为草稿"""
    data = request.get_json()
    year = data.get('year')
    month = data.get('month')

    if not year or not month:
        return jsonify({'error': '缺少 year/month'}), 400

    plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
    if not plan or plan.status != 'published':
        return jsonify({'error': '该月计划未发布，无需取消'}), 400

    plan.status = 'draft'
    plan.published_at = None
    plan.updated_at = datetime.now()

    # 将该月 confirmed 的课程回退为 scheduled
    start_date, end_date = _month_range(year, month)
    ClassSchedule.query.filter(
        ClassSchedule.scheduled_date >= start_date,
        ClassSchedule.scheduled_date < end_date,
        ClassSchedule.status == 'confirmed'
    ).update({'status': 'scheduled'}, synchronize_session=False)

    db.session.commit()

    return jsonify({
        'message': f'{year}年{month}月计划已取消发布，回退为草稿',
        'plan': plan.to_dict()
    })


# ==================== 草稿保存 ====================

@schedule_bp.route('/save-draft', methods=['POST'])
def save_draft():
    """保存当前月度课表为草稿"""
    data = request.get_json()
    year = data.get('year')
    month = data.get('month')
    
    if not year or not month:
        return jsonify({'error': '缺少 year/month'}), 400
    
    plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
    if not plan:
        plan = MonthlyPlan(year=year, month=month, status='draft')
        db.session.add(plan)
    
    plan.updated_at = datetime.now()
    # 如果已发布，不允许再存草稿（需要先解除发布）
    if plan.status == 'published':
        return jsonify({'error': '该月计划已发布，无法保存草稿'}), 400
    
    db.session.commit()
    
    return jsonify({
        'message': '草稿已保存',
        'plan': plan.to_dict()
    })


# ==================== 约束条件管理 ====================

def _get_or_create_plan(year, month):
    """获取或创建月度计划（辅助函数）"""
    plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
    if not plan:
        plan = MonthlyPlan(year=year, month=month, status='draft')
        db.session.add(plan)
        db.session.flush()  # 获取 ID
    return plan


@schedule_bp.route('/constraints', methods=['GET'])
def get_constraints():
    """获取指定月份的所有约束条件"""
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    
    if not year or not month:
        return jsonify({'error': '缺少 year/month'}), 400
    
    plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
    if not plan:
        return jsonify([])
    
    constraints = ScheduleConstraint.query.filter_by(
        monthly_plan_id=plan.id
    ).order_by(ScheduleConstraint.created_at).all()
    
    return jsonify([c.to_dict() for c in constraints])


@schedule_bp.route('/constraints', methods=['POST'])
def add_constraint():
    """添加约束条件"""
    data = request.get_json()
    year = data.get('year')
    month = data.get('month')
    description = data.get('description', '').strip()
    
    if not year or not month:
        return jsonify({'error': '缺少 year/month'}), 400
    if not description:
        return jsonify({'error': '约束描述不能为空'}), 400
    
    plan = _get_or_create_plan(year, month)
    
    constraint = ScheduleConstraint(
        monthly_plan_id=plan.id,
        constraint_type=data.get('constraint_type', 'custom'),
        description=description,
        parsed_data=_json.dumps(data.get('parsed_data')) if data.get('parsed_data') else None,
        is_active=True
    )
    db.session.add(constraint)
    db.session.commit()
    
    return jsonify(constraint.to_dict()), 201


@schedule_bp.route('/constraints/<int:constraint_id>', methods=['PUT'])
def update_constraint(constraint_id):
    """更新约束条件（主要用于切换启用/禁用）"""
    c = ScheduleConstraint.query.get_or_404(constraint_id)
    data = request.get_json()
    
    if 'is_active' in data:
        c.is_active = data['is_active']
    if 'description' in data:
        c.description = data['description']
    if 'constraint_type' in data:
        c.constraint_type = data['constraint_type']
    if 'parsed_data' in data:
        c.parsed_data = _json.dumps(data['parsed_data']) if data['parsed_data'] else None
    
    db.session.commit()
    return jsonify(c.to_dict())


@schedule_bp.route('/constraints/<int:constraint_id>', methods=['DELETE'])
def delete_constraint(constraint_id):
    """删除约束条件"""
    c = ScheduleConstraint.query.get_or_404(constraint_id)
    db.session.delete(c)
    db.session.commit()
    return jsonify({'message': '约束已删除'})




@schedule_bp.route('/check-holiday', methods=['GET'])
def check_holiday():
    """检查指定日期是否为节假日"""
    date_str = request.args.get('date')
    if not date_str:
        return jsonify({'error': '缺少date参数'}), 400
    
    result = is_holiday(date_str)
    return jsonify({
        'date': date_str,
        'is_holiday': result
    })


@schedule_bp.route('/<int:schedule_id>', methods=['DELETE'])
def delete_schedule(schedule_id):
    """删除单个课程安排"""
    schedule = ClassSchedule.query.get_or_404(schedule_id)
    
    # 状态保护：已完成的课程不允许删除
    if schedule.status == 'completed':
        return jsonify({'error': '已完成的课程不能删除'}), 400
    
    # 如果是合班主课程，恢复其他课程的原始状态
    import json as _json_mod
    merged_schedules = ClassSchedule.query.filter_by(merged_with=schedule_id).all()
    for ms in merged_schedules:
        if ms.merge_snapshot:
            try:
                snap = _json_mod.loads(ms.merge_snapshot)
                if snap.get('scheduled_date'):
                    ms.scheduled_date = date.fromisoformat(snap['scheduled_date'])
                if 'combo_id' in snap:
                    ms.combo_id = snap['combo_id']
                if 'combo_id_2' in snap:
                    ms.combo_id_2 = snap['combo_id_2']
                ms.status = snap.get('status', 'scheduled')
                ms.conflict_type = snap.get('conflict_type')
                ms.notes = snap.get('notes')
            except (ValueError, _json_mod.JSONDecodeError):
                pass
            ms.merge_snapshot = None
        ms.merged_with = None
    
    db.session.delete(schedule)
    db.session.commit()
    
    return jsonify({'message': '课程已删除', 'id': schedule_id})


@schedule_bp.route('/<int:schedule_id>', methods=['GET'])
def get_schedule_detail(schedule_id):
    """获取单个排课详情"""
    schedule = ClassSchedule.query.get_or_404(schedule_id)
    return jsonify(schedule.to_dict())


# ==================== Excel 导出 ====================

def _status_cn(status, merged_with=None, notes=None):
    """状态中文映射"""
    if merged_with or (notes and '合班主记录' in (notes or '')):
        return '合班'
    mapping = {
        'scheduled': '已排课',
        'completed': '已完成',
        'confirmed': '已确认',
        'conflict': '冲突',
        'planning': '预排',
        'cancelled': '已取消',
        'merged': '合班',
    }
    return mapping.get(status, status or '-')


def _display_topic_name(schedule):
    """获取拼接仪式前缀的课题显示名（开班仪式+X / 结课典礼+X）"""
    topic_name = schedule.topic.name if schedule.topic else '-'
    seq = schedule.topic.sequence if schedule.topic else None
    total = schedule.class_.project.topics.count() if schedule.class_ and schedule.class_.project else 0
    if seq and total > 0:
        if seq == 1:
            return f'开班仪式+{topic_name}'
        elif seq == total:
            return f'结课典礼+{topic_name}'
    return topic_name


def _apply_header_style(ws, header_row, col_count):
    """为表头行应用蓝底白字加粗样式"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_font = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=11)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        bottom=Side(style='thin', color='B0C4DE')
    )
    for col in range(1, col_count + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border


def _auto_width(ws, min_width=8, max_width=30):
    """根据内容自动调整列宽"""
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        try:
            first_cell = col[0]
            # 跳过 MergedCell（合并单元格占位对象没有 column_letter）
            if isinstance(first_cell, MergedCell):
                continue
            col_letter = first_cell.column_letter
            lengths = []
            for cell in col:
                if isinstance(cell, MergedCell):
                    continue
                val = str(cell.value) if cell.value is not None else ''
                # 中文字符按2倍宽度估算
                char_len = sum(2 if ord(c) > 127 else 1 for c in val)
                lengths.append(char_len)
            best = max(lengths) if lengths else min_width
            ws.column_dimensions[col_letter].width = max(min_width, min(best + 2, max_width))
        except (AttributeError, TypeError):
            continue


@schedule_bp.route('/export/<int:year>/<int:month>', methods=['GET'])
def export_month_excel(year, month):
    """导出月度课表 Excel（给讲师、班主任等工作人员看）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    from flask import send_file

    # 月份起止
    start_date = date(year, month, 1)
    end_date = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)

    schedules = ClassSchedule.query.filter(
        ClassSchedule.scheduled_date >= start_date,
        ClassSchedule.scheduled_date < end_date
    ).order_by(ClassSchedule.scheduled_date).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f'{year}年{month}月排课'

    # ---- 标题行 ----
    title_text = f'北清商学院 {year}年{month}月 排课计划'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(name='Microsoft YaHei', bold=True, size=16, color='1E293B')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    # 副标题
    plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
    status_text = '已发布' if plan and plan.status == 'published' else '草稿'
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
    sub_cell = ws.cell(row=2, column=1, value=f'状态: {status_text}  |  导出时间: {datetime.now().strftime("%Y-%m-%d %H:%M")}')
    sub_cell.font = Font(name='Microsoft YaHei', size=10, color='64748B')
    sub_cell.alignment = Alignment(horizontal='center')

    # ---- 表头 ----
    headers = ['周次', '日期(周六)', '日期(周日)', '班级', '课题', '周六讲师', '周六课程', '周日讲师', '周日课程', '班主任', '状态']
    header_row = 4
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=col_idx, value=h)
    _apply_header_style(ws, header_row, len(headers))

    # ---- 按周分组填充数据 ----
    # 构建周结构
    curr_week_start = start_date - timedelta(days=start_date.weekday())
    week_idx = 0
    row = header_row + 1

    conflict_font = Font(name='Microsoft YaHei', color='DC2626', bold=True)
    completed_fill = PatternFill(start_color='F0FDF4', end_color='F0FDF4', fill_type='solid')
    normal_font = Font(name='Microsoft YaHei', size=10)
    center_align = Alignment(horizontal='center', vertical='center')

    while curr_week_start < end_date:
        week_end = curr_week_start + timedelta(days=6)
        week_idx += 1

        sat = curr_week_start + timedelta(days=(5 - curr_week_start.weekday()) % 7)
        sun = sat + timedelta(days=1)

        week_schedules = [s for s in schedules if curr_week_start <= s.scheduled_date <= week_end]

        if not week_schedules:
            curr_week_start += timedelta(weeks=1)
            continue

        for s in week_schedules:
            day1_teacher = s.combo.teacher.name if s.combo and s.combo.teacher else '待定'
            day1_course = s.combo.course.name if s.combo and s.combo.course else '待定'
            day2_teacher = s.combo_2.teacher.name if s.combo_2 and s.combo_2.teacher else '待定'
            day2_course = s.combo_2.course.name if s.combo_2 and s.combo_2.course else '待定'
            homeroom = s.class_.homeroom.name if s.class_ and s.class_.homeroom else '未分配'
            status = _status_cn(s.status, s.merged_with, s.notes)

            # 使用实际排课日期而非计算的周六
            actual_sat = s.scheduled_date
            actual_sun = s.scheduled_date + timedelta(days=1)

            values = [
                f'第{week_idx}周',
                actual_sat.strftime('%m/%d'),
                actual_sun.strftime('%m/%d'),
                s.class_.name if s.class_ else '-',
                _display_topic_name(s),
                day1_teacher,
                day1_course,
                day2_teacher,
                day2_course,
                homeroom,
                status
            ]

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = normal_font
                cell.alignment = center_align

            # 冲突行红色高亮状态
            if s.status == 'conflict':
                ws.cell(row=row, column=11).font = conflict_font
            # 已完成行浅绿底
            if s.status in ('completed', 'confirmed'):
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col_idx).fill = completed_fill

            row += 1

        curr_week_start += timedelta(weeks=1)

    _auto_width(ws)

    # 写入流
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'排课计划_{year}年{month}月.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@schedule_bp.route('/export/class/<int:class_id>', methods=['GET'])
def export_class_excel(class_id):
    """导出单个班级课表 Excel（给学生看）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    from flask import send_file

    cls = Class.query.get_or_404(class_id)
    schedules = ClassSchedule.query.filter_by(class_id=class_id)\
        .order_by(ClassSchedule.scheduled_date).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f'{cls.name} 课程表'

    # ---- 标题 ----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    title_cell = ws.cell(row=1, column=1, value=f'{cls.name} 课程安排表')
    title_cell.font = Font(name='Microsoft YaHei', bold=True, size=16, color='1E293B')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    # 班级信息行
    homeroom = cls.homeroom.name if cls.homeroom else '未分配'
    project_name = cls.project.name if cls.project else '-'
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    info_cell = ws.cell(row=2, column=1, value=f'项目: {project_name}  |  班主任: {homeroom}  |  导出时间: {datetime.now().strftime("%Y-%m-%d %H:%M")}')
    info_cell.font = Font(name='Microsoft YaHei', size=10, color='64748B')
    info_cell.alignment = Alignment(horizontal='center')

    # ---- 表头 ----
    headers = ['序号', '日期(周六)', '日期(周日)', '课题', '周六讲师', '周六课程', '周日讲师', '周日课程']
    header_row = 4
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=col_idx, value=h)
    _apply_header_style(ws, header_row, len(headers))

    # ---- 数据行 ----
    normal_font = Font(name='Microsoft YaHei', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    completed_fill = PatternFill(start_color='F0FDF4', end_color='F0FDF4', fill_type='solid')

    row = header_row + 1
    for idx, s in enumerate(schedules, 1):
        sat = s.scheduled_date
        sun = sat + timedelta(days=1) if sat else None

        day1_teacher = s.combo.teacher.name if s.combo and s.combo.teacher else '待定'
        day1_course = s.combo.course.name if s.combo and s.combo.course else '待定'
        day2_teacher = s.combo_2.teacher.name if s.combo_2 and s.combo_2.teacher else '待定'
        day2_course = s.combo_2.course.name if s.combo_2 and s.combo_2.course else '待定'

        values = [
            idx,
            sat.strftime('%Y-%m-%d') if sat else '-',
            sun.strftime('%Y-%m-%d') if sun else '-',
            _display_topic_name(s),
            day1_teacher,
            day1_course,
            day2_teacher,
            day2_course,
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = normal_font
            cell.alignment = center_align

        # 已完成行浅绿底
        if s.status in ('completed', 'confirmed'):
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row, column=col_idx).fill = completed_fill

        row += 1

    _auto_width(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'课程表_{cls.name}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ==================== 智能排课算法核心函数 ====================
from config import Config as _cfg

def _get_last_class_date(class_id, before_date=None):
    """获取班级最后一次上课/排课的日期（跨所有月份，但只看 before_date 之前的记录）"""
    query = ClassSchedule.query.filter(
        ClassSchedule.class_id == class_id,
        ClassSchedule.status.in_(['completed', 'confirmed', 'scheduled', 'planning', 'merged'])
    )
    if before_date:
        query = query.filter(ClassSchedule.scheduled_date < before_date)
    last = query.order_by(ClassSchedule.scheduled_date.desc()).first()
    return last.scheduled_date if last else None


def _calculate_urgency(class_id, reference_date):
    """
    计算班级排课紧迫度。
    返回: (urgency_score, last_date)
    - urgency > 1.0 → 超过目标间隔，越大越急
    - urgency = 999  → 新班级从未排过课
    """
    last_date = _get_last_class_date(class_id, before_date=reference_date)
    if last_date is None:
        return 999.0, None
    days_since = (reference_date - last_date).days
    if days_since <= 0:
        return 0.01, last_date
    urgency = days_since / _cfg.TARGET_INTERVAL_DAYS
    return urgency, last_date


def _is_blocked(check_date, constraints):
    """检查日期是否在封锁日期列表中"""
    blocked = constraints.get('blocked_dates', [])
    for b in blocked:
        b_date = b.get('date') if isinstance(b, dict) else b
        if b_date == check_date.isoformat():
            return True
    return False


def _check_homeroom_unavailable(cls, sat, constraints):
    """检查班主任是否在该周有请假/不可用"""
    if not cls.homeroom_id or not cls.homeroom:
        return None
    homeroom_unavailable = constraints.get('homeroom_unavailable', [])
    hrm_name = cls.homeroom.name
    week_monday = sat - timedelta(days=5)
    week_sunday = week_monday + timedelta(days=6)

    for u in homeroom_unavailable:
        if u.get('homeroom_name') != hrm_name:
            continue
        dates_raw = u.get('dates', [])
        parsed = []
        for item in dates_raw:
            if isinstance(item, str):
                try:
                    parsed.append(date.fromisoformat(item))
                except Exception:
                    if '~' in item:
                        parts = item.split('~')
                        try:
                            d1 = date.fromisoformat(parts[0])
                            d2 = date.fromisoformat(parts[1])
                            d = d1
                            while d <= d2:
                                parsed.append(d)
                                d = d + timedelta(days=1)
                        except Exception:
                            pass
            elif isinstance(item, dict):
                f = item.get('from') or item.get('start')
                t = item.get('to') or item.get('end')
                try:
                    d1 = date.fromisoformat(f)
                    d2 = date.fromisoformat(t)
                    d = d1
                    while d <= d2:
                        parsed.append(d)
                        d = d + timedelta(days=1)
                except Exception:
                    pass
        for pd in parsed:
            if week_monday <= pd <= week_sunday:
                return f'班主任 {hrm_name} 在该周请假/不可用'
    return None


def _score_candidate(cls, sat, last_date, combo1, combo2, assigned_map, constraints):
    """
    评估某个候选周六对某班级的综合得分。
    
    参数:
        cls: Class 对象
        sat: 候选周六日期
        last_date: 该班级上次上课日期（可为 None）
        combo1: Day1(周六) 的 TeacherCourseCombo
        combo2: Day2(周日) 的 TeacherCourseCombo
        assigned_map: 本轮已分配记录 {class_id: {date, combo1_teacher_id, combo2_teacher_id}}
        constraints: 前端传入的约束条件 dict
    
    返回: (score, is_hard_conflict, conflict_reasons, merge_suggestions)
        score: 0.0 ~ 1.0
        is_hard_conflict: bool
        conflict_reasons: [str]
        merge_suggestions: [{target_class_id, target_class_name, topic_name}]  合班建议
    """
    conflict_reasons = []
    merge_suggestions = []
    sun = sat + timedelta(days=1)

    # --- H2: 节假日（双重保险，候选池理论上已过滤）---
    if is_holiday(sat) or is_holiday(sun):
        return (0.0, True, ['节假日冲突'], [])

    # --- H6: 封锁日期 ---
    blocked = constraints.get('blocked_dates', [])
    for b in blocked:
        b_date = b.get('date') if isinstance(b, dict) else b
        if b_date == sat.isoformat() or b_date == sun.isoformat():
            reason = b.get('reason', '人工约束') if isinstance(b, dict) else '人工约束'
            return (0.0, True, [f'封锁日期({reason})'], [])

    # --- S1: 间隔硬下限（< MIN_INTERVAL_DAYS 直接拒绝）---
    if last_date:
        interval = (sat - last_date).days
        if interval < _cfg.MIN_INTERVAL_DAYS:
            return (0.0, True, [f'间隔过短({interval}天 < {_cfg.MIN_INTERVAL_DAYS}天)'], [])
    
    # --- H3: 班主任不能同日双班 ---
    homeroom_conflict = False
    if cls.homeroom_id:
        # 查数据库已有记录
        existing = ClassSchedule.query.join(Class).filter(
            Class.homeroom_id == cls.homeroom_id,
            ClassSchedule.scheduled_date == sat,
            ClassSchedule.class_id != cls.id
        ).first()
        if existing:
            homeroom_conflict = True
            conflict_reasons.append(f'班主任撞课({existing.class_.name})')

        # 查本轮 assigned_map
        for a_cls_id, a_info in assigned_map.items():
            if a_info['date'] == sat and a_cls_id != cls.id:
                a_cls = Class.query.get(a_cls_id)
                if a_cls and a_cls.homeroom_id == cls.homeroom_id:
                    homeroom_conflict = True
                    conflict_reasons.append(f'班主任撞课({a_cls.name})[本轮]')

    # --- H7: 班主任请假 ---
    if cls.homeroom_id and not homeroom_conflict:
        hrm_reason = _check_homeroom_unavailable(cls, sat, constraints)
        if hrm_reason:
            homeroom_conflict = True
            conflict_reasons.append(hrm_reason)

    # --- H4: 讲师 Day1(周六) 不能撞课 ---
    teacher_conflict = False
    if combo1 and combo1.teacher:
        t1_id = combo1.teacher_id
        # 查数据库 — combo_id 维度
        t1_db = ClassSchedule.query.join(
            TeacherCourseCombo, ClassSchedule.combo_id == TeacherCourseCombo.id
        ).filter(
            TeacherCourseCombo.teacher_id == t1_id,
            ClassSchedule.scheduled_date == sat,
            ClassSchedule.class_id != cls.id
        ).first()
        if t1_db:
            teacher_conflict = True
            conflict_reasons.append(f'周六讲师 {combo1.teacher.name} 撞课({t1_db.class_.name})')
        # 查本轮 assigned_map — combo1_teacher_id 维度
        for a_cls_id, a_info in assigned_map.items():
            if a_info['date'] == sat and a_cls_id != cls.id:
                if a_info.get('combo1_teacher_id') == t1_id:
                    a_cls = Class.query.get(a_cls_id)
                    a_name = a_cls.name if a_cls else '?'
                    teacher_conflict = True
                    conflict_reasons.append(f'周六讲师 {combo1.teacher.name} 撞课({a_name})[本轮]')

    # --- H5: 讲师 Day2(周日) 不能撞课 ---
    if combo2 and combo2.teacher:
        t2_id = combo2.teacher_id
        # 查数据库 — combo_id_2 维度（同一 scheduled_date 的其他记录）
        schedules_on_sat = ClassSchedule.query.filter(
            ClassSchedule.scheduled_date == sat,
            ClassSchedule.class_id != cls.id
        ).all()
        for existing_s in schedules_on_sat:
            if existing_s.combo_id_2:
                c2 = TeacherCourseCombo.query.get(existing_s.combo_id_2)
                if c2 and c2.teacher_id == t2_id:
                    teacher_conflict = True
                    conflict_reasons.append(f'周日讲师 {combo2.teacher.name} 撞课({existing_s.class_.name})')
                    break
        # 查本轮 assigned_map — combo2_teacher_id 维度
        for a_cls_id, a_info in assigned_map.items():
            if a_info['date'] == sat and a_cls_id != cls.id:
                if a_info.get('combo2_teacher_id') == t2_id:
                    a_cls = Class.query.get(a_cls_id)
                    a_name = a_cls.name if a_cls else '?'
                    teacher_conflict = True
                    conflict_reasons.append(f'周日讲师 {combo2.teacher.name} 撞课({a_name})[本轮]')

    # --- H8: 讲师请假 ---
    unavailable = constraints.get('teacher_unavailable', [])
    t1_name = combo1.teacher.name if combo1 and combo1.teacher else None
    t2_name = combo2.teacher.name if combo2 and combo2.teacher else None
    sat_str = sat.isoformat()
    sun_str = sun.isoformat()
    if t1_name:
        for u in unavailable:
            if u.get('teacher_name') == t1_name and sat_str in u.get('dates', []):
                teacher_conflict = True
                conflict_reasons.append(f'周六讲师 {t1_name} 请假')
                break
    if t2_name:
        for u in unavailable:
            if u.get('teacher_name') == t2_name and sun_str in u.get('dates', []):
                teacher_conflict = True
                conflict_reasons.append(f'周日讲师 {t2_name} 请假')
                break

    # --- S4: 教室容量 ---
    db_count = ClassSchedule.query.filter(
        ClassSchedule.scheduled_date == sat,
        ClassSchedule.status.in_(['scheduled', 'planning', 'confirmed', 'completed', 'conflict'])
    ).count()
    # 注意：不计 merged，因为合班共用教室不额外占位
    assigned_count = sum(1 for info in assigned_map.values() if info['date'] == sat)
    total_on_sat = db_count + assigned_count

    # --- 合班建议：同课题检测（不依赖教室满不满） ---
    if combo1:
        current_topic_id = combo1.topic_id
        # 查当天同课题的班级（DB中）
        same_topic_on_sat = ClassSchedule.query.filter(
            ClassSchedule.scheduled_date == sat,
            ClassSchedule.topic_id == current_topic_id,
            ClassSchedule.class_id != cls.id,
            ClassSchedule.status.in_(['scheduled', 'planning', 'merged'])
        ).all()
        for s in same_topic_on_sat:
            merge_suggestions.append({
                'target_schedule_id': s.id,
                'target_class_id': s.class_id,
                'target_class_name': s.class_.name if s.class_ else '未知',
                'topic_name': s.topic.name if s.topic else '未知',
            })
        # 查本轮 assigned_map 中的同课题
        for a_cls_id, a_info in assigned_map.items():
            if a_info['date'] == sat and a_cls_id != cls.id:
                if a_info.get('topic_id') == current_topic_id:
                    a_cls = Class.query.get(a_cls_id)
                    merge_suggestions.append({
                        'target_class_id': a_cls_id,
                        'target_class_name': a_cls.name if a_cls else '未知',
                        'topic_name': combo1.topic.name if combo1.topic else '未知',
                    })

    if total_on_sat >= _cfg.MAX_CLASSES_PER_SATURDAY:
        conflict_reasons.append(f'教室已满({total_on_sat}/{_cfg.MAX_CLASSES_PER_SATURDAY})')
        # 教室满时：如果有合班建议，不完全拒绝（可以通过合班释放教室）
        if not merge_suggestions:
            return (0.0, True, conflict_reasons, [])
        # 有合班建议时继续评分，但 conflict_score 设为 0

    # --- 评分计算 ---

    # S1: 间隔分
    if last_date:
        interval = (sat - last_date).days
        interval_score = max(0.0, 1.0 - abs(interval - _cfg.TARGET_INTERVAL_DAYS) / _cfg.TARGET_INTERVAL_DAYS)
    else:
        interval_score = 1.0  # 新班级

    # 冲突分
    has_any_conflict = homeroom_conflict or teacher_conflict or len(conflict_reasons) > 0
    conflict_score = 0.0 if has_any_conflict else 1.0

    # 均衡分
    balance_score = 1.0 / (1 + total_on_sat)

    # 月份匹配分（如果候选跨月）
    in_month_score = 1.0  # 默认都在目标月内

    total = (
        interval_score * _cfg.SCORE_INTERVAL_WEIGHT +
        conflict_score * _cfg.SCORE_CONFLICT_WEIGHT +
        balance_score * _cfg.SCORE_BALANCE_WEIGHT +
        in_month_score * _cfg.SCORE_IN_MONTH_WEIGHT
    )

    return (total, has_any_conflict, conflict_reasons, merge_suggestions)


def _generate_suggestions(assignment, issue_type, all_assignments):
    """为某个排课问题生成建议方案"""
    suggestions = []
    if issue_type == 'too_long':
        suggestions.append(f"检查上月是否有空余档期可提前安排")
        suggestions.append(f"如无法提前，建议与讲师沟通加课")
    elif issue_type == 'conflict':
        if '班主任' in str(assignment.get('conflicts', [])):
            suggestions.append('调整到班主任可到场的周末')
            suggestions.append('或临时更换班主任')
        if '讲师' in str(assignment.get('conflicts', [])):
            suggestions.append('为该班更换讲师')
            suggestions.append('将该班移至讲师有空的周六')
        if '教室' in str(assignment.get('conflicts', [])):
            suggestions.append('建议与同课题班级合班上课')
    elif issue_type == 'no_slot':
        suggestions.append('建议手动安排到下月初')
        suggestions.append('检查是否可以更换讲师腾出档期')
    return suggestions


def _build_quality_report(assignments):
    """构建排课质量报告"""
    intervals = [a['interval_days'] for a in assignments if a.get('interval_days') is not None]

    if intervals:
        interval_scores = [
            max(0, 1.0 - abs(d - _cfg.TARGET_INTERVAL_DAYS) / _cfg.TARGET_INTERVAL_DAYS)
            for d in intervals
        ]
        overall_score = int(sum(interval_scores) / len(interval_scores) * 100)
    else:
        overall_score = 100

    good = [a for a in assignments if a.get('interval_days') and _cfg.MIN_INTERVAL_DAYS <= a['interval_days'] <= _cfg.MAX_INTERVAL_DAYS]
    long_ = [a for a in assignments if a.get('interval_days') and a['interval_days'] > _cfg.MAX_INTERVAL_DAYS]
    short = [a for a in assignments if a.get('interval_days') and a['interval_days'] < _cfg.MIN_INTERVAL_DAYS]
    conflict_items = [a for a in assignments if a.get('conflicts')]
    skipped = [a for a in assignments if a.get('assigned_date') is None]

    issues = []
    for a in long_:
        issues.append({
            'class_name': a.get('class_name'),
            'type': 'interval_too_long',
            'severity': 'warning',
            'detail': f"距上次课({a.get('last_date', '?')})已过{a['interval_days']}天，超出建议范围({_cfg.MAX_INTERVAL_DAYS}天)",
            'suggestions': _generate_suggestions(a, 'too_long', assignments)
        })
    for a in conflict_items:
        issues.append({
            'class_name': a.get('class_name'),
            'type': 'conflict',
            'severity': 'error',
            'detail': '；'.join(a['conflicts']),
            'suggestions': _generate_suggestions(a, 'conflict', assignments)
        })
    for a in skipped:
        issues.append({
            'class_name': a.get('class_name'),
            'type': 'no_slot',
            'severity': 'critical',
            'detail': a.get('skip_reason', '当月无可用档期'),
            'suggestions': _generate_suggestions(a, 'no_slot', assignments)
        })

    return {
        'overall_score': overall_score,
        'summary': {
            'total': len(assignments),
            'scheduled': len(assignments) - len(skipped),
            'good_interval': len(good),
            'long_interval': len(long_),
            'short_interval': len(short),
            'conflicts': len(conflict_items),
            'skipped': len(skipped),
        },
        'assignments': assignments,
        'issues': issues,
    }



def _run_scheduling_algorithm(year, month, constraints, conflict_mode='smart', overrides=None, skip_class_ids=None):
    """
    核心排课算法（共享逻辑：preview 和 generate 都调此函数）。
    
    参数:
        year, month: 目标月份
        constraints: 约束条件 dict
        conflict_mode: 'smart'(评分驱动) / 'postpone'(顺延) / 'mark'(标记冲突)
        overrides: {class_id_str: new_date_str} 用户在预览中的手动调整
        skip_class_ids: set of class IDs to skip (e.g. merged target classes)

    返回: (assignments, quality_report)
        assignments: 排课分配详情列表
        quality_report: 质量报告 dict
    """
    if overrides is None:
        overrides = {}
    if skip_class_ids is None:
        skip_class_ids = set()

    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)

    # Step 1: 构建候选日期池（含月前1周和月后1周的弹性空间）
    pool_start = start_date - timedelta(days=7)
    pool_end = end_date + timedelta(days=7)
    candidate_saturdays = []
    d = pool_start
    while d.weekday() != 5:
        d += timedelta(days=1)
    while d < pool_end:
        sun = d + timedelta(days=1)
        if not is_holiday(d) and not is_holiday(sun):
            if not _is_blocked(d, constraints) and not _is_blocked(sun, constraints):
                candidate_saturdays.append(d)
        d += timedelta(days=7)

    if not candidate_saturdays:
        return [], _build_quality_report([])

    # Step 2: 获取活跃班级 + 计算紧迫度 + 排序
    active_classes = Class.query.filter(Class.status.in_(['active', 'planning'])).all()
    ref_date = candidate_saturdays[0]

    class_infos = []
    for cls in active_classes:
        # 跳过被合班的 target 班级
        if cls.id in skip_class_ids:
            continue
        urgency, last_date = _calculate_urgency(cls.id, ref_date)
        class_infos.append({'cls': cls, 'urgency': urgency, 'last_date': last_date})

    # 按紧迫度降序 — 最急的先排，获得最好的档期
    class_infos.sort(key=lambda x: x['urgency'], reverse=True)

    # Step 3: 为每个班级评分选日期
    assigned_map = {}  # {class_id: {date, combo1_teacher_id, combo2_teacher_id, topic_id}}
    assignments = []

    for info in class_infos:
        cls = info['cls']
        last_date = info['last_date']

        # 确定下一个未上课题
        all_topics = cls.project.topics.all()  # relationship 已定义 order_by=Topic.sequence
        next_topic = None

        # 查已完成的最大 sequence（planning 不算，因为只是投射）
        last_completed = ClassSchedule.query.filter(
            ClassSchedule.class_id == cls.id,
            ClassSchedule.status.in_(['completed', 'confirmed', 'scheduled'])
        ).join(Topic).order_by(Topic.sequence.desc()).first()

        if not last_completed:
            if all_topics:
                next_topic = all_topics[0]
        else:
            current_seq = last_completed.topic.sequence
            for t in all_topics:
                if t.sequence > current_seq:
                    next_topic = t
                    break

        if not next_topic:
            continue  # 已结课

        # 选 combo
        all_combos = TeacherCourseCombo.query.filter_by(topic_id=next_topic.id)\
            .order_by(TeacherCourseCombo.priority.desc(), TeacherCourseCombo.id.desc()).all()

        combo1 = None
        combo2 = None
        if all_combos:
            combo1 = all_combos[0]
            combo2 = combo1  # 默认兜底
            for c in all_combos:
                if c.teacher_id != combo1.teacher_id:
                    combo2 = c
                    break

        # 序列化可用组合供前端编辑
        combos_list = [{'id': c.id, 'teacher_name': c.teacher.name if c.teacher else '?', 'course_name': c.course.name if c.course else '?'} for c in all_combos]

        # 检查用户是否在预览中手动指定了日期
        override_date = overrides.get(str(cls.id))
        if override_date:
            try:
                forced_sat = date.fromisoformat(override_date)
                interval_days = (forced_sat - last_date).days if last_date else None
                assignments.append({
                    'class_id': cls.id,
                    'class_name': cls.name,
                    'topic_id': next_topic.id,
                    'topic_name': next_topic.name,
                    'display_topic_name': f'开班仪式+{next_topic.name}' if next_topic.sequence == 1 else (f'结课典礼+{next_topic.name}' if next_topic.sequence == len(all_topics) else next_topic.name),
                    'combo_id': combo1.id if combo1 else None,
                    'combo_id_2': combo2.id if combo2 else None,
                    'combo1_teacher_name': combo1.teacher.name if combo1 and combo1.teacher else None,
                    'combo2_teacher_name': combo2.teacher.name if combo2 and combo2.teacher else None,
                    'combo1_course_name': combo1.course.name if combo1 and combo1.course else None,
                    'combo2_course_name': combo2.course.name if combo2 and combo2.course else None,
                    'assigned_date': forced_sat.isoformat(),
                    'last_date': last_date.isoformat() if last_date else None,
                    'interval_days': interval_days,
                    'score': 0.8,
                    'conflicts': [],
                    'merge_suggestions': [],
                    'homeroom_name': cls.homeroom.name if cls.homeroom else '未分配',
                    'urgency': info['urgency'],
                    'is_override': True,
                    'available_combos': combos_list,
                })
                assigned_map[cls.id] = {
                    'date': forced_sat,
                    'combo1_teacher_id': combo1.teacher_id if combo1 else None,
                    'combo2_teacher_id': combo2.teacher_id if combo2 else None,
                    'topic_id': next_topic.id,
                }
                continue
            except Exception:
                pass  # 无效的 override date，按正常流程走

        # 评分选日期
        if conflict_mode == 'smart':
            best_sat = None
            best_score = -1
            best_conflicts = []
            best_merge_suggestions = []

            for sat in candidate_saturdays:
                score, is_hard, reasons, merges = _score_candidate(
                    cls, sat, last_date, combo1, combo2,
                    assigned_map, constraints
                )
                if score > best_score:
                    best_score = score
                    best_sat = sat
                    best_conflicts = reasons if is_hard else []
                    best_merge_suggestions = merges  # 始终保留合班建议（不论是否有硬冲突）

        elif conflict_mode == 'mark':
            # 兼容旧逻辑：找第一个位置（即使有冲突也标记）
            best_sat = None
            best_score = 0
            best_conflicts = []
            best_merge_suggestions = []
            for sat in candidate_saturdays:
                score, is_hard, reasons, merges = _score_candidate(
                    cls, sat, last_date, combo1, combo2,
                    assigned_map, constraints
                )
                if not is_hard:
                    best_sat = sat
                    best_score = score
                    best_conflicts = []
                    best_merge_suggestions = []
                    break
                elif best_sat is None:
                    best_sat = sat
                    best_score = score
                    best_conflicts = reasons
                    best_merge_suggestions = merges

        else:  # postpone
            best_sat = None
            best_score = 0
            best_conflicts = []
            best_merge_suggestions = []
            for sat in candidate_saturdays:
                score, is_hard, reasons, merges = _score_candidate(
                    cls, sat, last_date, combo1, combo2,
                    assigned_map, constraints
                )
                if not is_hard:
                    best_sat = sat
                    best_score = score
                    best_conflicts = []
                    best_merge_suggestions = []
                    break
            # postpone 模式下如果全部有硬冲突，best_sat 保持 None

        if best_sat:
            interval_days = (best_sat - last_date).days if last_date else None
            assigned_map[cls.id] = {
                'date': best_sat,
                'combo1_teacher_id': combo1.teacher_id if combo1 else None,
                'combo2_teacher_id': combo2.teacher_id if combo2 else None,
                'topic_id': next_topic.id,
            }
            # 确定 status
            if best_conflicts:
                final_status = 'conflict'
            else:
                final_status = 'scheduled'

            assignments.append({
                'class_id': cls.id,
                'class_name': cls.name,
                'topic_id': next_topic.id,
                'topic_name': next_topic.name,
                'display_topic_name': f'开班仪式+{next_topic.name}' if next_topic.sequence == 1 else (f'结课典礼+{next_topic.name}' if next_topic.sequence == len(all_topics) else next_topic.name),
                'combo_id': combo1.id if combo1 else None,
                'combo_id_2': combo2.id if combo2 else None,
                'combo1_teacher_name': combo1.teacher.name if combo1 and combo1.teacher else None,
                'combo2_teacher_name': combo2.teacher.name if combo2 and combo2.teacher else None,
                'combo1_course_name': combo1.course.name if combo1 and combo1.course else None,
                'combo2_course_name': combo2.course.name if combo2 and combo2.course else None,
                'assigned_date': best_sat.isoformat(),
                'last_date': last_date.isoformat() if last_date else None,
                'interval_days': interval_days,
                'score': best_score,
                'conflicts': best_conflicts,
                'merge_suggestions': best_merge_suggestions,
                'homeroom_name': cls.homeroom.name if cls.homeroom else '未分配',
                'urgency': info['urgency'],
                'status': final_status,
                'available_combos': combos_list,
            })
        else:
            assignments.append({
                'class_id': cls.id,
                'class_name': cls.name,
                'topic_id': next_topic.id,
                'topic_name': next_topic.name,
                'display_topic_name': f'开班仪式+{next_topic.name}' if next_topic.sequence == 1 else (f'结课典礼+{next_topic.name}' if next_topic.sequence == len(all_topics) else next_topic.name),
                'assigned_date': None,
                'last_date': last_date.isoformat() if last_date else None,
                'interval_days': None,
                'score': 0,
                'conflicts': [],
                'merge_suggestions': [],
                'homeroom_name': cls.homeroom.name if cls.homeroom else '未分配',
                'urgency': info['urgency'],
                'skip_reason': '所有候选日期均不可用',
                'available_combos': combos_list,
            })

    quality_report = _build_quality_report(assignments)
    return assignments, quality_report


@schedule_bp.route('/generate-preview', methods=['POST'])
def generate_schedule_preview():
    """预览排课结果（不入库），返回质量报告和排课分配详情"""
    data = request.get_json()
    year = data.get('year')
    month = data.get('month')
    constraints = data.get('constraints', {})
    conflict_mode = data.get('conflict_mode', 'smart')
    merges = data.get('merges', [])

    if not year or not month:
        return jsonify({'error': 'Missing year/month'}), 400

    try:
        # 预览模式：先在 session 中删除该月草稿记录，运行算法，再 rollback
        start_d = date(year, month, 1)
        end_d = date(year, month + 1, 1) if month < 12 else date(year + 1, 1, 1)

        # 删除该月草稿记录（session级别，不 commit）
        ClassSchedule.query.filter(
            ClassSchedule.scheduled_date >= start_d,
            ClassSchedule.scheduled_date < end_d,
            ClassSchedule.status.in_(['scheduled', 'planning', 'conflict'])
        ).delete(synchronize_session='fetch')
        db.session.flush()  # 让算法查询能看到删除后的状态

        # 合班：收集被合并的 target 班级 ID，算法将跳过它们
        merged_target_ids = set()
        for m in merges:
            merged_target_ids.add(m.get('target_class_id'))

        assignments, quality_report = _run_scheduling_algorithm(
            year, month, constraints, conflict_mode,
            skip_class_ids=merged_target_ids
        )

        # 为合班的 source 班级标注合班信息
        for m in merges:
            for a in assignments:
                if a.get('class_id') == m.get('class_id'):
                    a['is_merged'] = True
                    a['merged_with_class_id'] = m.get('target_class_id')
                    a['merged_with_class_name'] = m.get('target_class_name', '')
                    # 使用用户选择的组合
                    if m.get('combo_id'):
                        a['combo_id'] = m['combo_id']
                    if m.get('combo_id_2'):
                        a['combo_id_2'] = m['combo_id_2']
                    break

        # 为被合并的 target 班级生成占位记录
        for m in merges:
            target_cls = Class.query.get(m.get('target_class_id'))
            if target_cls:
                source_a = next((a for a in assignments if a.get('class_id') == m.get('class_id')), None)
                assignments.append({
                    'class_id': m['target_class_id'],
                    'class_name': target_cls.name,
                    'topic_name': source_a.get('topic_name', '') if source_a else '',
                    'assigned_date': m.get('date') or (source_a.get('assigned_date') if source_a else None),
                    'is_merged_target': True,
                    'merged_into_class_id': m.get('class_id'),
                    'merged_into_class_name': source_a.get('class_name', '') if source_a else '',
                    'interval_days': None,
                    'conflicts': [],
                    'skip_reason': None,
                })

        # 重新构建质量报告（合班后少了冲突）
        quality_report = _build_quality_report(
            [a for a in assignments if not a.get('is_merged_target')]
        )

        # 回滚！预览不持久化任何更改
        db.session.rollback()

        # 获取当月已有排课记录（含 planning 初始计划）
        existing = ClassSchedule.query.filter(
            ClassSchedule.scheduled_date >= start_d,
            ClassSchedule.scheduled_date < end_d,
            ClassSchedule.status.in_(['completed', 'confirmed', 'scheduled', 'planning', 'merged'])
        ).all()
        existing_schedules = []
        for s in existing:
            existing_schedules.append({
                'id': s.id,
                'class_id': s.class_id,
                'class_name': s.class_.name if s.class_ else '?',
                'topic_name': s.topic.name if s.topic else '?',
                'display_topic_name': _display_topic_name(s),
                'scheduled_date': s.scheduled_date.isoformat(),
                'status': s.status,
                'homeroom_name': s.class_.homeroom.name if s.class_ and s.class_.homeroom else '未分配',
                'combo1_teacher_name': s.combo.teacher.name if s.combo and s.combo.teacher else '待定',
                'combo1_course_name': s.combo.course.name if s.combo and s.combo.course else '待定',
                'combo2_teacher_name': s.combo_2.teacher.name if s.combo_2 and s.combo_2.teacher else None,
                'combo2_course_name': s.combo_2.course.name if s.combo_2 and s.combo_2.course else None,
            })

        # 获取候选周六 + 不可用日期及原因
        all_saturdays = _get_all_saturdays_with_reasons(year, month, constraints)

        return jsonify({
            'success': True,
            'preview': assignments,
            'existing_schedules': existing_schedules,
            'quality_report': quality_report,
            'candidate_saturdays': [
                sat.isoformat() for sat in _get_candidate_saturdays(year, month, constraints)
            ],
            'all_saturdays': all_saturdays,
            'merges_applied': len(merges),
        })
    except Exception as e:
        db.session.rollback()  # 确保预览异常时也回滚
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def _get_candidate_saturdays(year, month, constraints):
    """获取候选周六列表（供前端预览调整时使用）"""
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)
    pool_start = start_date - timedelta(days=7)
    pool_end = end_date + timedelta(days=7)
    result = []
    d = pool_start
    while d.weekday() != 5:
        d += timedelta(days=1)
    while d < pool_end:
        sun = d + timedelta(days=1)
        if not is_holiday(d) and not is_holiday(sun):
            if not _is_blocked(d, constraints) and not _is_blocked(sun, constraints):
                result.append(d)
        d += timedelta(days=7)
    return result


def _get_all_saturdays_with_reasons(year, month, constraints):
    """获取月份范围内所有周六及其可用状态和原因，供前端显示"""
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)
    pool_start = start_date - timedelta(days=7)
    pool_end = end_date + timedelta(days=7)

    result = []
    d = pool_start
    while d.weekday() != 5:
        d += timedelta(days=1)
    while d < pool_end:
        sun = d + timedelta(days=1)
        item = {'date': d.isoformat(), 'sun_date': sun.isoformat(), 'available': True, 'reason': None}

        if is_holiday(d):
            item['available'] = False
            item['reason'] = f'{d.strftime("%m/%d")}(六) 节假日'
        elif is_holiday(sun):
            item['available'] = False
            item['reason'] = f'{sun.strftime("%m/%d")}(日) 节假日'
        elif _is_blocked(d, constraints):
            item['available'] = False
            item['reason'] = '约束条件排除'
        elif _is_blocked(sun, constraints):
            item['available'] = False
            item['reason'] = '约束条件排除(周日)'

        result.append(item)
        d += timedelta(days=7)
    return result


@schedule_bp.route('/generate', methods=['POST'])
def generate_schedule():
    """根据约束条件生成/重新生成月度排课（智能评分驱动）"""
    data = request.get_json()
    year = data.get('year')
    month = data.get('month')
    constraints = data.get('constraints', {})
    conflict_mode = data.get('conflict_mode', 'smart')
    overrides = data.get('overrides', {})
    combo_overrides = data.get('combo_overrides', {})  # {class_id_str: {combo1: id, combo2: id}}
    merges = data.get('merges', [])

    if not year or not month:
        return jsonify({'error': 'Missing year/month'}), 400

    try:
        # 1. 清除该月所有草稿状态的排课
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, month + 1, 1)

        # 先清除其他月份对即将删除记录的 merged_with 引用
        to_delete_ids = [s.id for s in ClassSchedule.query.filter(
            ClassSchedule.scheduled_date >= start_date,
            ClassSchedule.scheduled_date < end_date,
            ClassSchedule.status.in_(['scheduled', 'planning', 'conflict', 'merged'])
        ).all()]
        if to_delete_ids:
            ClassSchedule.query.filter(
                ClassSchedule.merged_with.in_(to_delete_ids)
            ).update({ClassSchedule.merged_with: None, ClassSchedule.merge_snapshot: None}, synchronize_session=False)

        ClassSchedule.query.filter(
            ClassSchedule.scheduled_date >= start_date,
            ClassSchedule.scheduled_date < end_date,
            ClassSchedule.status.in_(['scheduled', 'planning', 'conflict', 'merged'])
        ).delete()

        db.session.flush()

        # 2. 合班：收集被合并的 target 班级 ID
        merged_target_ids = set()
        for m in merges:
            merged_target_ids.add(m.get('target_class_id'))

        # 3. 运行核心排课算法
        assignments, quality_report = _run_scheduling_algorithm(
            year, month, constraints, conflict_mode, overrides,
            skip_class_ids=merged_target_ids
        )

        # 3. 将结果写入数据库
        generated_count = 0
        skipped_classes_info = []

        for a in assignments:
            if a.get('assigned_date') is None:
                skipped_classes_info.append({
                    'class_name': a.get('class_name', '未知'),
                    'reason': a.get('skip_reason', '无可用档期')
                })
                continue

            assigned_sat = date.fromisoformat(a['assigned_date'])

            # 应用用户的组合调整
            cls_id_str = str(a['class_id'])
            final_combo_id = a.get('combo_id')
            final_combo_id_2 = a.get('combo_id_2')
            if cls_id_str in combo_overrides:
                co = combo_overrides[cls_id_str]
                if 'combo1' in co:
                    final_combo_id = co['combo1']
                if 'combo2' in co:
                    final_combo_id_2 = co['combo2']

            # 确定冲突类型
            conflict_type_val = None
            final_status = a.get('status', 'scheduled')
            final_notes = 'AI智能排课'

            if a.get('conflicts'):
                final_status = 'conflict'
                final_notes = '；'.join(a['conflicts'])
                notes_text = final_notes
                if '班主任' in notes_text:
                    conflict_type_val = 'homeroom'
                elif '讲师' in notes_text:
                    conflict_type_val = 'teacher'
                elif '节假日' in notes_text:
                    conflict_type_val = 'holiday'

            new_schedule = ClassSchedule(
                class_id=a['class_id'],
                topic_id=a['topic_id'],
                combo_id=final_combo_id,
                combo_id_2=final_combo_id_2,
                scheduled_date=assigned_sat,
                week_number=0,
                status=final_status,
                conflict_type=conflict_type_val,
                notes=final_notes
            )
            db.session.add(new_schedule)
            generated_count += 1

        db.session.flush()  # 让 new_schedule 获得 ID

        # 4. 合班记录写入
        import json as _json
        for m in merges:
            source_class_id = m.get('class_id')
            target_class_id = m.get('target_class_id')
            merge_combo_id = m.get('combo_id')
            merge_combo_id_2 = m.get('combo_id_2')

            # 找到 source 班级的排课记录
            source_schedule = ClassSchedule.query.filter(
                ClassSchedule.class_id == source_class_id,
                ClassSchedule.scheduled_date >= start_date,
                ClassSchedule.scheduled_date < end_date,
                ClassSchedule.status.in_(['scheduled', 'conflict'])
            ).first()

            if source_schedule:
                # 为 target 班级创建合班记录
                target_cls = Class.query.get(target_class_id)
                target_topic_id = source_schedule.topic_id

                # 保存 source 的原始状态快照（拆分时恢复用）
                source_snapshot = _json.dumps({
                    'scheduled_date': source_schedule.scheduled_date.isoformat(),
                    'combo_id': source_schedule.combo_id,
                    'combo_id_2': source_schedule.combo_id_2,
                    'status': source_schedule.status,
                    'notes': source_schedule.notes or 'AI智能排课'
                })

                merged_schedule = ClassSchedule(
                    class_id=target_class_id,
                    topic_id=target_topic_id,
                    combo_id=merge_combo_id or source_schedule.combo_id,
                    combo_id_2=merge_combo_id_2 or source_schedule.combo_id_2,
                    scheduled_date=source_schedule.scheduled_date,
                    week_number=0,
                    status='merged',
                    merged_with=source_schedule.id,
                    merge_snapshot=_json.dumps({
                        'scheduled_date': source_schedule.scheduled_date.isoformat(),
                        'combo_id': merge_combo_id or source_schedule.combo_id,
                        'combo_id_2': merge_combo_id_2 or source_schedule.combo_id_2,
                        'status': 'scheduled',
                        'notes': None
                    }),
                    notes=f'合班: 与{source_schedule.class_.name if source_schedule.class_ else "?"}合班上课'
                )
                db.session.add(merged_schedule)
                generated_count += 1

                # 更新 source 的 notes 和保存快照
                source_schedule.merge_snapshot = source_snapshot
                source_schedule.notes = f'{source_schedule.notes or "AI智能排课"} ｜ 合班: 与{target_cls.name if target_cls else "?"}合班'

        db.session.commit()

        # 4. 自动创建/更新月度计划为草稿
        plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
        if not plan:
            plan = MonthlyPlan(year=year, month=month, status='draft')
            db.session.add(plan)
        plan.updated_at = datetime.now()
        db.session.commit()

        result_msg = f'已生成 {generated_count} 节课程安排'
        if skipped_classes_info:
            result_msg += f'。有 {len(skipped_classes_info)} 个班级未排课'

        return jsonify({
            'success': True,
            'message': result_msg,
            'skipped': skipped_classes_info,
            'plan': plan.to_dict(),
            'quality_report': quality_report,
        })

    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

