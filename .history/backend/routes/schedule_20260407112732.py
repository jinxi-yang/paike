"""
排课调度API - 含节假日检查、调整、合班功能
"""
from flask import Blueprint, jsonify, request, current_app
from datetime import date, datetime, timedelta
from models import db, ClassSchedule, Class, MonthlyPlan, TeacherCourseCombo, Topic, ScheduleConstraint, Homeroom, MergeConfig, City, Teacher
from routes.ai import call_ai_extract
import requests
import json as _json
import threading
import uuid
import copy

schedule_bp = Blueprint('schedule', __name__)

# ==================== 异步任务基础设施 ====================
_task_store = {}  # {task_id: {'status': 'running'|'done'|'error', 'progress': str, 'result': dict, 'error': str}}


@schedule_bp.route('/task-status/<task_id>', methods=['GET'])
def get_task_status(task_id):
    """查询异步排课任务的执行状态"""
    task = _task_store.get(task_id)
    if not task:
        return jsonify({'status': 'error', 'error': '任务不存在或已过期'}), 404
    return jsonify(task)


@schedule_bp.route('/task-cancel/<task_id>', methods=['POST'])
def cancel_task(task_id):
    """取消正在运行的排课任务"""
    task = _task_store.get(task_id)
    if not task:
        return jsonify({'error': '任务不存在'}), 404
    if task['status'] != 'running':
        return jsonify({'message': '任务已结束，无需取消'})
    task['cancelled'] = True
    task['progress'] = '正在取消...'
    return jsonify({'message': '取消请求已发送'})


def _month_range(year, month):
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)
    return start_date, end_date


def _guess_suggestion(reason):
    text = reason or ''
    if '节假日' in text:
        return '建议顺延到下一周周六/周日'
    if '请假' in text or '不可用' in text:
        return '建议改用同课题备选讲师，或顺延一周'
    if '班主任' in text:
        return '建议调整到班主任可到场的下一周末'
    if '撞课' in text:
        return '建议同周内错开班级，或调整到下一周'
    if '日期被排除' in text:
        return '建议改到最近可用周末'
    return '建议人工调整时间或师资后再发布'


def _build_publish_checklist(year, month):
    start_date, end_date = _month_range(year, month)
    month_schedules = ClassSchedule.query.filter(
        ClassSchedule.scheduled_date >= start_date,
        ClassSchedule.scheduled_date < end_date
    ).all()

    unresolved = []
    pending = []
    resolved = []

    for s in month_schedules:
        if s.conflict_type:
            reason = s.notes or '冲突原因未记录'
            same_day_count = ClassSchedule.query.filter(
                ClassSchedule.scheduled_date == s.scheduled_date
            ).count()
            unresolved.append({
                'schedule_id': s.id,
                'class_name': s.class_.name if s.class_ else None,
                'topic_name': s.topic.name if s.topic else None,
                'date': s.scheduled_date.isoformat() if s.scheduled_date else None,
                'reason': reason,
                'suggestion': _guess_suggestion(reason),
                'impact_scope': f'影响日期 {s.scheduled_date.isoformat()}，同日排课 {max(same_day_count - 1, 0)} 条'
            })
        elif s.status in ['scheduled']:
            # 只有真正需要关注的才放入 pending（缺讲师、缺日期等）
            needs_attention = False
            attention_reason = ''
            if not s.combo_id:
                needs_attention = True
                attention_reason = '未分配讲师（教-课组合）'
            elif not s.scheduled_date:
                needs_attention = True
                attention_reason = '未设定上课日期'

            if needs_attention:
                pending.append({
                    'schedule_id': s.id,
                    'class_name': s.class_.name if s.class_ else None,
                    'topic_name': s.topic.name if s.topic else None,
                    'date': s.scheduled_date.isoformat() if s.scheduled_date else None,
                    'reason': attention_reason,
                    'suggestion': '建议发布前补全信息',
                    'impact_scope': '仅影响当前班级'
                })
            else:
                resolved.append({
                    'schedule_id': s.id,
                    'class_name': s.class_.name if s.class_ else None,
                    'topic_name': s.topic.name if s.topic else None,
                    'date': s.scheduled_date.isoformat() if s.scheduled_date else None
                })
        else:
            resolved.append({
                'schedule_id': s.id,
                'class_name': s.class_.name if s.class_ else None,
                'topic_name': s.topic.name if s.topic else None,
                'date': s.scheduled_date.isoformat() if s.scheduled_date else None
            })

    # 主动扫描：检查同日同讲师/同班主任未标记的冲突
    # Bug 5 fix: 排除合班记录（merged_with 不为空的次记录），避免合班共用讲师/班主任被误报
    seen_conflict_ids = set(item['schedule_id'] for item in unresolved)
    from collections import defaultdict
    by_date = defaultdict(list)
    for s in month_schedules:
        if s.id not in seen_conflict_ids and s.status not in ('completed', 'cancelled') and not s.merged_with:
            by_date[s.scheduled_date].append(s)

    for dt, day_schedules in by_date.items():
        if len(day_schedules) < 2:
            continue
        # 检查周六讲师冲突
        teacher_day1 = defaultdict(list)
        for s in day_schedules:
            if s.combo and s.combo.teacher_id:
                teacher_day1[s.combo.teacher_id].append(s)
        for tid, slist in teacher_day1.items():
            if len(slist) >= 2:
                names = ', '.join(s.class_.name for s in slist if s.class_)
                teacher_name = slist[0].combo.teacher.name if slist[0].combo and slist[0].combo.teacher else '未知'
                pending.append({
                    'schedule_id': slist[0].id,
                    'class_name': names,
                    'topic_name': None,
                    'date': dt.isoformat() if dt else None,
                    'reason': f'周六讲师 {teacher_name} 同日为 {len(slist)} 个班级授课',
                    'suggestion': '建议错开日期或更换讲师',
                    'impact_scope': f'涉及 {len(slist)} 个班级'
                })
        # 检查周日讲师冲突
        teacher_day2 = defaultdict(list)
        for s in day_schedules:
            if s.combo_2 and s.combo_2.teacher_id:
                teacher_day2[s.combo_2.teacher_id].append(s)
        for tid, slist in teacher_day2.items():
            if len(slist) >= 2:
                names = ', '.join(s.class_.name for s in slist if s.class_)
                teacher_name = slist[0].combo_2.teacher.name if slist[0].combo_2 and slist[0].combo_2.teacher else '未知'
                pending.append({
                    'schedule_id': slist[0].id,
                    'class_name': names,
                    'topic_name': None,
                    'date': dt.isoformat() if dt else None,
                    'reason': f'周日讲师 {teacher_name} 同日为 {len(slist)} 个班级授课',
                    'suggestion': '建议错开日期或更换讲师',
                    'impact_scope': f'涉及 {len(slist)} 个班级'
                })
        # 检查班主任冲突 — 使用有效班主任（override优先）
        homeroom_map = defaultdict(list)
        for s in day_schedules:
            effective_hr = s.homeroom_override_id or (s.class_.homeroom_id if s.class_ else None)
            if effective_hr:
                homeroom_map[effective_hr].append(s)
        for hid, slist in homeroom_map.items():
            if len(slist) >= 2:
                names = ', '.join(s.class_.name for s in slist if s.class_)
                hr_obj = Homeroom.query.get(hid)
                hr_name = hr_obj.name if hr_obj else '未知'
                pending.append({
                    'schedule_id': slist[0].id,
                    'class_name': names,
                    'topic_name': None,
                    'date': dt.isoformat() if dt else None,
                    'reason': f'班主任 {hr_name} 同日负责 {len(slist)} 个班级',
                    'suggestion': '如为合班属正常情况，否则建议错开日期',
                    'impact_scope': f'涉及 {len(slist)} 个班级'
                })

    return {
        'resolved': resolved,
        'pending': pending,
        'unresolved': unresolved
    }

# ==================== 节假日相关 ====================

import os
import json

# 加载本地节假日缓存（按年份动态查找 holidays_{year}.json）
_local_holiday_data = {}
_holiday_files_loaded = set()
_backend_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def _load_holiday_file(year):
    """按年份加载对应的节假日缓存文件"""
    if year in _holiday_files_loaded:
        return
    _holiday_files_loaded.add(year)
    local_path = os.path.join(_backend_dir, f'holidays_{year}.json')
    try:
        with open(local_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            _local_holiday_data.update(data)
            print(f"Loaded {len(data)} holiday records from: {local_path}")
    except FileNotFoundError:
        print(f"Info: No holiday file for {year}: {local_path}")
    except Exception as e:
        print(f"Warning: Could not load holiday file {local_path}: {e}")

# 启动时预加载当前年份
from datetime import datetime as _dt
_load_holiday_file(_dt.now().year)

_holiday_cache = {}

# 记录已刷新过的月份，避免同一次排课session重复调API
_refreshed_months = set()

def _try_refresh_holidays_for_month(year, month):
    """排课前一次性尝试从API批量更新该月的节假日数据到缓存。
    失败不影响排课，静默回退到本地JSON。每月只刷新一次。"""
    month_key = f'{year}-{month:02d}'
    if month_key in _refreshed_months:
        return  # 已刷过
    _refreshed_months.add(month_key)
    
    # 收集该月所有周六周日
    start = date(year, month, 1)
    end = date(year, month + 1, 1) if month < 12 else date(year + 1, 1, 1)
    end_ext = end + timedelta(days=7)  # 延伸一周
    dates_to_check = []
    d = start
    while d < end_ext:
        if d.weekday() in (5, 6):  # 周六日
            date_str = d.isoformat()
            if date_str not in _holiday_cache and date_str not in _local_holiday_data:
                mm_dd = date_str[5:]
                if mm_dd not in _local_holiday_data:
                    dates_to_check.append(d)
        d += timedelta(days=1)
    
    if not dates_to_check:
        return  # 所有日期都已有本地数据
    
    print(f'[holiday] 尝试从API更新 {month_key} 的 {len(dates_to_check)} 个日期...', flush=True)
    success = 0
    for check_d in dates_to_check:
        try:
            resp = requests.get(
                f"https://timor.tech/api/holiday/info/{check_d.isoformat()}",
                timeout=2
            )
            if resp.status_code == 200:
                data = resp.json()
                if data.get('code') == 0:
                    holiday_info = data.get('holiday')
                    is_hol = holiday_info.get('holiday', False) if holiday_info else False
                    _holiday_cache[check_d.isoformat()] = is_hol
                    success += 1
            else:
                # API不可用(403等)，停止后续尝试
                print(f'[holiday] API返回 {resp.status_code}，使用本地数据', flush=True)
                return
        except Exception:
            print(f'[holiday] API超时/异常，使用本地数据', flush=True)
            return
    print(f'[holiday] 成功更新 {success}/{len(dates_to_check)} 个日期', flush=True)

def is_holiday(check_date):
    """
    检查某日期是否为节假日
    优先使用本地 holidays_{year}.json（按年份自动加载）
    其次使用 timor.tech API
    最后兜底逻辑：默认为工作日（排课日）
    """
    if isinstance(check_date, str):
        check_date = date.fromisoformat(check_date)
    
    # 动态加载该年份的节假日缓存
    _load_holiday_file(check_date.year)
    
    date_str = check_date.isoformat()
    mm_dd = date_str[5:] # MM-DD
    
    # 1. 检查内存缓存
    if date_str in _holiday_cache:
        return _holiday_cache[date_str]
    
    # 2. 检查本地文件缓存 (优先)
    # 尝试 YYYY-MM-DD
    if date_str in _local_holiday_data:
        record = _local_holiday_data[date_str]
        # treat official holidays and any record that represents a make-up/adjusted workday as restricted
        is_makeup = ('after' in record) or ('补班' in (record.get('name') or '')) or record.get('workday', False)
        is_hol = bool(record.get('holiday', False) or is_makeup)
        _holiday_cache[date_str] = is_hol
        return is_hol

    # 尝试 MM-DD (holidays_2026.json 格式)
    if mm_dd in _local_holiday_data:
        record = _local_holiday_data[mm_dd]
        is_makeup = ('after' in record) or ('补班' in (record.get('name') or '')) or record.get('workday', False)
        is_hol = bool(record.get('holiday', False) or is_makeup)
        # 注意：这里我们存入缓存的是完整日期
        _holiday_cache[date_str] = is_hol
        return is_hol

    # 3. 不在本地节假日数据中 = 正常日（可排课）
    # holidays_2026.json 已包含所有法定节假日和调休，无需外部API
    _holiday_cache[date_str] = False
    return False

def find_next_available_saturday(start_date):
    """找到下一个周六"""
    if isinstance(start_date, str):
        start_date = date.fromisoformat(start_date)
    
    # 找到下一个周六 (weekday 5 = Saturday)
    days_until_saturday = (5 - start_date.weekday()) % 7
    if days_until_saturday == 0 and start_date.weekday() != 5:
        days_until_saturday = 7
    
    return start_date + timedelta(days=days_until_saturday)


# ==================== 月度排课查询 ====================

@schedule_bp.route('/month/<int:year>/<int:month>', methods=['GET'])
def get_month_schedule(year, month):
    """获取指定月份的所有排课"""
    # 自动同步状态（纠正错误的completed标记）
    from .classes import sync_class_statuses
    sync_class_statuses()
    
    # 计算月份的起止日期
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)
    
    schedules = ClassSchedule.query.filter(
        ClassSchedule.scheduled_date >= start_date,
        ClassSchedule.scheduled_date < end_date
    ).order_by(ClassSchedule.scheduled_date).all()
    
    # 生成该月所有周的结构
    weeks_data = []
    
    # 找到该月第一天所在的周一
    # weekday(): 0=Mon, 6=Sun
    curr_week_start = start_date - timedelta(days=start_date.weekday())
    
    while curr_week_start < end_date:
        week_end = curr_week_start + timedelta(days=6)
        week_key = curr_week_start.isoformat()
        
        # 查找该周的课程，并附加间隔信息
        week_schedules = []
        for s in schedules:
            if not (curr_week_start <= s.scheduled_date <= week_end):
                continue
            s_dict = s.to_dict()
            # 查询该班级上一次课的日期（在当前课之前的最晚一次）
            prev = ClassSchedule.query.filter(
                ClassSchedule.class_id == s.class_id,
                ClassSchedule.scheduled_date < s.scheduled_date,
                ClassSchedule.status.in_(['completed', 'scheduled'])
            ).order_by(ClassSchedule.scheduled_date.desc()).first()
            if prev:
                s_dict['last_date'] = prev.scheduled_date.isoformat()
                s_dict['interval_days'] = (s.scheduled_date - prev.scheduled_date).days
            else:
                s_dict['last_date'] = None
                s_dict['interval_days'] = None
            week_schedules.append(s_dict)
        
        weeks_data.append({
            'week_start': week_key,
            'week_end': week_end.isoformat(),
            'schedules': week_schedules
        })
        
        curr_week_start += timedelta(weeks=1)
    
    # week_start > end_date 时停止，但可能最后一周跨月，只要 week_start < end_date 就会包含
    
    # 获取月度计划状态
    plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
    
    return jsonify({
        'year': year,
        'month': month,
        'plan_status': plan.status if plan else 'draft',
        'published_at': plan.published_at.isoformat() if plan and plan.published_at else None,
        'last_saved_at': plan.updated_at.isoformat() if plan and plan.updated_at else None,
        'weeks': weeks_data
    })


# ==================== 冲突重检公共函数 ====================

def _recheck_conflicts_for_dates(dates):
    """对指定日期集合内的所有活跃排课记录重新检测冲突，更新 conflict_type。
    
    应在所有修改排课数据的写操作后调用，确保 DB 中的 conflict_type 始终准确。
    调用前请确保已 flush 当前事务（让新数据可查询）。
    """
    if not dates:
        return
    for check_date in dates:
        if check_date is None:
            continue
        same_day = ClassSchedule.query.filter(
            ClassSchedule.scheduled_date == check_date,
            ClassSchedule.status.in_(['scheduled', 'completed']),
            ClassSchedule.merged_with.is_(None)
        ).all()

        for s in same_day:
            new_conflict = None
            conflict_notes = []
            # 检查讲师冲突 —— 周六(combo)
            if s.combo_id:
                c1 = TeacherCourseCombo.query.get(s.combo_id)
                if c1 and c1.teacher_id:
                    for other in same_day:
                        if other.id == s.id:
                            continue
                        if other.combo_id:
                            oc1 = TeacherCourseCombo.query.get(other.combo_id)
                            if oc1 and oc1.teacher_id == c1.teacher_id:
                                new_conflict = 'teacher'
                                t_name = c1.teacher.name if c1.teacher else '?'
                                o_name = other.class_.name if other.class_ else '?'
                                conflict_notes.append(f'讲师 {t_name} 撞课 ({o_name})')
                                break
            # 检查讲师冲突 —— 周日(combo_2)
            if not new_conflict and s.combo_id_2:
                c2 = TeacherCourseCombo.query.get(s.combo_id_2)
                if c2 and c2.teacher_id:
                    for other in same_day:
                        if other.id == s.id:
                            continue
                        if other.combo_id_2:
                            oc2 = TeacherCourseCombo.query.get(other.combo_id_2)
                            if oc2 and oc2.teacher_id == c2.teacher_id:
                                new_conflict = 'teacher'
                                t_name = c2.teacher.name if c2.teacher else '?'
                                o_name = other.class_.name if other.class_ else '?'
                                conflict_notes.append(f'讲师 {t_name} 撞课 ({o_name})')
                                break

            # 检查班主任冲突
            if not new_conflict:
                eff_hr = s.homeroom_override_id or (s.class_.homeroom_id if s.class_ else None)
                if eff_hr:
                    for other in same_day:
                        if other.id == s.id:
                            continue
                        other_hr = other.homeroom_override_id or (other.class_.homeroom_id if other.class_ else None)
                        if other_hr == eff_hr:
                            new_conflict = 'homeroom'
                            hr_obj = Homeroom.query.get(eff_hr)
                            hr_name = hr_obj.name if hr_obj else '?'
                            o_name = other.class_.name if other.class_ else '?'
                            conflict_notes.append(f'班主任 {hr_name} 撞课 ({o_name})')
                            break

            # 更新 conflict_type
            if s.conflict_type != new_conflict:
                s.conflict_type = new_conflict
                if new_conflict and conflict_notes:
                    # 设置冲突 notes（保留合班相关的 notes）
                    if s.notes and '合班' in s.notes:
                        pass  # 保留合班信息
                    else:
                        s.notes = '；'.join(conflict_notes)
                elif not new_conflict:
                    # 冲突解除，清除过时的冲突 notes
                    if s.notes and ('撞课' in s.notes or '手动' in s.notes or '冲突' in s.notes):
                        if '合班' not in s.notes:
                            s.notes = None


# ==================== 手动增加课次 ====================

@schedule_bp.route('/', methods=['POST'])
def create_schedule():
    """手动增加一个课次/活动"""
    data = request.get_json()
    class_id = data.get('class_id')
    topic_id = data.get('topic_id')
    scheduled_date_str = data.get('scheduled_date')

    if not class_id or not topic_id or not scheduled_date_str:
        return jsonify({'error': '班级、课题和日期为必填项'}), 400

    cls = Class.query.get(class_id)
    if not cls:
        return jsonify({'error': f'班级 ID {class_id} 不存在'}), 404

    topic = Topic.query.get(topic_id)
    if not topic:
        return jsonify({'error': f'课题 ID {topic_id} 不存在'}), 404

    # 课题唯一性校验：非"其他"课题在班级中只能排一次
    if not topic.is_other:
        existing = ClassSchedule.query.filter_by(
            class_id=class_id, topic_id=topic_id
        ).first()
        if existing:
            return jsonify({'error': f'课题「{topic.name}」已在本班课表中，普通课题只能排课一次。如需重复排课请使用【其他】课题。'}), 400

    try:
        scheduled_date = date.fromisoformat(scheduled_date_str)
    except ValueError:
        return jsonify({'error': f'日期格式无效: {scheduled_date_str}'}), 400

    combo_id = data.get('combo_id')
    combo_id_2 = data.get('combo_id_2')

    # 讲师/课程与课题一致性校验
    if combo_id:
        c1 = TeacherCourseCombo.query.get(combo_id)
        if c1 and c1.topic_id != topic_id:
            return jsonify({'error': '第一天分配的讲师/课程不属于该课题，请重新选择'}), 400
    if combo_id_2:
        c2 = TeacherCourseCombo.query.get(combo_id_2)
        if c2 and c2.topic_id != topic_id:
            return jsonify({'error': '第二天分配的讲师/课程不属于该课题，请重新选择'}), 400

    homeroom_override_id = data.get('homeroom_override_id')
    notes = data.get('notes', '').strip() or None

    # 构建备注（非周六时自动添加提示）
    if scheduled_date.weekday() != 5 and not notes:
        weekday_names = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
        notes = f'手动添加（{weekday_names[scheduled_date.weekday()]}）'

    # 地点校验：检查是否选择了已用过的非默认地点
    location_id_raw = data.get('location_id')
    location_id_val = int(location_id_raw) if location_id_raw else None
    if location_id_val and cls.city_id and location_id_val != cls.city_id:
        used_locations = _get_used_non_default_locations(class_id, cls.city_id)
        if location_id_val in used_locations:
            loc = City.query.get(location_id_val)
            return jsonify({'error': f'该班级已去过地点「{loc.name if loc else location_id_val}」，不允许再次选择'}), 400

    # 推迟受影响的后续课次（避开节假日并产生累加推延）
    postpone_weeks = data.get('postpone_weeks', 0)
    postponed_count = 0
    if postpone_weeks and postpone_weeks > 0:
        current_shift_days = postpone_weeks * 7
        affected_schedules = ClassSchedule.query.filter(
            ClassSchedule.class_id == class_id,
            ClassSchedule.scheduled_date >= scheduled_date
        ).order_by(ClassSchedule.scheduled_date).all()
        
        for sch in affected_schedules:
            # 加上当前的累积推迟天数
            new_date = sch.scheduled_date + timedelta(days=current_shift_days)
            original_new_date = new_date
            
            # 如果新日期落在节假日，继续往后找下一个可用周六
            while is_holiday(new_date):
                new_date += timedelta(weeks=1)
            
            # 计算因节假日额外多推迟的天数，并累加到全局偏移量中，确保后续的课跟着一起往后推，保持间隔
            extra_days = (new_date - original_new_date).days
            if extra_days > 0:
                current_shift_days += extra_days
                
            sch.scheduled_date = new_date
            postponed_count += 1

    new_schedule = ClassSchedule(
        class_id=class_id,
        topic_id=topic_id,
        combo_id=combo_id if combo_id else None,
        combo_id_2=combo_id_2 if combo_id_2 else None,
        scheduled_date=scheduled_date,
        week_number=0,
        status='scheduled',
        notes=notes,
        homeroom_override_id=homeroom_override_id if homeroom_override_id else None,
        location_id=location_id_val,
        has_opening=data.get('has_opening', False),
        has_team_building=data.get('has_team_building', False),
        has_closing=data.get('has_closing', False),
        day2_has_opening=data.get('day2_has_opening', False),
        day2_has_team_building=data.get('day2_has_team_building', False),
        day2_has_closing=data.get('day2_has_closing', False),
    )
    db.session.add(new_schedule)
    db.session.commit()

    # 创建后重新计算所有日期的顺序和本课次序号
    _resequence_topics_by_date(class_id)
    
    # 重新查询最新状态（可能 week_number 已变）
    new_schedule = ClassSchedule.query.get(new_schedule.id)

    result = new_schedule.to_dict()
    if postponed_count > 0:
        result['postponed_count'] = postponed_count
        result['postpone_message'] = f'已将 {postponed_count} 个后续课次推迟 {postpone_weeks} 周'

    return jsonify(result), 201


# ==================== 排课调整 ====================

@schedule_bp.route('/adjust', methods=['POST'])
def adjust_schedule():
    """调整单节课的日期或教-课组合"""
    data = request.get_json()
    schedule_id = data.get('schedule_id')
    force = data.get('force', False)
    
    schedule = ClassSchedule.query.get_or_404(schedule_id)
    
    # 状态保护：已取消的课程不允许调整 (已放开 completed 的限制以支持历史调整)
    if schedule.status == 'cancelled':
        return jsonify({'error': '该课程已取消，无法调整'}), 400
    
    # Bug 7 fix: 合班记录调整保护 — 提醒用户需先拆分
    if schedule.merged_with:
        return jsonify({'error': '该课程已合班，请先拆分后再调整'}), 400
    # 检查是否为合班主记录（被其他记录引用）
    has_secondary = ClassSchedule.query.filter_by(merged_with=schedule_id).count()
    if has_secondary > 0 and not force:
        return jsonify({
            'warning': f'该课程是合班主记录（含 {has_secondary} 个合班班级），调整将影响合班关系。建议先拆分后再调整，或强制保存。',
            'is_merge_master': True
        }), 200
    
    # 课题变更校验：非"其他"课题在班级中只能排一次
    if 'topic_id' in data:
        new_topic_id = data['topic_id']
        if new_topic_id and new_topic_id != schedule.topic_id:
            new_topic = Topic.query.get(new_topic_id)
            if new_topic and not new_topic.is_other:
                existing = ClassSchedule.query.filter(
                    ClassSchedule.class_id == schedule.class_id,
                    ClassSchedule.topic_id == new_topic_id,
                    ClassSchedule.id != schedule.id
                ).first()
                if existing:
                    return jsonify({'error': f'课题「{new_topic.name}」已在本班课表中，普通课题只能使用一次。如需重复请使用【其他】课题。'}), 400
            # 在非force模式下也设置 topic_id
            schedule.topic_id = new_topic_id

    # 讲师/课程与课题一致性校验（基础结构要求，不允许强制跳过）
    target_topic_id = schedule.topic_id
    if data.get('combo_id'):
        c1 = TeacherCourseCombo.query.get(data['combo_id'])
        if c1 and c1.topic_id != target_topic_id:
            return jsonify({'error': '第一天分配的讲师/课程不属于该课题，请重新选择'}), 400
    if data.get('combo_id_2'):
        c2 = TeacherCourseCombo.query.get(data['combo_id_2'])
        if c2 and c2.topic_id != target_topic_id:
            return jsonify({'error': '第二天分配的讲师/课程不属于该课题，请重新选择'}), 400

    # 强制模式：跳过冲突检查直接应用修改，但仍走重检测流程
    if force:
        conflict_notes = []
        if 'new_date' in data:
            schedule.scheduled_date = date.fromisoformat(data['new_date'])
        if 'combo_id' in data:
            old_combo_id = schedule.combo_id
            schedule.combo_id = data['combo_id']
            new_combo = TeacherCourseCombo.query.get(data['combo_id']) if data['combo_id'] else None
            if new_combo:
                conflict_notes.append(f'周六讲师手动改为 {new_combo.teacher.name}')
        if 'combo_id_2' in data:
            schedule.combo_id_2 = data['combo_id_2']
            new_combo2 = TeacherCourseCombo.query.get(data['combo_id_2']) if data['combo_id_2'] else None
            if new_combo2:
                conflict_notes.append(f'周日讲师手动改为 {new_combo2.teacher.name}')
        # 记录手动调整信息到notes，但不设conflict_type（由重检测决定）
        schedule.notes = '手动调整: ' + '; '.join(conflict_notes) if conflict_notes else schedule.notes
        # 不再 return，继续走下面的冲突重检测流程
    
    else:
        # 非强制模式：正常调整（含冲突检查）
        # 调整日期
        if 'new_date' in data:
            new_date = date.fromisoformat(data['new_date'])
            
            date_actually_changed = (schedule.scheduled_date != new_date)
            # 检查是否是周六 (现在是警告不是报错)
            if date_actually_changed and new_date.weekday() != 5:
                pass # 允许前端安排在非周末
            
            # 禁止移到已过去的日期
            if new_date < date.today():
                return jsonify({'error': f'不能将课程移到已过去的日期 ({new_date.isoformat()})'}), 400
            
            # 检查节假日
            if is_holiday(new_date):
                return jsonify({
                    'warning': '该日期是节假日',
                    'holiday': True,
                    'proceed': False
                }), 200
            
            # 检查班主任冲突 — 使用有效班主任（homeroom_override_id 优先于 Class.homeroom_id）
            class_obj = schedule.class_
            effective_hr_id = schedule.homeroom_override_id or (class_obj.homeroom_id if class_obj else None)
            if effective_hr_id:
                # 查找同日同班主任的其他课程
                same_day_schedules = ClassSchedule.query.filter(
                    ClassSchedule.scheduled_date == new_date,
                    ClassSchedule.id != schedule_id
                ).all()
                homeroom_conflicts = []
                for other_s in same_day_schedules:
                    other_effective_hr = other_s.homeroom_override_id or (other_s.class_.homeroom_id if other_s.class_ else None)
                    if other_effective_hr == effective_hr_id:
                        homeroom_conflicts.append(other_s)
                if homeroom_conflicts:
                    return jsonify({
                        'warning': '班主任在该日期已有其他课程',
                        'conflict_type': 'homeroom',
                        'conflicts': [c.to_dict() for c in homeroom_conflicts],
                        'proceed': True
                    }), 200
            
            # 检查讲师冲突（如果已分配教-课组合，考虑两天）
            # ★ 关键：如果用户同时提交了新combo，应该用新combo检查冲突，而不是旧的！
            #   否则用户想通过换讲师来解除冲突，但后端还在用旧讲师检查，永远报冲突。
            check_combo1_id = data.get('combo_id', schedule.combo_id)
            check_combo2_id = data.get('combo_id_2', schedule.combo_id_2)
            
            if check_combo1_id:
                check_combo1 = TeacherCourseCombo.query.get(check_combo1_id)
                if check_combo1 and check_combo1.teacher_id:
                    teacher_id = check_combo1.teacher_id
                    # 周六讲师只与其他记录的周六讲师(combo_id)比较
                    teacher_conflicts = ClassSchedule.query.join(
                        TeacherCourseCombo, ClassSchedule.combo_id == TeacherCourseCombo.id
                    ).filter(
                        TeacherCourseCombo.teacher_id == teacher_id,
                        ClassSchedule.scheduled_date == new_date,
                        ClassSchedule.id != schedule_id
                    ).all()

                    if teacher_conflicts:
                        return jsonify({
                            'warning': f'讲师 {check_combo1.teacher.name} 在该周六已有其他课程',
                            'conflict_type': 'teacher',
                            'conflicts': [c.to_dict() for c in teacher_conflicts],
                            'proceed': True
                        }), 200
            if check_combo2_id:
                check_combo2 = TeacherCourseCombo.query.get(check_combo2_id)
                if check_combo2 and check_combo2.teacher_id:
                    teacher2_id = check_combo2.teacher_id
                    # 周日讲师只与其他记录的周日讲师(combo_id_2)比较
                    teacher_conflicts2 = ClassSchedule.query.filter(
                        ClassSchedule.combo_id_2.isnot(None),
                        ClassSchedule.scheduled_date == new_date,
                        ClassSchedule.id != schedule_id
                    ).all()
                    teacher_conflicts2 = [s for s in teacher_conflicts2 if s.combo_2 and s.combo_2.teacher_id == teacher2_id]
                    if teacher_conflicts2:
                        return jsonify({
                            'warning': f'讲师 {check_combo2.teacher.name} 在该周日已有其他课程',
                            'conflict_type': 'teacher',
                            'conflicts': [c.to_dict() for c in teacher_conflicts2],
                            'proceed': True
                        }), 200
            
            # optional: 如果必要，还可检查新星期天是否是节假日
            sun_date = new_date + timedelta(days=1)
            if is_holiday(sun_date):
                return jsonify({
                    'warning': '新周日日期为节假日',
                    'holiday': True,
                    'proceed': False
                }), 200
            
            schedule.scheduled_date = new_date
        
        # 调整教-课组合（周六）
        if 'combo_id' in data:
            new_combo_id = data['combo_id']
            # 检查新周六讲师是否有冲突（只与其他记录的周六讲师比较）
            if new_combo_id:
                new_combo = TeacherCourseCombo.query.get(new_combo_id)
                if new_combo:
                    teacher_conflicts = ClassSchedule.query.join(
                        TeacherCourseCombo, ClassSchedule.combo_id == TeacherCourseCombo.id
                    ).filter(
                        TeacherCourseCombo.teacher_id == new_combo.teacher_id,
                        ClassSchedule.scheduled_date == schedule.scheduled_date,
                        ClassSchedule.id != schedule_id
                    ).all()
                    
                    if teacher_conflicts:
                        return jsonify({
                            'warning': f'讲师 {new_combo.teacher.name} 在该周六已有其他课程',
                            'conflict_type': 'teacher',
                            'conflicts': [c.to_dict() for c in teacher_conflicts],
                            'proceed': True
                        }), 200
            
            schedule.combo_id = new_combo_id
        # 调整教-课组合（周日）
        if 'combo_id_2' in data:
            new_combo2_id = data['combo_id_2']
            # 检查新周日讲师是否有冲突（只与其他记录的周日讲师比较）
            if new_combo2_id:
                new_combo2 = TeacherCourseCombo.query.get(new_combo2_id)
                if new_combo2:
                    teacher_conflicts2 = ClassSchedule.query.filter(
                        ClassSchedule.combo_id_2.isnot(None),
                        ClassSchedule.scheduled_date == schedule.scheduled_date,
                        ClassSchedule.id != schedule_id
                    ).all()
                    teacher_conflicts2 = [s for s in teacher_conflicts2 if s.combo_2 and s.combo_2.teacher_id == new_combo2.teacher_id]
                    if teacher_conflicts2:
                        return jsonify({
                            'warning': f'讲师 {new_combo2.teacher.name} 在该周日已有其他课程',
                            'conflict_type': 'teacher',
                            'conflicts': [c.to_dict() for c in teacher_conflicts2],
                            'proceed': True
                        }), 200
            schedule.combo_id_2 = new_combo2_id
        
        # 调整班主任
        if 'homeroom_override_id' in data:
            new_hr_id = data['homeroom_override_id']
            if new_hr_id:
                schedule.homeroom_override_id = int(new_hr_id)
            else:
                schedule.homeroom_override_id = None
        
        # 调整上课地点
        if 'location_id' in data:
            new_loc_id = data['location_id']
            if new_loc_id:
                new_loc_id = int(new_loc_id)
                # 检查是否为已用过的非默认地点
                class_obj_loc = schedule.class_
                default_city_id = class_obj_loc.city_id if class_obj_loc else None
                if default_city_id and new_loc_id != default_city_id:
                    used_locs = _get_used_non_default_locations(schedule.class_id, default_city_id, exclude_schedule_id=schedule.id)
                    if new_loc_id in used_locs:
                        loc = City.query.get(new_loc_id)
                        return jsonify({'error': f'该班级已去过地点「{loc.name if loc else new_loc_id}」，不允许再次选择'}), 400
                schedule.location_id = new_loc_id
            else:
                schedule.location_id = None

        # 更新备注
        if 'notes' in data:
            schedule.notes = data['notes']
    
    # ── 保存后：对同日所有课程重新检测冲突 ──
    db.session.flush()

    dates_to_check = {schedule.scheduled_date}
    if 'date' in data and data['date']:
        old_date_str = data.get('_old_date')
        if old_date_str:
            dates_to_check.add(date.fromisoformat(old_date_str))

    _recheck_conflicts_for_dates(dates_to_check)
    db.session.commit()
    
    # 日期变动后重排课题序号
    _resequence_topics_by_date(schedule.class_id)
    
    return jsonify(schedule.to_dict())


@schedule_bp.route('/move-week', methods=['POST'])
def move_to_week():
    """将课程移动到上一周或下一周（含冲突/节假日检测）"""
    data = request.get_json()
    schedule_id = data.get('schedule_id')
    direction = data.get('direction', 'next')  # 'next' or 'prev'
    force = data.get('force', False)  # 前端确认后强制移动
    
    schedule = ClassSchedule.query.get_or_404(schedule_id)
    
    # 状态保护（放开 completed）
    if schedule.status == 'cancelled':
        return jsonify({'error': '该课程已取消，无法移动'}), 400
    
    if direction == 'next':
        new_date = schedule.scheduled_date + timedelta(weeks=1)
    else:
        new_date = schedule.scheduled_date - timedelta(weeks=1)
    
    # 确保是周六
    new_date = find_next_available_saturday(new_date - timedelta(days=1))
    
    # 禁止移到已过去的日期
    if new_date < date.today():
        return jsonify({'error': f'不能将课程移到已过去的日期 ({new_date.isoformat()})'}), 400
    
    # 检查目标日期的冲突
    warnings = []
    
    # 检查节假日
    if is_holiday(new_date):
        warnings.append(f'{new_date.isoformat()} 为节假日')
    
    # 检查班主任冲突 — 使用有效班主任（homeroom_override_id 优先）
    effective_hr_id = schedule.homeroom_override_id or (schedule.class_.homeroom_id if schedule.class_ else None)
    if effective_hr_id:
        same_day_others = ClassSchedule.query.filter(
            ClassSchedule.scheduled_date == new_date,
            ClassSchedule.class_id != schedule.class_id
        ).all()
        for other_s in same_day_others:
            other_hr = other_s.homeroom_override_id or (other_s.class_.homeroom_id if other_s.class_ else None)
            if other_hr == effective_hr_id:
                warnings.append(f'班主任与 {other_s.class_.name if other_s.class_ else "?"} 撞课')
                break
    
    # 检查讲师冲突（周六）
    if schedule.combo:
        teacher_conflict = ClassSchedule.query.filter(
            ClassSchedule.scheduled_date == new_date,
            ClassSchedule.combo_id.isnot(None),
            ClassSchedule.id != schedule.id
        ).all()
        for tc in teacher_conflict:
            if tc.combo and tc.combo.teacher_id == schedule.combo.teacher_id:
                warnings.append(f'周六讲师 {schedule.combo.teacher.name} 与 {tc.class_.name} 撞课')
                break
    
    # 检查讲师冲突（周日）
    if schedule.combo_2:
        teacher_conflict_2 = ClassSchedule.query.filter(
            ClassSchedule.scheduled_date == new_date,
            ClassSchedule.combo_id_2.isnot(None),
            ClassSchedule.id != schedule.id
        ).all()
        for tc in teacher_conflict_2:
            if tc.combo_2 and tc.combo_2.teacher_id == schedule.combo_2.teacher_id:
                warnings.append(f'周日讲师 {schedule.combo_2.teacher.name} 与 {tc.class_.name} 撞课')
                break
    
    # 如果有警告且未强制，返回警告让前端确认
    if warnings and not force:
        return jsonify({
            'confirm_required': True,
            'new_date': new_date.isoformat(),
            'warnings': warnings,
            'message': '移动目标日期存在以下问题，是否仍要移动？'
        })
    
    old_date = schedule.scheduled_date
    schedule.scheduled_date = new_date
    # 移动后由公共函数统一重检（旧日期 + 新日期），不再手动标记
    db.session.flush()
    _recheck_conflicts_for_dates({old_date, new_date})
    db.session.commit()
    
    return jsonify(schedule.to_dict())


# ==================== 课题交换 ====================

@schedule_bp.route('/swap-topics', methods=['POST'])
def swap_topics():
    """交换同一班级内两个未完成课次的日期和地点"""
    data = request.get_json()
    id_a = data.get('schedule_id_a')
    id_b = data.get('schedule_id_b')

    if not id_a or not id_b or id_a == id_b:
        return jsonify({'error': '请选择两个不同的课次'}), 400

    sch_a = ClassSchedule.query.get(id_a)
    sch_b = ClassSchedule.query.get(id_b)
    if not sch_a or not sch_b:
        return jsonify({'error': '课次记录不存在'}), 404

    # 必须是同一班级
    if sch_a.class_id != sch_b.class_id:
        return jsonify({'error': '只能交换同一班级的课次'}), 400

    # 不能是已完成或已取消
    for s, label in [(sch_a, 'A'), (sch_b, 'B')]:
        if s.status in ('completed', 'cancelled'):
            return jsonify({'error': f'课次 {s.topic.name if s.topic else label} 状态为 {s.status}，无法交换'}), 400

    # 不能是已过去的日期
    today = date.today()
    for s in [sch_a, sch_b]:
        if s.scheduled_date and s.scheduled_date < today:
            return jsonify({'error': f'课次 {s.topic.name if s.topic else "?"} 的日期已过去，无法交换'}), 400

    # 合班记录不允许交换
    for s in [sch_a, sch_b]:
        if s.merged_with:
            return jsonify({'error': f'课次 {s.topic.name if s.topic else "?"} 已合班，请先拆分'}), 400

    # 执行交换：互换日期和地点
    sch_a.scheduled_date, sch_b.scheduled_date = sch_b.scheduled_date, sch_a.scheduled_date
    sch_a.location_id, sch_b.location_id = sch_b.location_id, sch_a.location_id

    db.session.flush()

    # 重新检测两个日期上的冲突
    dates_to_check = set()
    if sch_a.scheduled_date:
        dates_to_check.add(sch_a.scheduled_date)
    if sch_b.scheduled_date:
        dates_to_check.add(sch_b.scheduled_date)
    _recheck_conflicts_for_dates(dates_to_check)

    db.session.commit()

    # 重排课题序号
    _resequence_topics_by_date(sch_a.class_id)

    return jsonify({
        'message': '交换成功',
        'schedule_a': sch_a.to_dict(),
        'schedule_b': sch_b.to_dict()
    })


# ==================== 合班功能 ====================

@schedule_bp.route('/merge-info', methods=['POST'])
def merge_info():
    """获取合班可选项（日期、讲师组合、班主任）供前端弹窗展示"""
    data = request.get_json()
    schedule_ids = data.get('schedule_ids', [])
    
    schedules_unordered = ClassSchedule.query.filter(ClassSchedule.id.in_(schedule_ids)).all()
    if len(schedules_unordered) < 2:
        return jsonify({'error': '至少需要两条记录'}), 400
    
    # 按 schedule_ids 顺序排列，确保第一个 ID 对应主班
    id_to_sch = {s.id: s for s in schedules_unordered}
    schedules = [id_to_sch[sid] for sid in schedule_ids if sid in id_to_sch]
    if len(schedules) < 2:
        return jsonify({'error': '至少需要两条记录'}), 400
    
    # 收集可选日期
    dates = list(set(s.scheduled_date.isoformat() for s in schedules if s.scheduled_date))
    dates.sort()
    
    # 收集可选周六组合
    combos_day1 = []
    seen_combo1 = set()
    for s in schedules:
        if s.combo and s.combo_id not in seen_combo1:
            seen_combo1.add(s.combo_id)
            combos_day1.append({
                'combo_id': s.combo_id,
                'label': f'{s.combo.teacher.name} - {s.combo.course_name}' if s.combo.teacher and s.combo.course_name else f'组合#{s.combo_id}',
                'from_class': s.class_.name if s.class_ else ''
            })
    
    # 收集可选周日组合
    combos_day2 = []
    seen_combo2 = set()
    for s in schedules:
        if s.combo_2 and s.combo_id_2 not in seen_combo2:
            seen_combo2.add(s.combo_id_2)
            combos_day2.append({
                'combo_id': s.combo_id_2,
                'label': f'{s.combo_2.teacher.name} - {s.combo_2.course_name}' if s.combo_2.teacher and s.combo_2.course_name else f'组合#{s.combo_id_2}',
                'from_class': s.class_.name if s.class_ else ''
            })
    
    # 收集可选班主任 — 使用有效班主任（override优先）
    homerooms = []
    seen_hr = set()
    for s in schedules:
        effective_hr_id = s.homeroom_override_id or (s.class_.homeroom_id if s.class_ else None)
        if effective_hr_id and effective_hr_id not in seen_hr:
            seen_hr.add(effective_hr_id)
            hr_obj = Homeroom.query.get(effective_hr_id)
            homerooms.append({
                'homeroom_id': effective_hr_id,
                'name': hr_obj.name if hr_obj else '未知',
                'from_class': s.class_.name if s.class_ else ''
            })
    
    # 收集可选上课地点（合班班级默认地点的合集）
    locations = []
    seen_loc = set()
    for s in schedules:
        cls = s.class_
        if cls and cls.city_id and cls.city_id not in seen_loc:
            seen_loc.add(cls.city_id)
            city_obj = City.query.get(cls.city_id)
            
            # 检查是否有任何参与班级已经去过这个地点（且该班级原本地点不是这里）
            is_disabled = False
            reason = ''
            for part_s in schedules:
                part_cls = part_s.class_
                if part_cls and part_cls.city_id and part_cls.city_id != cls.city_id:
                    used_locs = _get_used_non_default_locations(part_s.class_id, part_cls.city_id, exclude_schedule_id=part_s.id)
                    if cls.city_id in used_locs:
                        is_disabled = True
                        reason = f'班级 {part_cls.name} 之前已去过该地点'
                        break

            locations.append({
                'location_id': cls.city_id,
                'name': city_obj.name if city_obj else '未知',
                'from_class': cls.name,
                'is_disabled': is_disabled,
                'reason': reason
            })
    
    return jsonify({
        'dates': dates,
        'combos_day1': combos_day1,
        'combos_day2': combos_day2,
        'homerooms': homerooms,
        'locations': locations,
        'schedules': [s.to_dict() for s in schedules]
    })


@schedule_bp.route('/merge', methods=['POST'])
def merge_classes():
    """合班操作 - 多个班级同一课题合并上课，统一讲师和日期"""
    import json
    data = request.get_json()
    schedule_ids = data.get('schedule_ids', [])
    merged_date = data.get('merged_date')
    merged_combo_id = data.get('merged_combo_id')
    merged_combo_id_2 = data.get('merged_combo_id_2')
    lead_homeroom = data.get('lead_homeroom_name', '')
    
    if len(schedule_ids) < 2:
        return jsonify({'error': '合班至少需要两个课程'}), 400
    
    schedules_unordered = ClassSchedule.query.filter(ClassSchedule.id.in_(schedule_ids)).all()
    
    if len(schedules_unordered) != len(schedule_ids):
        return jsonify({'error': '部分课程不存在'}), 404
    
    # 按 schedule_ids 顺序排列，确保第一个 ID 成为主记录
    id_to_sch = {s.id: s for s in schedules_unordered}
    schedules = [id_to_sch[sid] for sid in schedule_ids if sid in id_to_sch]
    
    # 验证是否为同一课题
    topic_ids = set(s.topic_id for s in schedules)
    if len(topic_ids) > 1:
        return jsonify({'error': '只能合并相同课题的课程'}), 400
    
    # 使用第一个作为主课表
    main_schedule = schedules[0]
    class_names = [s.class_.name for s in schedules if s.class_]
    
    # 统一日期
    if merged_date:
        unified_date = date.fromisoformat(merged_date)
    else:
        unified_date = main_schedule.scheduled_date
        
    # 验证合班地点是否已被任一班级使用过
    merged_location_id = data.get('merged_location_id')
    if merged_location_id:
        merged_loc_id_val = int(merged_location_id)
        for s in schedules:
            cls = s.class_
            default_city_id = cls.city_id if cls else None
            if default_city_id and merged_loc_id_val != default_city_id:
                # 排除当前要被合并的记录，因为如果原来选的就是这个地点，不算违规
                used_locs = _get_used_non_default_locations(s.class_id, default_city_id, exclude_schedule_id=s.id)
                if merged_loc_id_val in used_locs:
                    loc = City.query.get(merged_loc_id_val)
                    return jsonify({'error': f'班级「{cls.name}」已去过地点「{loc.name if loc else merged_loc_id_val}」，不允许再次选择此地点进行合班'}), 400
    
    # 合班前保存每条记录的快照（拆分时用于恢复）
    for s in schedules:
        snapshot = {
            'scheduled_date': s.scheduled_date.isoformat() if s.scheduled_date else None,
            'combo_id': s.combo_id,
            'combo_id_2': s.combo_id_2,
            'status': s.status,
            'conflict_type': s.conflict_type,
            'notes': s.notes,
            'homeroom_override_id': s.homeroom_override_id,  # 保存合班前的班主任设置
            'location_id': s.location_id  # 保存合班前的地点设置
        }
        s.merge_snapshot = json.dumps(snapshot, ensure_ascii=False)
    
    # 解析班主任名称为 ID
    lead_homeroom_id = None
    if lead_homeroom:
        hr = Homeroom.query.filter_by(name=lead_homeroom).first()
        if hr:
            lead_homeroom_id = hr.id
    
    # 更新所有记录的统一信息
    for s in schedules:
        s.scheduled_date = unified_date
        if merged_combo_id:
            s.combo_id = int(merged_combo_id)
        if merged_combo_id_2:
            s.combo_id_2 = int(merged_combo_id_2)
        # 合班班主任作为临时覆盖保存到每条记录
        if lead_homeroom_id:
            s.homeroom_override_id = lead_homeroom_id
        # 合班上课地点
        merged_location_id = data.get('merged_location_id')
        if merged_location_id:
            s.location_id = int(merged_location_id)
    
    # 设置合班关联
    for s in schedules[1:]:
        s.merged_with = main_schedule.id
        s.notes = f'合班至 {main_schedule.class_.name}' + (f'（带班: {lead_homeroom}）' if lead_homeroom else '')
    
    main_schedule.notes = f'合班主记录（含 {", ".join(class_names[1:])}）' + (f'（带班: {lead_homeroom}）' if lead_homeroom else '')
    
    # 合班后重检冲突（合班可能改变日期/班主任/讲师）
    db.session.flush()
    affected_dates = set(s.scheduled_date for s in schedules if s.scheduled_date)
    _recheck_conflicts_for_dates(affected_dates)

    # 自动持久化合班关系到 merge_config（重排时保留合班）
    try:
        mc_year = unified_date.year
        mc_month = unified_date.month
        plan = _get_or_create_plan(mc_year, mc_month)
        topic_id = main_schedule.topic_id
        primary_class_id = main_schedule.class_id
        for s in schedules[1:]:
            existing_mc = MergeConfig.query.filter_by(
                monthly_plan_id=plan.id,
                topic_id=topic_id,
                merged_class_id=s.class_id
            ).first()
            if not existing_mc:
                mc = MergeConfig(
                    monthly_plan_id=plan.id,
                    topic_id=topic_id,
                    primary_class_id=primary_class_id,
                    merged_class_id=s.class_id,
                    combo_id=int(merged_combo_id) if merged_combo_id else None,
                    combo_id_2=int(merged_combo_id_2) if merged_combo_id_2 else None
                )
                db.session.add(mc)
    except Exception:
        pass  # merge_config 写入失败不影响合班本身

    db.session.commit()
    
    return jsonify({
        'main_schedule': main_schedule.to_dict(),
        'merged_schedules': [s.to_dict() for s in schedules[1:]]
    })


@schedule_bp.route('/unmerge/<int:schedule_id>', methods=['POST'])
def unmerge_class(schedule_id):
    """取消合班 - 级联清除所有相关合班记录，并从快照恢复原始状态"""
    import json
    schedule = ClassSchedule.query.get_or_404(schedule_id)
    
    def _restore_from_snapshot(record):
        """从 merge_snapshot 恢复记录的原始状态"""
        if not record.merge_snapshot:
            return
        try:
            snap = json.loads(record.merge_snapshot)
            if snap.get('scheduled_date'):
                record.scheduled_date = date.fromisoformat(snap['scheduled_date'])
            if 'combo_id' in snap:
                record.combo_id = snap['combo_id']
            if 'combo_id_2' in snap:
                record.combo_id_2 = snap['combo_id_2']
            record.status = snap.get('status', 'scheduled')
            record.conflict_type = snap.get('conflict_type')
            record.notes = snap.get('notes')
            # 恢复班主任设置（合班前的状态）
            record.homeroom_override_id = snap.get('homeroom_override_id')
            # 恢复地点设置
            record.location_id = snap.get('location_id')
        except (json.JSONDecodeError, ValueError):
            pass
        record.merge_snapshot = None
        record.merged_with = None
    
    # 情况1: 如果是主记录（被其他记录引用），同时恢复所有次记录
    secondary_records = ClassSchedule.query.filter_by(merged_with=schedule_id).all()
    for sec in secondary_records:
        _restore_from_snapshot(sec)
    
    # 情况2: 如果是次记录（引用其他主记录），也检查主记录
    if schedule.merged_with:
        main_record = ClassSchedule.query.get(schedule.merged_with)
        if main_record:
            # 检查主记录下还有没有其他次记录
            remaining_records = ClassSchedule.query.filter(
                ClassSchedule.merged_with == schedule.merged_with,
                ClassSchedule.id != schedule_id
            ).all()
            if len(remaining_records) == 0:
                # 没有其他次记录了，主记录也恢复
                _restore_from_snapshot(main_record)
            else:
                # Bug 8 fix: 还有其他次记录时，更新主记录 notes 中的班级列表
                remaining_names = [r.class_.name for r in remaining_records if r.class_]
                if remaining_names:
                    main_record.notes = f'合班主记录（含 {", ".join(remaining_names)}）'
    
    # 恢复当前记录
    _restore_from_snapshot(schedule)
    
    # 拆分后重检冲突（日期可能恢复到不同的位置）
    db.session.flush()
    affected_dates = set()
    affected_dates.add(schedule.scheduled_date)
    for sec in secondary_records:
        if sec.scheduled_date:
            affected_dates.add(sec.scheduled_date)
    if schedule.merged_with:
        main_record = ClassSchedule.query.get(schedule.merged_with)
        if main_record and main_record.scheduled_date:
            affected_dates.add(main_record.scheduled_date)
    _recheck_conflicts_for_dates(affected_dates)

    # 清理 merge_config 中对应的合班关系
    try:
        topic_id = schedule.topic_id
        if secondary_records:
            # 拆的是主记录 → 删除该主班的所有 merge_config
            for sec in secondary_records:
                MergeConfig.query.filter_by(
                    topic_id=topic_id,
                    primary_class_id=schedule.class_id,
                    merged_class_id=sec.class_id
                ).delete()
        else:
            # 拆的是次记录 → 只删除这个并入班的 merge_config
            MergeConfig.query.filter_by(
                topic_id=topic_id,
                merged_class_id=schedule.class_id
            ).delete()
    except Exception:
        pass  # merge_config 清理失败不影响拆分本身

    db.session.commit()
    return jsonify({'message': '拆分成功，已恢复原始状态', 'schedule': schedule.to_dict()})


# ==================== 月度计划发布 ====================

@schedule_bp.route('/publish-checklist', methods=['GET'])
def publish_checklist():
    """发布前冲突处置清单"""
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    if not year or not month:
        return jsonify({'error': 'Missing year/month'}), 400

    checklist = _build_publish_checklist(year, month)
    return jsonify({
        'year': year,
        'month': month,
        'checklist': checklist
    })


@schedule_bp.route('/publish', methods=['POST'])
def publish_month():
    """发布月度计划"""
    data = request.get_json()
    year = data.get('year')
    month = data.get('month')
    force_publish = bool(data.get('force_publish', False))
    force_note = (data.get('force_note') or '').strip()

    checklist = _build_publish_checklist(year, month)
    unresolved = checklist.get('unresolved', [])
    if unresolved and not force_publish:
        return jsonify({
            'error': '存在无法自动解决的冲突，默认阻止发布',
            'code': 'UNRESOLVED_CONFLICTS',
            'checklist': checklist
        }), 409

    if force_publish and unresolved and not force_note:
        return jsonify({
            'error': '强制发布必须填写备注',
            'code': 'FORCE_NOTE_REQUIRED'
        }), 400

    # find or create monthly plan
    plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
    if not plan:
        plan = MonthlyPlan(year=year, month=month)
        db.session.add(plan)

    plan.status = 'published'
    plan.published_at = datetime.now()

    # update month schedules
    start_date, end_date = _month_range(year, month)

    # 状态简化：发布只改 MonthlyPlan.status，不再改 ClassSchedule.status
    # ClassSchedule.status 保持 scheduled/completed 不变

    if force_publish and unresolved:
        for item in unresolved:
            sid = item.get('schedule_id')
            schedule = ClassSchedule.query.get(sid)
            if schedule:
                schedule.notes = f"{schedule.notes or ''}\n[强制发布备注] {force_note}".strip()

    db.session.commit()

    return jsonify({
        'message': f'{year}年{month}月计划已发布',
        'plan': plan.to_dict(),
        'forced': force_publish,
        'forced_conflict_count': len(unresolved) if force_publish else 0
    })


@schedule_bp.route('/unpublish', methods=['POST'])
def unpublish_month():
    """取消发布月度计划，回退为草稿"""
    data = request.get_json()
    year = data.get('year')
    month = data.get('month')

    if not year or not month:
        return jsonify({'error': '缺少 year/month'}), 400

    plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
    if not plan or plan.status != 'published':
        return jsonify({'error': '该月计划未发布，无需取消'}), 400

    plan.status = 'draft'
    plan.published_at = None
    plan.updated_at = datetime.now()

    # 状态简化：取消发布只改 MonthlyPlan.status，不再改 ClassSchedule.status
    start_date, end_date = _month_range(year, month)

    db.session.commit()

    return jsonify({
        'message': f'{year}年{month}月计划已取消发布，回退为草稿',
        'plan': plan.to_dict()
    })


# ==================== 草稿保存 ====================

@schedule_bp.route('/save-draft', methods=['POST'])
def save_draft():
    """保存当前月度课表为草稿"""
    data = request.get_json()
    year = data.get('year')
    month = data.get('month')
    
    if not year or not month:
        return jsonify({'error': '缺少 year/month'}), 400
    
    plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
    if not plan:
        plan = MonthlyPlan(year=year, month=month, status='draft')
        db.session.add(plan)
    
    plan.updated_at = datetime.now()
    # 如果已发布，不允许再存草稿（需要先解除发布）
    if plan.status == 'published':
        return jsonify({'error': '该月计划已发布，无法保存草稿'}), 400
    
    db.session.commit()
    
    return jsonify({
        'message': '草稿已保存',
        'plan': plan.to_dict()
    })


# ==================== 约束条件管理 ====================

def _get_or_create_plan(year, month):
    """获取或创建月度计划（辅助函数）"""
    plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
    if not plan:
        plan = MonthlyPlan(year=year, month=month, status='draft')
        db.session.add(plan)
        db.session.flush()  # 获取 ID
    return plan


def _load_db_constraints(year, month):
    """从数据库加载当月所有启用的约束条件，转换为排课算法所需的 dict 格式"""
    plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
    if not plan:
        return {}
    
    db_constraints = ScheduleConstraint.query.filter_by(
        monthly_plan_id=plan.id,
        is_active=True
    ).all()
    
    if not db_constraints:
        return {}
    
    result = {
        'blocked_dates': [],
        'teacher_unavailable': [],
        'homeroom_unavailable': []
    }
    
    for c in db_constraints:
        # 优先使用 parsed_data（AI 结构化解析的结果）
        if c.parsed_data:
            try:
                parsed = _json.loads(c.parsed_data) if isinstance(c.parsed_data, str) else c.parsed_data
                # parsed_data 可能直接就是约束的一部分
                if isinstance(parsed, dict):
                    for key in ['blocked_dates', 'teacher_unavailable', 'homeroom_unavailable']:
                        if key in parsed and isinstance(parsed[key], list):
                            result[key].extend(parsed[key])
                continue
            except (ValueError, TypeError):
                pass
        
        # 没有 parsed_data 则根据 constraint_type 尝试从 description 推断
        desc = c.description or ''
        if c.constraint_type == 'blocked_date':
            # 尝试从描述中提取日期
            import re
            date_matches = re.findall(r'(\d{4}-\d{2}-\d{2})', desc)
            for d in date_matches:
                result['blocked_dates'].append({'date': d, 'reason': desc})
        elif c.constraint_type == 'teacher_unavailable':
            import re
            date_matches = re.findall(r'(\d{4}-\d{2}-\d{2})', desc)
            if date_matches:
                # 尝试提取讲师名（描述开头通常是讲师名）
                name_match = re.match(r'^(.+?)[\s:：]', desc)
                teacher_name = name_match.group(1) if name_match else desc.split()[0] if desc.split() else '未知'
                result['teacher_unavailable'].append({
                    'teacher_name': teacher_name,
                    'dates': date_matches,
                    'reason': desc
                })
        elif c.constraint_type == 'homeroom_unavailable':
            import re
            date_matches = re.findall(r'(\d{4}-\d{2}-\d{2})', desc)
            if date_matches:
                name_match = re.match(r'^(.+?)[\s:：]', desc)
                hr_name = name_match.group(1) if name_match else desc.split()[0] if desc.split() else '未知'
                result['homeroom_unavailable'].append({
                    'homeroom_name': hr_name,
                    'dates': date_matches,
                    'reason': desc
                })
        # custom 类型无法自动解析，跳过
    
    return result


def _merge_constraints(frontend_constraints, db_constraints):
    """合并前端传入的约束和数据库加载的约束（去重）"""
    merged = {
        'blocked_dates': list(frontend_constraints.get('blocked_dates', [])),
        'teacher_unavailable': list(frontend_constraints.get('teacher_unavailable', [])),
        'homeroom_unavailable': list(frontend_constraints.get('homeroom_unavailable', []))
    }
    
    # 去重合并 blocked_dates（按 date 字段去重）
    existing_dates = {b.get('date') if isinstance(b, dict) else b for b in merged['blocked_dates']}
    for b in db_constraints.get('blocked_dates', []):
        d = b.get('date') if isinstance(b, dict) else b
        if d not in existing_dates:
            merged['blocked_dates'].append(b)
            existing_dates.add(d)
    
    # 去重合并 teacher_unavailable（按 teacher_name 合并 dates）
    teacher_map = {}
    for t in merged['teacher_unavailable']:
        name = t.get('teacher_name', '')
        if name not in teacher_map:
            teacher_map[name] = {'teacher_name': name, 'dates': list(t.get('dates', [])), 'reason': t.get('reason', '')}
        else:
            existing = set(teacher_map[name]['dates'])
            for d in t.get('dates', []):
                if d not in existing:
                    teacher_map[name]['dates'].append(d)
    for t in db_constraints.get('teacher_unavailable', []):
        name = t.get('teacher_name', '')
        if name not in teacher_map:
            teacher_map[name] = {'teacher_name': name, 'dates': list(t.get('dates', [])), 'reason': t.get('reason', '')}
        else:
            existing = set(teacher_map[name]['dates'])
            for d in t.get('dates', []):
                if d not in existing:
                    teacher_map[name]['dates'].append(d)
    merged['teacher_unavailable'] = list(teacher_map.values())
    
    # 去重合并 homeroom_unavailable（按 homeroom_name 合并 dates）
    hr_map = {}
    for h in merged['homeroom_unavailable']:
        name = h.get('homeroom_name', '')
        if name not in hr_map:
            hr_map[name] = {'homeroom_name': name, 'dates': list(h.get('dates', [])), 'reason': h.get('reason', '')}
        else:
            existing = set(hr_map[name]['dates'])
            for d in h.get('dates', []):
                if d not in existing:
                    hr_map[name]['dates'].append(d)
    for h in db_constraints.get('homeroom_unavailable', []):
        name = h.get('homeroom_name', '')
        if name not in hr_map:
            hr_map[name] = {'homeroom_name': name, 'dates': list(h.get('dates', [])), 'reason': h.get('reason', '')}
        else:
            existing = set(hr_map[name]['dates'])
            for d in h.get('dates', []):
                if d not in existing:
                    hr_map[name]['dates'].append(d)
    merged['homeroom_unavailable'] = list(hr_map.values())
    
    # 保留前端约束中的其他字段（如 preferred_dates、merge_suggestions 等）
    for key in frontend_constraints:
        if key not in merged:
            merged[key] = frontend_constraints[key]
    
    return merged


@schedule_bp.route('/constraints', methods=['GET'])
def get_constraints():
    """获取指定月份的所有约束条件"""
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    
    if not year or not month:
        return jsonify({'error': '缺少 year/month'}), 400
    
    plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
    if not plan:
        return jsonify([])
    
    constraints = ScheduleConstraint.query.filter_by(
        monthly_plan_id=plan.id
    ).order_by(ScheduleConstraint.created_at).all()
    
    return jsonify([c.to_dict() for c in constraints])


@schedule_bp.route('/constraints', methods=['POST'])
def add_constraint():
    """添加约束条件 — AI解析成功才保存"""
    data = request.get_json()
    year = data.get('year')
    month = data.get('month')
    description = data.get('description', '').strip()
    
    if not year or not month:
        return jsonify({'error': '缺少 year/month'}), 400
    if not description:
        return jsonify({'error': '约束描述不能为空'}), 400
    
    # 1. 构建排课上下文
    teachers = Teacher.query.all()
    homerooms = Homeroom.query.all()
    classes = Class.query.filter(Class.status != 'completed').all()
    
    context = {
        'current_month': f'{year}-{str(month).zfill(2)}',
        'teachers': [t.name for t in teachers],
        'homeroom_teachers': [h.name for h in homerooms],
        'classes': [c.name for c in classes]
    }
    
    # 2. 调用AI解析
    parsed, error = call_ai_extract(description, context)
    if error:
        return jsonify({'error': f'AI解析失败: {error}'}), 500
    
    # 3. AI成功 → 保存约束 + 解析结果
    plan = _get_or_create_plan(year, month)
    
    constraint = ScheduleConstraint(
        monthly_plan_id=plan.id,
        constraint_type=data.get('constraint_type', 'custom'),
        description=description,
        parsed_data=_json.dumps(parsed, ensure_ascii=False) if parsed else None,
        is_active=True
    )
    db.session.add(constraint)
    db.session.commit()
    
    return jsonify(constraint.to_dict()), 201


@schedule_bp.route('/constraints/<int:constraint_id>', methods=['PUT'])
def update_constraint(constraint_id):
    """更新约束条件（主要用于切换启用/禁用）"""
    c = ScheduleConstraint.query.get_or_404(constraint_id)
    data = request.get_json()
    
    if 'is_active' in data:
        c.is_active = data['is_active']
    if 'description' in data:
        c.description = data['description']
    if 'constraint_type' in data:
        c.constraint_type = data['constraint_type']
    if 'parsed_data' in data:
        c.parsed_data = _json.dumps(data['parsed_data']) if data['parsed_data'] else None
    
    db.session.commit()
    return jsonify(c.to_dict())


@schedule_bp.route('/constraints/<int:constraint_id>', methods=['DELETE'])
def delete_constraint(constraint_id):
    """删除约束条件"""
    c = ScheduleConstraint.query.get_or_404(constraint_id)
    db.session.delete(c)
    db.session.commit()
    return jsonify({'message': '约束已删除'})


# ==================== 合班配置管理 ====================

@schedule_bp.route('/merge-config', methods=['GET'])
def get_merge_configs():
    """获取指定月份的所有合班配置"""
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)

    if not year or not month:
        return jsonify({'error': '缺少 year/month'}), 400

    plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
    if not plan:
        return jsonify([])

    configs = MergeConfig.query.filter_by(monthly_plan_id=plan.id).all()
    return jsonify([c.to_dict() for c in configs])


@schedule_bp.route('/merge-config', methods=['POST'])
def add_merge_config():
    """添加合班配置"""
    data = request.get_json()
    year = data.get('year')
    month = data.get('month')
    topic_id = data.get('topic_id')
    primary_class_id = data.get('primary_class_id')
    merged_class_id = data.get('merged_class_id')

    if not all([year, month, topic_id, primary_class_id, merged_class_id]):
        return jsonify({'error': '缺少必要参数'}), 400

    if primary_class_id == merged_class_id:
        return jsonify({'error': '主班和并入班不能相同'}), 400

    plan = _get_or_create_plan(year, month)

    # 检查是否已存在
    existing = MergeConfig.query.filter_by(
        monthly_plan_id=plan.id,
        topic_id=topic_id,
        merged_class_id=merged_class_id
    ).first()
    if existing:
        return jsonify({'error': '该合班关系已存在'}), 409

    config = MergeConfig(
        monthly_plan_id=plan.id,
        topic_id=topic_id,
        primary_class_id=primary_class_id,
        merged_class_id=merged_class_id,
        combo_id=data.get('combo_id'),
        combo_id_2=data.get('combo_id_2')
    )
    db.session.add(config)
    db.session.commit()

    return jsonify(config.to_dict()), 201


@schedule_bp.route('/merge-config/<int:config_id>', methods=['DELETE'])
def delete_merge_config(config_id):
    """删除合班配置"""
    config = MergeConfig.query.get_or_404(config_id)
    db.session.delete(config)
    db.session.commit()
    return jsonify({'message': '合班配置已删除'})



@schedule_bp.route('/check-holiday', methods=['GET'])
def check_holiday():
    """检查指定日期是否为节假日"""
    date_str = request.args.get('date')
    if not date_str:
        return jsonify({'error': '缺少date参数'}), 400
    
    result = is_holiday(date_str)
    return jsonify({
        'date': date_str,
        'is_holiday': result
    })


@schedule_bp.route('/<int:schedule_id>', methods=['DELETE'])
def delete_schedule(schedule_id):
    """删除单个课程安排"""
    schedule = ClassSchedule.query.get_or_404(schedule_id)
    
    # 历史遗留限制已移除：现在允许删除已完成的课程（兼容历史排课录入）
    
    # 如果是合班主课程，恢复其他课程的原始状态
    import json as _json_mod
    merged_schedules = ClassSchedule.query.filter_by(merged_with=schedule_id).all()
    for ms in merged_schedules:
        if ms.merge_snapshot:
            try:
                snap = _json_mod.loads(ms.merge_snapshot)
                if snap.get('scheduled_date'):
                    ms.scheduled_date = date.fromisoformat(snap['scheduled_date'])
                if 'combo_id' in snap:
                    ms.combo_id = snap['combo_id']
                if 'combo_id_2' in snap:
                    ms.combo_id_2 = snap['combo_id_2']
                ms.status = snap.get('status', 'scheduled')
                ms.conflict_type = snap.get('conflict_type')
                ms.notes = snap.get('notes')
            except (ValueError, _json_mod.JSONDecodeError):
                pass
            ms.merge_snapshot = None
        ms.merged_with = None
    
    # 记录被删记录的日期（删除前），以及恢复的合班次记录的日期
    affected_dates = set()
    if schedule.scheduled_date:
        affected_dates.add(schedule.scheduled_date)
    for ms in merged_schedules:
        if ms.scheduled_date:
            affected_dates.add(ms.scheduled_date)
    
    cid = schedule.class_id
    db.session.delete(schedule)
    db.session.flush()
    _recheck_conflicts_for_dates(affected_dates)
    db.session.commit()
    
    if cid:
        _resequence_topics_by_date(cid)
    
    return jsonify({'message': '课程已删除', 'id': schedule_id})


@schedule_bp.route('/<int:schedule_id>', methods=['GET'])
def get_schedule_detail(schedule_id):
    """获取单个排课详情"""
    schedule = ClassSchedule.query.get_or_404(schedule_id)
    return jsonify(schedule.to_dict())


# ==================== 可选地点查询 ====================

def _get_used_non_default_locations(class_id, default_city_id, exclude_schedule_id=None):
    """获取班级已使用过的非默认地点 ID 集合（含合班场景）"""
    used = set()
    query = ClassSchedule.query.filter(
        ClassSchedule.class_id == class_id,
        ClassSchedule.location_id.isnot(None),
        ClassSchedule.location_id != default_city_id
    )
    if exclude_schedule_id:
        query = query.filter(ClassSchedule.id != exclude_schedule_id)
    for s in query.all():
        used.add(s.location_id)
    return used


@schedule_bp.route('/<int:schedule_id>/available-locations', methods=['GET'])
def get_available_locations(schedule_id):
    """获取排课记录可选的上课地点（含历史排除逻辑）"""
    schedule = ClassSchedule.query.get_or_404(schedule_id)
    cls = schedule.class_
    if not cls:
        return jsonify({'error': '排课记录对应的班级不存在'}), 404
    
    default_city_id = cls.city_id
    all_cities = City.query.order_by(City.id).all()
    
    # 获取已使用过的非默认地点（排除当前记录本身）
    used_locations = _get_used_non_default_locations(cls.id, default_city_id, exclude_schedule_id=schedule_id)
    
    locations = []
    for city in all_cities:
        is_default = (city.id == default_city_id)
        is_disabled = (not is_default and city.id in used_locations)
        reason = '该班级已去过此地点' if is_disabled else None
        locations.append({
            'id': city.id,
            'name': city.name,
            'is_default': is_default,
            'is_disabled': is_disabled,
            'reason': reason
        })
    
    return jsonify({
        'default_location_id': default_city_id,
        'current_location_id': schedule.location_id,
        'locations': locations
    })


@schedule_bp.route('/class/<int:class_id>/available-locations', methods=['GET'])
def get_class_available_locations(class_id):
    """获取班级可选的上课地点（增加课次时使用，含历史排除逻辑）"""
    cls = Class.query.get_or_404(class_id)
    default_city_id = cls.city_id
    all_cities = City.query.order_by(City.id).all()
    
    used_locations = _get_used_non_default_locations(class_id, default_city_id)
    
    locations = []
    for city in all_cities:
        is_default = (city.id == default_city_id)
        is_disabled = (not is_default and city.id in used_locations)
        reason = '该班级已去过此地点' if is_disabled else None
        locations.append({
            'id': city.id,
            'name': city.name,
            'is_default': is_default,
            'is_disabled': is_disabled,
            'reason': reason
        })
    
    return jsonify({
        'default_location_id': default_city_id,
        'locations': locations
    })


# ==================== Excel 导出 ====================

def _status_cn(status, merged_with=None, notes=None, schedules_map=None):
    """状态中文映射
    schedules_map: 可选的 {id: ClassSchedule} 映射，避免 N+1 查询
    """
    if merged_with:
        # 次记录：查找主班名称
        main_sch = None
        if schedules_map and merged_with in schedules_map:
            main_sch = schedules_map[merged_with]
        else:
            main_sch = ClassSchedule.query.get(merged_with)
        if main_sch and main_sch.class_:
            return f'已合班至{main_sch.class_.name}'
        return '已合班'
    if notes and '合班主记录' in (notes or ''):
        # 主记录：提取合班班级名
        import re
        m = re.search(r'合班主记录（含\s*(.+?)）', notes or '')
        if m:
            return f'合班主记录（含{m.group(1)}）'
        return '合班主记录'
    mapping = {
        'scheduled': '✓',
        'completed': '已完成',
        'cancelled': '已取消',
    }
    return mapping.get(status, status or '-')


def _display_topic_name(schedule):
    """获取拼接仪式前缀的课题显示名（开班仪式+X / 结课典礼+X）"""
    topic_name = schedule.topic.name if schedule.topic else '-'
    seq = schedule.topic.sequence if schedule.topic else None
    total = schedule.class_.project.topics.count() if schedule.class_ and schedule.class_.project else 0
    if seq and total > 0:
        if seq == 1:
            return f'开班仪式+{topic_name}'
        elif seq == total:
            return f'结课典礼+{topic_name}'
    return topic_name


def _apply_header_style(ws, header_row, col_count):
    """为表头行应用蓝底白字加粗样式"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_font = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=11)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        bottom=Side(style='thin', color='B0C4DE')
    )
    for col in range(1, col_count + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border


def _auto_width(ws, min_width=8, max_width=30):
    """根据内容自动调整列宽"""
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        try:
            first_cell = col[0]
            # 跳过 MergedCell（合并单元格占位对象没有 column_letter）
            if isinstance(first_cell, MergedCell):
                continue
            col_letter = first_cell.column_letter
            lengths = []
            for cell in col:
                if isinstance(cell, MergedCell):
                    continue
                val = str(cell.value) if cell.value is not None else ''
                # 中文字符按2倍宽度估算
                char_len = sum(2 if ord(c) > 127 else 1 for c in val)
                lengths.append(char_len)
            best = max(lengths) if lengths else min_width
            ws.column_dimensions[col_letter].width = max(min_width, min(best + 2, max_width))
        except (AttributeError, TypeError):
            continue


@schedule_bp.route('/export/<int:year>/<int:month>', methods=['GET'])
def export_month_excel(year, month):
    """导出月度课表 Excel（给讲师、班主任等工作人员看）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    from flask import send_file

    # 月份起止
    start_date = date(year, month, 1)
    end_date = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)

    schedules = ClassSchedule.query.filter(
        ClassSchedule.scheduled_date >= start_date,
        ClassSchedule.scheduled_date < end_date
    ).order_by(ClassSchedule.scheduled_date).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f'{year}年{month}月排课'

    # ---- 标题行 ----
    title_text = f'北清商学院 {year}年{month}月 排课计划'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(name='Microsoft YaHei', bold=True, size=16, color='1E293B')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    # 副标题
    plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
    status_text = '已发布' if plan and plan.status == 'published' else '草稿'
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
    sub_cell = ws.cell(row=2, column=1, value=f'状态: {status_text}  |  导出时间: {datetime.now().strftime("%Y-%m-%d %H:%M")}')
    sub_cell.font = Font(name='Microsoft YaHei', size=10, color='64748B')
    sub_cell.alignment = Alignment(horizontal='center')

    # ---- 表头 ----
    headers = ['周次', '日期(周六)', '日期(周日)', '班级', '课题', '周六讲师', '周六课程', '周日讲师', '周日课程', '班主任', '状态']
    header_row = 4
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=col_idx, value=h)
    _apply_header_style(ws, header_row, len(headers))

    # ---- 按周分组填充数据 ----
    # 构建周结构
    curr_week_start = start_date - timedelta(days=start_date.weekday())
    week_idx = 0
    row = header_row + 1

    conflict_font = Font(name='Microsoft YaHei', color='DC2626', bold=True)
    completed_fill = PatternFill(start_color='F0FDF4', end_color='F0FDF4', fill_type='solid')
    merged_fill = PatternFill(start_color='FAF5FF', end_color='FAF5FF', fill_type='solid')
    merged_font = Font(name='Microsoft YaHei', size=10, color='7C3AED')
    normal_font = Font(name='Microsoft YaHei', size=10)
    center_align = Alignment(horizontal='center', vertical='center')

    # 预构建 ID→记录 映射，避免 _status_cn 中的 N+1 查询
    schedules_map = {s.id: s for s in schedules}

    while curr_week_start < end_date:
        week_end = curr_week_start + timedelta(days=6)
        week_idx += 1

        sat = curr_week_start + timedelta(days=(5 - curr_week_start.weekday()) % 7)
        sun = sat + timedelta(days=1)

        week_schedules = [s for s in schedules if curr_week_start <= s.scheduled_date <= week_end]

        if not week_schedules:
            curr_week_start += timedelta(weeks=1)
            continue

        for s in week_schedules:
            day1_teacher = s.combo.teacher.name if s.combo and s.combo.teacher else '待定'
            day1_course = s.combo.course_name if s.combo else '待定'
            day2_teacher = s.combo_2.teacher.name if s.combo_2 and s.combo_2.teacher else '待定'
            day2_course = s.combo_2.course_name if s.combo_2 else '待定'
            homeroom = s.class_.homeroom.name if s.class_ and s.class_.homeroom else '未分配'
            status = _status_cn(s.status, s.merged_with, s.notes, schedules_map=schedules_map)

            # 使用实际排课日期而非计算的周六
            actual_sat = s.scheduled_date
            actual_sun = s.scheduled_date + timedelta(days=1)

            values = [
                f'第{week_idx}周',
                actual_sat.strftime('%m/%d'),
                actual_sun.strftime('%m/%d'),
                s.class_.name if s.class_ else '-',
                _display_topic_name(s),
                day1_teacher,
                day1_course,
                day2_teacher,
                day2_course,
                homeroom,
                status
            ]

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = normal_font
                cell.alignment = center_align

            # 有冲突时：状态列显示冲突原因 + 红色
            if s.conflict_type:
                conflict_desc = s.conflict_type
                # 尝试从 notes 取更详细的冲突描述
                if s.notes and '撞课' in s.notes:
                    conflict_desc = s.notes
                ws.cell(row=row, column=11, value=conflict_desc)
                ws.cell(row=row, column=11).font = conflict_font
            # 已完成行浅绿底
            if s.status == 'completed':
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col_idx).fill = completed_fill
            # 合班次记录行紫色底
            if s.merged_with:
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col_idx).fill = merged_fill
                    ws.cell(row=row, column=col_idx).font = merged_font

            row += 1

        curr_week_start += timedelta(weeks=1)

    _auto_width(ws)

    # 写入流
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'排课计划_{year}年{month}月.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@schedule_bp.route('/export/class/<int:class_id>', methods=['GET'])
def export_class_excel(class_id):
    """导出单个班级课表 Excel（给学生看）"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    from flask import send_file

    cls = Class.query.get_or_404(class_id)
    schedules = ClassSchedule.query.filter_by(class_id=class_id)\
        .order_by(ClassSchedule.scheduled_date).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f'{cls.name} 课程表'

    # ---- 标题 ----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    title_cell = ws.cell(row=1, column=1, value=f'{cls.name} 课程安排表')
    title_cell.font = Font(name='Microsoft YaHei', bold=True, size=16, color='1E293B')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    # 班级信息行
    homeroom = cls.homeroom.name if cls.homeroom else '未分配'
    project_name = cls.project.name if cls.project else '-'
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    info_cell = ws.cell(row=2, column=1, value=f'项目: {project_name}  |  班主任: {homeroom}  |  导出时间: {datetime.now().strftime("%Y-%m-%d %H:%M")}')
    info_cell.font = Font(name='Microsoft YaHei', size=10, color='64748B')
    info_cell.alignment = Alignment(horizontal='center')

    # ---- 表头 ----
    headers = ['序号', '日期(周六)', '日期(周日)', '课题', '周六讲师', '周六课程', '周日讲师', '周日课程']
    header_row = 4
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=col_idx, value=h)
    _apply_header_style(ws, header_row, len(headers))

    # ---- 数据行 ----
    normal_font = Font(name='Microsoft YaHei', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    completed_fill = PatternFill(start_color='F0FDF4', end_color='F0FDF4', fill_type='solid')

    row = header_row + 1
    for idx, s in enumerate(schedules, 1):
        sat = s.scheduled_date
        sun = sat + timedelta(days=1) if sat else None

        day1_teacher = s.combo.teacher.name if s.combo and s.combo.teacher else '待定'
        day1_course = s.combo.course_name if s.combo else '待定'
        day2_teacher = s.combo_2.teacher.name if s.combo_2 and s.combo_2.teacher else '待定'
        day2_course = s.combo_2.course_name if s.combo_2 else '待定'

        values = [
            idx,
            sat.strftime('%Y-%m-%d') if sat else '-',
            sun.strftime('%Y-%m-%d') if sun else '-',
            _display_topic_name(s),
            day1_teacher,
            day1_course,
            day2_teacher,
            day2_course,
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = normal_font
            cell.alignment = center_align

        # 已完成行浅绿底
        if s.status == 'completed':
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row, column=col_idx).fill = completed_fill

        row += 1

    _auto_width(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'课程表_{cls.name}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ==================== 智能排课算法核心函数 ====================
from config import Config as _cfg

def _get_last_class_date(class_id, before_date=None):
    """获取班级最后一次上课/排课的日期（跨所有月份，但只看 before_date 之前的记录）"""
    query = ClassSchedule.query.filter(
        ClassSchedule.class_id == class_id,
        ClassSchedule.status.in_(['completed', 'scheduled'])
    )
    if before_date:
        query = query.filter(ClassSchedule.scheduled_date < before_date)
    last = query.order_by(ClassSchedule.scheduled_date.desc()).first()
    return last.scheduled_date if last else None


def _cleanup_stale_scheduled_records():
    """清理全库中课题已完成但仍有scheduled的残留记录。
    
    场景：班级创建时 auto_schedule_class 一口气排了所有课题到未来多个月，
    后来智能排课把某课题提前排在更早月份并完成了，但其他月份的旧 scheduled 记录残留。
    
    清理条件：同一 class_id + topic_id 既有 completed 又有 scheduled 记录时，
    删除 scheduled 的那条（课已经上过了，旧计划没意义了）。
    
    安全性：不删除有合班关系（merged_with）的记录，避免破坏合班数据。
    """
    # 找出所有"班级+课题"已有 completed 记录的组合
    completed_pairs = db.session.query(
        ClassSchedule.class_id, ClassSchedule.topic_id
    ).filter(
        ClassSchedule.status == 'completed'
    ).distinct().all()
    
    if not completed_pairs:
        return 0
    
    # 构建集合用于快速查找
    completed_set = set((cp.class_id, cp.topic_id) for cp in completed_pairs)
    
    # 找出匹配的 scheduled 残留记录（排除有合班关系的）
    stale = ClassSchedule.query.filter(
        ClassSchedule.status == 'scheduled',
        ClassSchedule.merged_with.is_(None)
    ).all()
    
    to_delete = []
    for s in stale:
        # "其他"课题(is_other)可重复排课，即使同课题已有completed也不算残留
        if s.topic and s.topic.is_other:
            continue
        if (s.class_id, s.topic_id) in completed_set:
            # 再检查：不删除有其他记录 merged_with 指向自己的（自己是合班主记录）
            has_dependents = ClassSchedule.query.filter(
                ClassSchedule.merged_with == s.id
            ).first()
            if not has_dependents:
                to_delete.append(s)
    
    for s in to_delete:
        db.session.delete(s)
    
    if to_delete:
        db.session.commit()
    
    return len(to_delete)


def _resequence_topics_by_date(class_id):
    """按排课日期重排课题的 sequence，确保序号和时间顺序一致。
    并在同时计算每个课次的实际上课顺序 (week_number)。
    在任何修改了排课日期的操作后调用。
    """
    schedules = ClassSchedule.query.filter_by(class_id=class_id)\
        .filter(ClassSchedule.scheduled_date.isnot(None))\
        .filter(db.or_(ClassSchedule.status != 'cancelled', ClassSchedule.status.is_(None)))\
        .order_by(ClassSchedule.scheduled_date).all()
    seen_topics = []
    changed = False

    # 更新课题 sequence
    for s in schedules:
        if s.topic_id and s.topic_id not in seen_topics:
            seen_topics.append(s.topic_id)
    if seen_topics:
        for i, tid in enumerate(seen_topics):
            topic = Topic.query.get(tid)
            if topic and topic.sequence != i + 1:
                topic.sequence = i + 1
                changed = True

    # 更新排课记录实际上课顺序 (week_number)
    # 对于合班记录(merged_with)，其周次顺序跟主记录独立，它相当于本班级自己的第N次课
    nth_class = 1
    for s in schedules:
        if s.week_number != nth_class:
            s.week_number = nth_class
            changed = True
        nth_class += 1

    if changed:
        db.session.commit()


def _optimize_combos_per_day(assignments, constraints, precomputed=None):
    """阶段二：按天优化科教组合分配，消除讲师冲突。
    
    贪心算法逐个排课时，先排的班级不知道后排班级会冲突，
    导致有些本可避免的讲师冲突留在了最终方案中。
    
    本函数在所有班级已经分配好日期后，对每一天的组合重新优化：
    对同一天的所有班级，穷举所有可行的 (周六组合, 周日组合) 分配，
    找到讲师冲突最少（理想为0）的方案。
    
    问题规模极小（每天最多7个班，每课题2-5个组合），穷举毫秒级。
    """
    from itertools import product as iter_product
    
    # 收集讲师请假约束（标准键名为 teacher_unavailable，格式为 {teacher_name, dates}）
    teacher_leave = set()  # {(teacher_id, date_str)}
    leave_list = constraints.get('teacher_unavailable', [])
    if isinstance(leave_list, list):
        from models import Teacher
        # 建立 name→id 映射以适配 teacher_unavailable 的 name-based 格式
        _teacher_name_to_id = {}
        for item in leave_list:
            if isinstance(item, dict):
                t_name = item.get('teacher_name')
                tid = item.get('teacher_id')  # 兼容旧格式
                if t_name and not tid:
                    if t_name not in _teacher_name_to_id:
                        t_obj = Teacher.query.filter_by(name=t_name).first()
                        _teacher_name_to_id[t_name] = t_obj.id if t_obj else None
                    tid = _teacher_name_to_id[t_name]
                if tid:
                    for d in item.get('dates', []):
                        teacher_leave.add((tid, d))
    
    # 预加载所有需要的组合数据（从 precomputed 或 DB）
    topic_ids = set(a.get('topic_id') for a in assignments if a.get('topic_id'))
    combos_by_topic = {}  # {topic_id: [combo1, combo2, ...]}
    if precomputed:
        # 从 precomputed.class_infos 收集所有 topic→combos 映射
        for info in precomputed.get('class_infos', []):
            tid = info.get('next_topic').id if info.get('next_topic') else None
            if tid and tid not in combos_by_topic:
                combos_by_topic[tid] = info.get('all_combos', [])
    # 兜底：对 precomputed 中没有的 topic 从 DB 查
    for tid in topic_ids:
        if tid not in combos_by_topic:
            combos_by_topic[tid] = TeacherCourseCombo.query.filter_by(topic_id=tid)\
                .order_by(TeacherCourseCombo.priority.desc(), TeacherCourseCombo.id.desc()).all()
    
    # 建立 combo_id -> combo 对象的映射（用于冲突检查）
    combo_cache = {}  # {combo_id: combo_obj}
    for combos in combos_by_topic.values():
        for c in combos:
            combo_cache[c.id] = c
    
    # 按日期分组（只看有日期且非合班目标的记录）
    by_date = {}
    for i, a in enumerate(assignments):
        d = a.get('assigned_date')
        if d and not a.get('is_merged_target') and not a.get('skip_reason'):
            by_date.setdefault(d, []).append(i)
    
    for date_str, indices in by_date.items():
        if len(indices) <= 1:
            continue  # 单个班级不可能有讲师冲突
        
        # 构建每个班级在这一天的可选组合对列表
        slot_options = []  # [(assignment_index, [(combo1_obj, combo2_obj), ...], is_locked)]
        
        for idx in indices:
            a = assignments[idx]
            
            # 用户手动锁定的组合不参与优化
            if a.get('user_locked_combo') or a.get('is_override'):
                c1 = combo_cache.get(a.get('combo_id'))
                c2 = combo_cache.get(a.get('combo_id_2'))
                slot_options.append((idx, [(c1, c2)], True))
                continue
            
            # 获取该课题的所有可用组合（从缓存）
            topic_id = a.get('topic_id')
            all_combos = combos_by_topic.get(topic_id, [])
            if not all_combos:
                continue
            
            # 过滤掉讲师请假的组合
            available = [c for c in all_combos 
                        if not (c.teacher_id and (c.teacher_id, date_str) in teacher_leave)]
            if not available:
                available = all_combos
            
            # 生成所有 (周六, 周日) 组合对
            pairs = []
            for c1 in available:
                for c2 in available:
                    pairs.append((c1, c2))
            
            # 排序：优先不同讲师、高优先级
            def pair_sort_key(p):
                same_teacher = 1 if (p[0].teacher_id == p[1].teacher_id) else 0
                # priority = -(p[0].priority + p[1].priority)
                priority = -((p[0].priority or 0) + (p[1].priority or 0))
                return (same_teacher, priority)
            pairs.sort(key=pair_sort_key)
            
            slot_options.append((idx, pairs, False))
        
        if len(slot_options) < 2:
            continue
        
        # 检查当前方案是否已经无冲突
        current_conflict_count = _count_teacher_conflicts_for_day(
            slot_options, assignments, combo_cache
        )
        if current_conflict_count == 0:
            continue  # 已经没冲突了，跳过
        
        print(f'[phase2] 日期 {date_str}: {len(indices)}个班级, 当前冲突 {current_conflict_count}', flush=True)
        
        # 穷举搜索：找讲师冲突为0的方案
        all_pair_lists = [so[1] for so in slot_options]
        
        # 限制搜索空间
        total = 1
        for pl in all_pair_lists:
            total *= len(pl)

        if total <= 50000:
            # 穷举搜索（小规模，毫秒级）
            best_assignment = None
            best_conflict_count = float('inf')
            
            for combo_tuple in iter_product(*all_pair_lists):
                sat_teachers = {}
                sun_teachers = {}
                conflicts = 0
                
                for i, (c1, c2) in enumerate(combo_tuple):
                    if c1 and c1.teacher_id:
                        if c1.teacher_id in sat_teachers:
                            conflicts += 1
                        sat_teachers[c1.teacher_id] = slot_options[i][0]
                    
                    if c2 and c2.teacher_id:
                        if c2.teacher_id in sun_teachers:
                            conflicts += 1
                        sun_teachers[c2.teacher_id] = slot_options[i][0]
                
                if conflicts < best_conflict_count:
                    best_conflict_count = conflicts
                    best_assignment = combo_tuple
                    if conflicts == 0:
                        break
        else:
            # 贪心策略（大规模，毫秒级）：按选择最少优先排序，逐个分配最优组合
            indexed_slots = list(enumerate(slot_options))
            indexed_slots.sort(key=lambda x: len(x[1][1]))  # 选择少的先排
            
            sat_teachers_used = {}  # {teacher_id: True}
            sun_teachers_used = {}
            best_assignment_list = [None] * len(slot_options)
            best_conflict_count = 0
            
            for orig_i, (idx, pairs, is_locked) in indexed_slots:
                best_pair = pairs[0]  # 默认
                best_pair_conflicts = float('inf')
                
                for c1, c2 in pairs:
                    c = 0
                    if c1 and c1.teacher_id and c1.teacher_id in sat_teachers_used:
                        c += 1
                    if c2 and c2.teacher_id and c2.teacher_id in sun_teachers_used:
                        c += 1
                    if c < best_pair_conflicts:
                        best_pair_conflicts = c
                        best_pair = (c1, c2)
                        if c == 0:
                            break
                
                best_assignment_list[orig_i] = best_pair
                best_conflict_count += best_pair_conflicts
                c1, c2 = best_pair
                if c1 and c1.teacher_id:
                    sat_teachers_used[c1.teacher_id] = True
                if c2 and c2.teacher_id:
                    sun_teachers_used[c2.teacher_id] = True
            
            best_assignment = tuple(best_assignment_list)
        
        # 应用最优方案
        if best_assignment and best_conflict_count < current_conflict_count:
            print(f'[phase2] → 优化后冲突: {best_conflict_count}', flush=True)
            for i, (c1, c2) in enumerate(best_assignment):
                idx = slot_options[i][0]
                is_locked = slot_options[i][2]
                if is_locked:
                    continue
                
                a = assignments[idx]
                old_c1_id = a.get('combo_id')
                old_c2_id = a.get('combo_id_2')
                new_c1_id = c1.id if c1 else None
                new_c2_id = c2.id if c2 else None
                
                if new_c1_id != old_c1_id or new_c2_id != old_c2_id:
                    # 先记录原始信息再覆盖
                    orig_t1 = a.get('combo1_teacher_name', '?')
                    orig_t2 = a.get('combo2_teacher_name', '?')
                    a['combo_id'] = new_c1_id
                    a['combo_id_2'] = new_c2_id
                    a['combo1_teacher_name'] = c1.teacher.name if c1 and c1.teacher else None
                    a['combo2_teacher_name'] = c2.teacher.name if c2 and c2.teacher else None
                    a['combo1_course_name'] = c1.course_name if c1 else None
                    a['combo2_course_name'] = c2.course_name if c2 else None
                    a['combo_switched'] = True
                    a['combo_switch_info'] = {
                        'original_combo1_teacher': orig_t1,
                        'original_combo2_teacher': orig_t2,
                        'reason': 'phase2_optimization'
                    }
                
                # 清除旧冲突标记中的讲师冲突（保留班主任等其他冲突）
                if best_conflict_count == 0:
                    old_conflicts = a.get('conflicts', [])
                    a['conflicts'] = [c for c in old_conflicts if '讲师' not in c]
                    if not a['conflicts']:
                        a['conflict_type'] = None
        else:
            print(f'[phase2] → 无法消除! 最优仍有 {best_conflict_count} 个冲突 (穷举了所有组合)', flush=True)


def _count_teacher_conflicts_for_day(slot_options, assignments, combo_cache):
    """统计一天内当前组合分配的讲师冲突数（使用缓存避免DB查询）"""
    sat_teachers = {}
    sun_teachers = {}
    conflicts = 0
    
    for so in slot_options:
        idx = so[0]
        a = assignments[idx]
        c1_id = a.get('combo_id')
        c2_id = a.get('combo_id_2')
        
        if c1_id:
            c1 = combo_cache.get(c1_id)
            if c1 and c1.teacher_id:
                if c1.teacher_id in sat_teachers:
                    conflicts += 1
                sat_teachers[c1.teacher_id] = idx
        
        if c2_id:
            c2 = combo_cache.get(c2_id)
            if c2 and c2.teacher_id:
                if c2.teacher_id in sun_teachers:
                    conflicts += 1
                sun_teachers[c2.teacher_id] = idx
    
    return conflicts


def _calculate_urgency(class_id, reference_date):
    """
    计算班级排课紧迫度。
    返回: (urgency_score, last_date)
    - urgency > 1.0 → 超过目标间隔，越大越急
    - urgency = 999  → 新班级从未排过课
    """
    last_date = _get_last_class_date(class_id, before_date=reference_date)
    if last_date is None:
        return 999.0, None
    days_since = (reference_date - last_date).days
    if days_since <= 0:
        return 0.01, last_date
    urgency = days_since / _cfg.TARGET_INTERVAL_DAYS
    return urgency, last_date


def _is_blocked(check_date, constraints):
    """检查日期是否在封锁日期列表中"""
    blocked = constraints.get('blocked_dates', [])
    for b in blocked:
        b_date = b.get('date') if isinstance(b, dict) else b
        if b_date == check_date.isoformat():
            return True
    return False


def _check_homeroom_unavailable(cls, sat, constraints):
    """检查班主任是否在上课日（周六/周日）有请假/不可用"""
    if not cls.homeroom_id or not cls.homeroom:
        return None
    homeroom_unavailable = constraints.get('homeroom_unavailable', [])
    hrm_name = cls.homeroom.name
    sun = sat + timedelta(days=1)  # 周日
    teaching_days = {sat, sun}  # 只匹配实际上课日

    for u in homeroom_unavailable:
        if u.get('homeroom_name') != hrm_name:
            continue
        dates_raw = u.get('dates', [])
        parsed = []
        for item in dates_raw:
            if isinstance(item, str):
                try:
                    parsed.append(date.fromisoformat(item))
                except Exception:
                    if '~' in item:
                        parts = item.split('~')
                        try:
                            d1 = date.fromisoformat(parts[0])
                            d2 = date.fromisoformat(parts[1])
                            d = d1
                            while d <= d2:
                                parsed.append(d)
                                d = d + timedelta(days=1)
                        except Exception:
                            pass
            elif isinstance(item, dict):
                f = item.get('from') or item.get('start')
                t = item.get('to') or item.get('end')
                try:
                    d1 = date.fromisoformat(f)
                    d2 = date.fromisoformat(t)
                    d = d1
                    while d <= d2:
                        parsed.append(d)
                        d = d + timedelta(days=1)
                except Exception:
                    pass
        for pd in parsed:
            if pd in teaching_days:
                day_label = '周六' if pd == sat else '周日'
                return f'班主任 {hrm_name} {pd.isoformat()}({day_label})请假'
    return None


def _score_candidate(cls, sat, last_date, combo1, combo2, assigned_map, constraints, homeroom_overrides=None, precomputed=None):
    """
    评估某个候选周六对某班级的综合得分。
    
    当 precomputed 传入时，使用预缓存的内存索引（纯dict查找，无DB查询）。
    当 precomputed 为 None 时，回退到DB查询（兼容旧调用方式）。
    """
    if homeroom_overrides is None:
        homeroom_overrides = {}
    conflict_reasons = []
    merge_suggestions = []
    sun = sat + timedelta(days=1)

    # 提取预计算索引（如果存在）
    _hr_by_date = precomputed.get('homeroom_by_date', {}) if precomputed else {}
    _t1_by_date = precomputed.get('teacher_day1_by_date', {}) if precomputed else {}
    _t2_by_date = precomputed.get('teacher_day2_by_date', {}) if precomputed else {}
    _rc_city = precomputed.get('room_count_by_date_city', {}) if precomputed else {}
    _rc_all = precomputed.get('room_count_by_date', {}) if precomputed else {}
    _topic_idx = precomputed.get('topic_by_date', {}) if precomputed else {}
    _cls_names = precomputed.get('class_name_cache', {}) if precomputed else {}
    use_cache = precomputed is not None

    # --- H2: 节假日（双重保险，候选池理论上已过滤）---
    if is_holiday(sat) or is_holiday(sun):
        return (0.0, True, ['节假日冲突'], [])

    # --- H6: 封锁日期 ---
    blocked = constraints.get('blocked_dates', [])
    for b in blocked:
        b_date = b.get('date') if isinstance(b, dict) else b
        if b_date == sat.isoformat() or b_date == sun.isoformat():
            reason = b.get('reason', '人工约束') if isinstance(b, dict) else '人工约束'
            return (0.0, True, [f'封锁日期({reason})'], [])

    # --- S1: 间隔硬下限（< MIN_INTERVAL_DAYS 直接拒绝）---
    if last_date:
        interval = (sat - last_date).days
        if interval < _cfg.MIN_INTERVAL_DAYS:
            return (0.0, True, [f'间隔过短({interval}天 < {_cfg.MIN_INTERVAL_DAYS}天)'], [])
    
    # --- H3: 班主任不能同日双班 ---
    effective_homeroom_id = homeroom_overrides.get(cls.id, cls.homeroom_id)
    homeroom_conflict = False
    if effective_homeroom_id:
        if use_cache:
            # 纯内存查找
            hr_classes = _hr_by_date.get(sat, {}).get(effective_homeroom_id, [])
            for cn in hr_classes:
                if cn != cls.name:
                    homeroom_conflict = True
                    conflict_reasons.append(f'班主任撞课({cn})')
                    break
        else:
            # 回退到DB查询
            same_day_records = ClassSchedule.query.filter(
                ClassSchedule.scheduled_date == sat,
                ClassSchedule.class_id != cls.id
            ).all()
            for existing in same_day_records:
                existing_effective = homeroom_overrides.get(
                    existing.class_id,
                    existing.homeroom_override_id or (existing.class_.homeroom_id if existing.class_ else None)
                )
                if existing_effective == effective_homeroom_id:
                    homeroom_conflict = True
                    conflict_reasons.append(f'班主任撞课({existing.class_.name if existing.class_ else "?"})')
                    break

        # 查本轮 assigned_map
        for a_cls_id, a_info in assigned_map.items():
            if a_info['date'] == sat and a_cls_id != cls.id:
                a_effective = homeroom_overrides.get(a_cls_id)
                if a_effective is None:
                    # 优先从 assigned_map 中读取 homeroom_id（已在分配时存入）
                    a_effective = a_info.get('homeroom_id')
                    if a_effective is None and not use_cache:
                        a_cls = Class.query.get(a_cls_id)
                        if a_cls:
                            a_effective = a_cls.homeroom_id
                if a_effective == effective_homeroom_id:
                    homeroom_conflict = True
                    a_name = _cls_names.get(a_cls_id) or (a_info.get('class_name') or '?')
                    conflict_reasons.append(f'班主任撞课({a_name})[本轮]')

    # --- H7: 班主任请假（硬约束：直接跳过该日期） ---
    # 注意：不再受 homeroom_conflict(H3) 遮蔽，始终独立检查
    if cls.homeroom_id:
        hrm_reason = _check_homeroom_unavailable(cls, sat, constraints)
        if hrm_reason:
            conflict_reasons.append(hrm_reason)
            return (0.0, True, conflict_reasons, [])

    # --- S5: 班主任连续带班间隔 >= 14天（软约束：降分警告） ---
    homeroom_gap_warning = False
    min_hr_gap = 14  # 天
    if effective_homeroom_id and not homeroom_conflict:
        if use_cache:
            # 检查前后各1-2周内是否有同班主任的排课
            for delta_weeks in [1, -1, 2, -2]:
                adj_sat = sat + timedelta(weeks=delta_weeks)
                hr_classes_adj = _hr_by_date.get(adj_sat, {}).get(effective_homeroom_id, [])
                if hr_classes_adj:
                    gap = abs(delta_weeks * 7)
                    if gap < min_hr_gap:
                        adj_cls_names = ', '.join(cn for cn in hr_classes_adj if cn != cls.name)
                        if adj_cls_names:
                            homeroom_gap_warning = True
                            conflict_reasons.append(f'班主任连续带班(间隔{gap}天<{min_hr_gap}天, {adj_sat.isoformat()} {adj_cls_names})')
                            break
        # 也检查本轮 assigned_map 中的分配
        for a_cls_id, a_info in assigned_map.items():
            if a_cls_id != cls.id:
                a_hr = homeroom_overrides.get(a_cls_id, a_info.get('homeroom_id'))
                if a_hr == effective_homeroom_id:
                    gap = abs((sat - a_info['date']).days)
                    if 0 < gap < min_hr_gap:
                        a_name = _cls_names.get(a_cls_id) or (a_info.get('class_name') or '?')
                        homeroom_gap_warning = True
                        conflict_reasons.append(f'班主任连续带班({a_name}, 间隔{gap}天<{min_hr_gap}天)[本轮]')

    # --- H4: 讲师 Day1(周六) 不能撞课 ---
    teacher_conflict = False
    if combo1 and combo1.teacher:
        t1_id = combo1.teacher_id
        if use_cache:
            t1_classes = _t1_by_date.get(sat, {}).get(t1_id, [])
            for cn in t1_classes:
                if cn != cls.name:
                    teacher_conflict = True
                    conflict_reasons.append(f'周六讲师 {combo1.teacher.name} 撞课({cn})')
                    break
        else:
            t1_db = ClassSchedule.query.join(
                TeacherCourseCombo, ClassSchedule.combo_id == TeacherCourseCombo.id
            ).filter(
                TeacherCourseCombo.teacher_id == t1_id,
                ClassSchedule.scheduled_date == sat,
                ClassSchedule.class_id != cls.id
            ).first()
            if t1_db:
                teacher_conflict = True
                conflict_reasons.append(f'周六讲师 {combo1.teacher.name} 撞课({t1_db.class_.name})')
        # 查本轮 assigned_map
        for a_cls_id, a_info in assigned_map.items():
            if a_info['date'] == sat and a_cls_id != cls.id:
                if a_info.get('combo1_teacher_id') == t1_id:
                    a_name = _cls_names.get(a_cls_id) or (a_info.get('class_name') or '?')
                    teacher_conflict = True
                    conflict_reasons.append(f'周六讲师 {combo1.teacher.name} 撞课({a_name})[本轮]')

    # --- H5: 讲师 Day2(周日) 不能撞课 ---
    if combo2 and combo2.teacher:
        t2_id = combo2.teacher_id
        if use_cache:
            t2_classes = _t2_by_date.get(sat, {}).get(t2_id, [])
            for cn in t2_classes:
                if cn != cls.name:
                    teacher_conflict = True
                    conflict_reasons.append(f'周日讲师 {combo2.teacher.name} 撞课({cn})')
                    break
        else:
            schedules_on_sat = ClassSchedule.query.filter(
                ClassSchedule.scheduled_date == sat,
                ClassSchedule.class_id != cls.id
            ).all()
            for existing_s in schedules_on_sat:
                if existing_s.combo_id_2:
                    c2 = TeacherCourseCombo.query.get(existing_s.combo_id_2)
                    if c2 and c2.teacher_id == t2_id:
                        teacher_conflict = True
                        conflict_reasons.append(f'周日讲师 {combo2.teacher.name} 撞课({existing_s.class_.name})')
                        break
        # 查本轮 assigned_map
        for a_cls_id, a_info in assigned_map.items():
            if a_info['date'] == sat and a_cls_id != cls.id:
                if a_info.get('combo2_teacher_id') == t2_id:
                    a_name = _cls_names.get(a_cls_id) or (a_info.get('class_name') or '?')
                    teacher_conflict = True
                    conflict_reasons.append(f'周日讲师 {combo2.teacher.name} 撞课({a_name})[本轮]')

    # --- H8: 讲师请假 ---
    unavailable = constraints.get('teacher_unavailable', [])
    t1_name = combo1.teacher.name if combo1 and combo1.teacher else None
    t2_name = combo2.teacher.name if combo2 and combo2.teacher else None
    sat_str = sat.isoformat()
    sun_str = sun.isoformat()
    if t1_name:
        for u in unavailable:
            if u.get('teacher_name') == t1_name and sat_str in u.get('dates', []):
                teacher_conflict = True
                conflict_reasons.append(f'周六讲师 {t1_name} 请假')
                break
    if t2_name:
        for u in unavailable:
            if u.get('teacher_name') == t2_name and sun_str in u.get('dates', []):
                teacher_conflict = True
                conflict_reasons.append(f'周日讲师 {t2_name} 请假')
                break

    # --- S4: 教室容量 ---
    cls_city_id = cls.city_id
    if use_cache:
        if cls_city_id:
            db_count = _rc_city.get((sat, cls_city_id), 0)
        else:
            db_count = _rc_all.get(sat, 0)
    else:
        if cls_city_id:
            db_count = ClassSchedule.query.filter(
                ClassSchedule.scheduled_date == sat,
                ClassSchedule.status.in_(['scheduled', 'completed']),
                ClassSchedule.merged_with.is_(None)
            ).join(Class).filter(Class.city_id == cls_city_id).count()
        else:
            db_count = ClassSchedule.query.filter(
                ClassSchedule.scheduled_date == sat,
                ClassSchedule.status.in_(['scheduled', 'completed']),
                ClassSchedule.merged_with.is_(None)
            ).count()

    assigned_count = 0
    for a_cls_id, info in assigned_map.items():
        if info['date'] == sat:
            if cls_city_id and info.get('city_id') != cls_city_id:
                continue
            assigned_count += 1
            
    total_on_sat = db_count + assigned_count

    city = cls.city_ref if cls.city_id else None
    city_name = city.name if city else '未知'
    max_rooms = city.max_classrooms if city else 99
    
    if total_on_sat >= max_rooms:
        return (0.0, True, [f'{city_name}教室已满({total_on_sat}/{max_rooms})'], [])

    # --- 合班建议：同课题检测（不依赖教室满不满） ---
    if combo1:
        current_topic_id = combo1.topic_id
        if use_cache:
            same_topic_records = _topic_idx.get((sat, current_topic_id), [])
            for rec in same_topic_records:
                if rec['class_id'] != cls.id:
                    merge_suggestions.append({
                        'target_schedule_id': rec.get('schedule_id'),
                        'target_class_id': rec['class_id'],
                        'target_class_name': rec['class_name'],
                        'topic_name': rec['topic_name'],
                    })
        else:
            same_topic_on_sat = ClassSchedule.query.filter(
                ClassSchedule.scheduled_date == sat,
                ClassSchedule.topic_id == current_topic_id,
                ClassSchedule.class_id != cls.id,
                ClassSchedule.status.in_(['scheduled'])
            ).all()
            for s in same_topic_on_sat:
                merge_suggestions.append({
                    'target_schedule_id': s.id,
                    'target_class_id': s.class_id,
                    'target_class_name': s.class_.name if s.class_ else '未知',
                    'topic_name': s.topic.name if s.topic else '未知',
                })
        # 查本轮 assigned_map 中的同课题
        for a_cls_id, a_info in assigned_map.items():
            if a_info['date'] == sat and a_cls_id != cls.id:
                if a_info.get('topic_id') == current_topic_id:
                    a_name = _cls_names.get(a_cls_id) or (a_info.get('class_name') or '?')
                    merge_suggestions.append({
                        'target_class_id': a_cls_id,
                        'target_class_name': a_name,
                        'topic_name': combo1.topic.name if combo1.topic else '未知',
                    })

    # --- 容量检查 (按城市) ---
    mem_count = 0
    for cid, data in assigned_map.items():
        if data['date'] == sat and data.get('city_id') == cls_city_id:
            mem_count += 1
    
    db_count = 0
    if precomputed:
        db_count = precomputed.get('rc_city', {}).get((sat, cls_city_id), 0)
    
    total_on_sat = db_count + mem_count
    city_cap = cls.city_ref.max_classrooms if (cls and cls.city_ref) else 99
    
    if total_on_sat >= city_cap:
        conflict_reasons.append(f'【{cls.city_ref.name if cls.city_ref else "本市"}】校区教室已满({total_on_sat}/{city_cap})')
        if not merge_suggestions:
            return (0.0, True, conflict_reasons, [])

    # --- 评分计算 ---

    # S1: 间隔分 (梯形区间安全区打分)
    # 只要在 30~45 天之间，满分 1.0；越界后急剧衰减
    if last_date:
        interval = (sat - last_date).days
        if 30 <= interval <= 45:
            interval_score = 1.0
        elif interval < 30:
            # 缩短太密集，严厉惩罚
            interval_score = max(0.0, 1.0 - (30 - interval) / 10.0)
        else: # interval > 45
            # 时间拖得越久，分数急剧下降
            interval_score = max(0.0, 1.0 - (interval - 45) / 20.0)
    else:
        interval_score = 1.0  # 新班级

    # 冲突分
    has_any_conflict = homeroom_conflict or teacher_conflict or homeroom_gap_warning or len(conflict_reasons) > 0
    conflict_score = 0.0 if has_any_conflict else 1.0

    # 均衡分
    balance_score = 1.0 / (1 + total_on_sat)

    # 月份匹配分（如果候选跨月）
    in_month_score = 1.0  # 默认都在目标月内

    # 原始日期亲和力 (Original Date Affinity)
    # 如果候选日期就是该记录原本安排的日期，给予高分以保持排课不被"破坏性重排"
    # "其他"课题(is_other)通常是后添加的手动安排，给予更高保护
    affinity_score = 0.0
    original_date_for_cls = precomputed.get('original_dates', {}).get(cls.id) if precomputed else None
    if original_date_for_cls and str(sat)[:10] == str(original_date_for_cls)[:10]:
        affinity_score = 1.0
        # "其他"课题额外保护: 提高亲和力使其更难被替换
        is_other_topic = precomputed.get('topic_is_other', {}).get(cls.id, False) if precomputed else False
        if is_other_topic:
            affinity_score = 2.0
    
    # 权重配置
    weight_interval = getattr(_cfg, 'SCORE_INTERVAL_WEIGHT', 0.4)
    weight_conflict = getattr(_cfg, 'SCORE_CONFLICT_WEIGHT', 0.4)
    weight_balance = getattr(_cfg, 'SCORE_BALANCE_WEIGHT', 0.1)
    weight_in_month = getattr(_cfg, 'SCORE_IN_MONTH_WEIGHT', 0.1)
    # 亲和力权重提高：确保在间隔安全区内原日期不被轻易替换
    weight_affinity = 0.35

    total = (
        interval_score * weight_interval +
        conflict_score * weight_conflict +
        balance_score * weight_balance +
        in_month_score * weight_in_month +
        affinity_score * weight_affinity
    )

    return (total, has_any_conflict, conflict_reasons, merge_suggestions)


def _recalculate_assignments_conflicts(assignments, constraints, homeroom_overrides=None, precomputed=None):
    """
    全量 O(N²) 纯内存对称冲突检测。不执行任何DB查询。
    所有数据从 assignments 列表（已含 combo_id, teacher_name, homeroom_id 等）中读取。
    """
    if homeroom_overrides is None:
        homeroom_overrides = {}

    # 收集所有有效分配（排除合班目标和无日期的）
    active = []
    for a in assignments:
        if a.get('is_merged_target') or not a.get('assigned_date'):
            continue
        active.append(a)

    # 按日期分组，用于 O(N²) 对称检测
    from collections import defaultdict
    by_date = defaultdict(list)
    for a in active:
        by_date[a['assigned_date']].append(a)

    # 按班级 combo_id 建立 teacher_id 快查（从 precomputed.class_infos 或 assignments 自身读取）
    # assignments 中已有 combo1_teacher_name / combo2_teacher_name，但需要 teacher_id 做精确比较
    # 从 precomputed.class_infos 的 all_combos 构建 combo_id→teacher_id 缓存
    combo_teacher_cache = {}  # {combo_id: teacher_id}
    if precomputed:
        for info in precomputed.get('class_infos', []):
            for c in info.get('all_combos', []):
                combo_teacher_cache[c.id] = c.teacher_id
            if info.get('combo1') and info['combo1']:
                combo_teacher_cache[info['combo1'].id] = info['combo1'].teacher_id
            if info.get('combo2') and info['combo2']:
                combo_teacher_cache[info['combo2'].id] = info['combo2'].teacher_id

    def _get_teacher_id(combo_id):
        if not combo_id:
            return None
        if combo_id in combo_teacher_cache:
            return combo_teacher_cache[combo_id]
        # 兜底：从DB查（仅在 precomputed 不全时触发）
        c = TeacherCourseCombo.query.get(combo_id)
        if c:
            combo_teacher_cache[combo_id] = c.teacher_id
            return c.teacher_id
        return None

    # 对每个分配做对称冲突检测
    for a in active:
        a_date = a['assigned_date']
        a_cls_id = a['class_id']
        a_hr = homeroom_overrides.get(a_cls_id, a.get('homeroom_id'))
        a_t1 = _get_teacher_id(a.get('combo_id'))
        a_t2 = _get_teacher_id(a.get('combo_id_2'))

        conflict_reasons = []

        # 与同日的其他班级做对称检查
        for other in by_date[a_date]:
            if other['class_id'] == a_cls_id:
                continue

            o_cls_id = other['class_id']
            # H3: 班主任冲突
            o_hr = homeroom_overrides.get(o_cls_id, other.get('homeroom_id'))
            if a_hr and o_hr and a_hr == o_hr:
                conflict_reasons.append(f'班主任撞课({other.get("class_name", "?")})')

            # H4: 周六讲师冲突
            o_t1 = _get_teacher_id(other.get('combo_id'))
            if a_t1 and o_t1 and a_t1 == o_t1:
                t_name = a.get('combo1_teacher_name') or '?'
                conflict_reasons.append(f'周六讲师 {t_name} 撞课({other.get("class_name", "?")})')

            # H5: 周日讲师冲突
            o_t2 = _get_teacher_id(other.get('combo_id_2'))
            if a_t2 and o_t2 and a_t2 == o_t2:
                t_name = a.get('combo2_teacher_name') or '?'
                conflict_reasons.append(f'周日讲师 {t_name} 撞课({other.get("class_name", "?")})')

        # 去重（同一冲突可能从多个维度重复报告）
        seen = set()
        unique_reasons = []
        for r in conflict_reasons:
            if r not in seen:
                seen.add(r)
                unique_reasons.append(r)

        a['conflicts'] = unique_reasons
        conflict_text = '；'.join(unique_reasons) if unique_reasons else ""
        if '班主任' in conflict_text:
            a['conflict_type'] = 'homeroom'
        elif '讲师' in conflict_text:
            a['conflict_type'] = 'teacher'
        elif conflict_text:
            a['conflict_type'] = 'other'
        else:
            a['conflict_type'] = None

        # 重算评分（基于间隔 + 冲突状态，纯计算无DB）
        has_conflict = len(unique_reasons) > 0
        interval = a.get('interval_days')
        if interval is not None:
            # 与 _score_candidate 保持一致的 30-45 天满分区间
            if 30 <= interval <= 45:
                interval_score = 1.0
            elif interval < 30:
                interval_score = max(0.0, 1.0 - (30 - interval) / 10.0)
            else:
                interval_score = max(0.0, 1.0 - (interval - 45) / 20.0)
        else:
            interval_score = 1.0
        conflict_score = 0.0 if has_conflict else 1.0
        # 简化均衡分（同日班级数）
        balance_score = 1.0 / (1 + len(by_date[a['assigned_date']]))

        a['score'] = (
            interval_score * _cfg.SCORE_INTERVAL_WEIGHT +
            conflict_score * _cfg.SCORE_CONFLICT_WEIGHT +
            balance_score * _cfg.SCORE_BALANCE_WEIGHT +
            1.0 * _cfg.SCORE_IN_MONTH_WEIGHT
        )


def _find_best_combo_for_saturday(cls, sat, last_date, all_combos,
                                   default_combo1, default_combo2,
                                   assigned_map, constraints, homeroom_overrides,
                                   user_locked=False, precomputed=None):
    """
    对某个候选周六，尝试不同的科教组合以找到最佳方案。

    策略:
      1. 先用默认组合评分
      2. 如果无冲突 → 直接返回（不浪费算力）
      3. 如果有讲师冲突 → 尝试其他组合
      4. 如果用户手动锁定了组合(user_locked) → 不尝试替换

    当所有组合的 priority 相同时，不偏袒任何组合，完全由评分决定。

    返回: (score, is_hard, reasons, merges, best_combo1, best_combo2, switched)
        switched: bool, 是否自动切换了组合
    """
    # Phase 1: 用默认组合评分
    score, is_hard, reasons, merges = _score_candidate(
        cls, sat, last_date, default_combo1, default_combo2,
        assigned_map, constraints, homeroom_overrides, precomputed=precomputed
    )

    best = {
        'score': score, 'is_hard': is_hard, 'reasons': reasons,
        'merges': merges, 'combo1': default_combo1, 'combo2': default_combo2,
        'switched': False
    }

    # 如果无冲突或用户锁定了组合或只有1个组合，直接返回
    if not is_hard or user_locked or len(all_combos) <= 1:
        return (best['score'], best['is_hard'], best['reasons'], best['merges'],
                best['combo1'], best['combo2'], best['switched'])

    # Phase 2: 检查是否有讲师冲突（只有讲师冲突换组合才有意义）
    has_teacher_conflict = any('讲师' in r for r in reasons)
    if not has_teacher_conflict:
        return (best['score'], best['is_hard'], best['reasons'], best['merges'],
                best['combo1'], best['combo2'], best['switched'])

    # Phase 3: 判断冲突在周六、周日还是两天
    sat_teacher_conflict = any('周六讲师' in r for r in reasons)
    sun_teacher_conflict = any('周日讲师' in r for r in reasons)

    # 检查是否所有组合的 priority 相同（相同时不偏袒，全部尝试）
    priorities = set(c.priority for c in all_combos)
    all_same_priority = len(priorities) <= 1

    # 构建候选组合对
    tried = set()
    if default_combo1 and default_combo2:
        tried.add((default_combo1.id, default_combo2.id))

    # combo1 候选：仅在周六有冲突或优先级全相同时尝试更换
    combo1_candidates = all_combos if (sat_teacher_conflict or all_same_priority) else [default_combo1]
    # combo2 候选：仅在周日有冲突或优先级全相同时尝试更换
    try_swap_combo2 = sun_teacher_conflict or all_same_priority

    for c1 in combo1_candidates:
        if c1 is None:
            continue
        # 为该 combo1 选配 combo2
        if try_swap_combo2:
            c2_candidates = all_combos
        else:
            # combo2 不需替换，根据 c1 确定一个默认的 c2
            c2 = c1  # 兜底
            for c in all_combos:
                if c.teacher_id != c1.teacher_id:
                    c2 = c
                    break
            c2_candidates = [c2]

        for c2 in c2_candidates:
            if c2 is None:
                continue
            pair_key = (c1.id, c2.id)
            if pair_key in tried:
                continue
            tried.add(pair_key)

            # 周六周日尽量不同讲师（除非只有1个组合可选）
            if c1.teacher_id == c2.teacher_id and len(all_combos) > 1:
                continue

            alt_score, alt_hard, alt_reasons, alt_merges = _score_candidate(
                cls, sat, last_date, c1, c2,
                assigned_map, constraints, homeroom_overrides, precomputed=precomputed
            )

            # 比较：无冲突 > 有冲突，同级别比分数
            curr_hard = best['is_hard']
            if (curr_hard and not alt_hard) or \
               (curr_hard == alt_hard and alt_score > best['score']):
                best = {
                    'score': alt_score, 'is_hard': alt_hard,
                    'reasons': alt_reasons, 'merges': alt_merges,
                    'combo1': c1, 'combo2': c2, 'switched': True
                }

    return (best['score'], best['is_hard'], best['reasons'], best['merges'],
            best['combo1'], best['combo2'], best['switched'])


def _generate_suggestions(assignment, issue_type, all_assignments):
    """为某个排课问题生成建议方案"""
    suggestions = []
    if issue_type == 'too_long':
        suggestions.append(f"检查上月是否有空余档期可提前安排")
        suggestions.append(f"如无法提前，建议与讲师沟通加课")
    elif issue_type == 'conflict':
        if '班主任' in str(assignment.get('conflicts', [])):
            suggestions.append('调整到班主任可到场的周末')
            suggestions.append('或临时更换班主任')
        if '讲师' in str(assignment.get('conflicts', [])):
            suggestions.append('为该班更换讲师')
            suggestions.append('将该班移至讲师有空的周六')
        if '教室' in str(assignment.get('conflicts', [])):
            suggestions.append('建议与同课题班级合班上课')
    elif issue_type == 'no_slot':
        suggestions.append('建议手动安排到下月初')
        suggestions.append('检查是否可以更换讲师腾出档期')
    return suggestions


def _build_quality_report(assignments):
    """构建排课质量报告（多维度加权评分）"""
    intervals = [a['interval_days'] for a in assignments if a.get('interval_days') is not None]

    # 维度1: 间隔合理性 (60%)
    if intervals:
        interval_scores = [
            max(0, 1.0 - abs(d - _cfg.TARGET_INTERVAL_DAYS) / _cfg.TARGET_INTERVAL_DAYS)
            for d in intervals
        ]
        interval_avg = int(sum(interval_scores) / len(interval_scores) * 100)
    else:
        interval_avg = 100

    # 维度2: 无冲突率 (30%) — 每个冲突扣15分
    conflict_items_pre = [a for a in assignments if a.get('conflicts')]
    conflict_free_score = max(0, 100 - len(conflict_items_pre) * 15)

    # 维度3: 覆盖率 (10%) — 成功排课占比
    total = len(assignments)
    skipped_pre = [a for a in assignments if a.get('assigned_date') is None]
    scheduled_count = total - len(skipped_pre)
    coverage_score = int(scheduled_count / total * 100) if total else 100

    overall_score = int(interval_avg * 0.6 + conflict_free_score * 0.3 + coverage_score * 0.1)

    good = [a for a in assignments if a.get('interval_days') and _cfg.MIN_INTERVAL_DAYS <= a['interval_days'] <= _cfg.MAX_INTERVAL_DAYS]
    long_ = [a for a in assignments if a.get('interval_days') and a['interval_days'] > _cfg.MAX_INTERVAL_DAYS]
    short = [a for a in assignments if a.get('interval_days') and a['interval_days'] < _cfg.MIN_INTERVAL_DAYS]
    conflict_items = [a for a in assignments if a.get('conflicts')]
    skipped = [a for a in assignments if a.get('assigned_date') is None]

    issues = []
    for a in long_:
        issues.append({
            'class_name': a.get('class_name'),
            'type': 'interval_too_long',
            'severity': 'warning',
            'detail': f"距上次课({a.get('last_date', '?')})已过{a['interval_days']}天，超出建议范围({_cfg.MAX_INTERVAL_DAYS}天)",
            'suggestions': _generate_suggestions(a, 'too_long', assignments)
        })
    for a in conflict_items:
        issues.append({
            'class_name': a.get('class_name'),
            'type': 'conflict',
            'severity': 'error',
            'detail': '；'.join(a['conflicts']),
            'suggestions': _generate_suggestions(a, 'conflict', assignments)
        })
    for a in skipped:
        issues.append({
            'class_name': a.get('class_name'),
            'type': 'no_slot',
            'severity': 'critical',
            'detail': a.get('skip_reason', '当月无可用档期'),
            'suggestions': _generate_suggestions(a, 'no_slot', assignments)
        })

    return {
        'overall_score': overall_score,
        'summary': {
            'total': len(assignments),
            'scheduled': len(assignments) - len(skipped),
            'good_interval': len(good),
            'long_interval': len(long_),
            'short_interval': len(short),
            'conflicts': len(conflict_items),
            'skipped': len(skipped),
        },
        'assignments': assignments,
        'issues': issues,
    }



def _run_best_of_n(year, month, constraints, n_rounds=30, task_id=None, **kwargs):
    """
    多次随机贪心优化：跑 n_rounds 轮，每轮随机打乱班级排序，取得分最高的结果。
    第1轮保持原始紧迫度排序，后续轮次随机打乱。
    
    性能优化：per-class 数据（紧迫度、下一课题、科教组合）只查一次DB，
    计算结果传给30轮复用，避免重复查询。
    """
    homeroom_overrides = kwargs.get('homeroom_overrides', {})
    import random
    import time
    t_start = time.time()

    def _update_progress(msg):
        """更新前端可见的进度信息"""
        if task_id and task_id in _task_store:
            _task_store[task_id]['progress'] = msg
        print(msg, flush=True)

    # ── 预计算：一次性查询所有per-class数据 ──
    _update_progress('正在预计算班级数据...')
    precomputed = _precompute_class_data(year, month, constraints, **kwargs)
    t_precompute = time.time()
    class_count = len(precomputed['class_infos'])
    _update_progress(f'预计算完成({class_count}个班级), 准备开始{n_rounds}轮优化...')

    best_result = None
    best_score = -1
    best_conflicts = float('inf')
    import copy

    for i in range(n_rounds):
        # 检查是否被取消
        if task_id and task_id in _task_store and _task_store[task_id].get('cancelled'):
            _update_progress(f'排课已取消（完成{i}/{n_rounds}轮）')
            break
        seed = None if i == 0 else random.randint(1, 999999)
        _update_progress(f'第{i+1}/{n_rounds}轮排课中... 当前最优: {best_score}分/{best_conflicts}个冲突')
        assignments, quality = _run_scheduling_algorithm(
            year, month, constraints, shuffle_seed=seed,
            precomputed=precomputed, **kwargs
        )
        # Phase 1 冲突检查
        p1_conflicts = sum(1 for a in assignments if a.get('conflicts'))
        
        if p1_conflicts == 0:
            # Phase 1 无冲突，直接评分
            _recalculate_assignments_conflicts(assignments, constraints, homeroom_overrides, precomputed=precomputed)
            quality = _build_quality_report(assignments)
            score = quality.get('overall_score', 0)
            conflict_count = quality.get('summary', {}).get('conflicts', 0)
        else:
            # Phase 1 有冲突 → 跑 Phase 2 尝试消除
            assignments = copy.deepcopy(assignments)
            _optimize_combos_per_day(assignments, constraints, precomputed=precomputed)
            _recalculate_assignments_conflicts(assignments, constraints, homeroom_overrides, precomputed=precomputed)
            quality = _build_quality_report(assignments)
            score = quality.get('overall_score', 0)
            conflict_count = quality.get('summary', {}).get('conflicts', 0)

        # 优先选冲突最少的，冲突数相同时选得分最高的
        if (conflict_count < best_conflicts or
            (conflict_count == best_conflicts and score > best_score)):
            best_conflicts = conflict_count
            best_score = score
            best_result = (assignments, quality)
        
        _update_progress(f'第{i+1}/{n_rounds}轮完成: 本轮{score}分/{conflict_count}冲突, 最优{best_score}分/{best_conflicts}冲突')

    t_rounds = time.time()
    actual_rounds = min(i + 1, n_rounds) if 'i' in dir() else n_rounds
    _update_progress(f'优化完成: {actual_rounds}轮, 最终{best_score}分/{best_conflicts}冲突')

    if best_result:
        assignments, quality = best_result

        # ── Phase 3: 冲突兜底顺延 ──
        # 如果30轮优化后仍有硬冲突（班主任撞课/讲师撞课/教室满），
        # 将这些班级顺延至下月第一个周六，清除冲突，留给下月智能排课处理。
        hard_conflict_assignments = [
            a for a in assignments
            if a.get('conflicts') and not a.get('is_merged_target')
            and any(kw in c for c in a['conflicts'] for kw in ('撞课', '请假', '教室已满'))
        ]

        if hard_conflict_assignments:
            _update_progress(f'Phase 3: {len(hard_conflict_assignments)}个班级有硬冲突，正在顺延至下月...')

            # 计算下月第一个可用周六
            if month == 12:
                next_month_1st = date(year + 1, 1, 1)
            else:
                next_month_1st = date(year, month + 1, 1)
            next_month_sat = next_month_1st
            while next_month_sat.weekday() != 5:
                next_month_sat += timedelta(days=1)
            # 跳过节假日
            while _is_blocked(next_month_sat, constraints):
                next_month_sat += timedelta(days=7)

            for a in hard_conflict_assignments:
                # 只顺延有"真正硬冲突"的（排除纯软约束警告如"连续带班"）
                hard_reasons = [c for c in a.get('conflicts', [])
                                if '撞课' in c or '请假' in c or '教室已满' in c]
                if not hard_reasons:
                    continue

                old_date = a.get('assigned_date', '?')
                a['assigned_date'] = next_month_sat.isoformat()
                a['is_overflow'] = True
                a['is_deferred'] = True
                a['deferred_reason'] = f"本月冲突无法消解({'; '.join(hard_reasons[:2])}), 已顺延至下月"
                # 清除冲突标记（它不在本月竞争资源了）
                a['conflicts'] = []
                a['conflict_type'] = None
                # 重新计算间隔
                if a.get('last_date'):
                    try:
                        last_d = date.fromisoformat(a['last_date'])
                        a['interval_days'] = (next_month_sat - last_d).days
                    except Exception:
                        pass

            # 顺延后重检剩余班级的冲突（被推走的班释放了资源，可能让其他班解冲突）
            _recalculate_assignments_conflicts(assignments, constraints, homeroom_overrides, precomputed=precomputed)
            quality = _build_quality_report(
                [a for a in assignments if not a.get('is_merged_target')]
            )
            deferred_count = sum(1 for a in assignments if a.get('is_deferred'))
            remaining_conflicts = quality.get('summary', {}).get('conflicts', 0)
            _update_progress(f'Phase 3 完成: {deferred_count}个班顺延至下月, 剩余{remaining_conflicts}个冲突')

        t_end = time.time()
        print(f'[perf] 总耗时: {t_end - t_start:.2f}s', flush=True)
        return assignments, quality
    return [], _build_quality_report([])


def _precompute_class_data(year, month, constraints, **kwargs):
    """一次性预计算所有班级的排课数据，避免30轮重复查询。
    
    month_schedules: 当月需要优化的 ClassSchedule 对象列表。
    课题从这些记录中读取（不重新推导），算法只优化日期和讲师组合。
    """
    skip_class_ids = kwargs.get('skip_class_ids', set()) or set()
    homeroom_overrides = kwargs.get('homeroom_overrides', {}) or {}
    combo_overrides = kwargs.get('combo_overrides', {}) or {}
    month_schedules = kwargs.get('month_schedules', [])  # 当月已有排课记录

    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)

    # 合班配置
    merge_groups = {}
    merged_class_set = set()
    plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
    if plan:
        merge_configs = MergeConfig.query.filter_by(monthly_plan_id=plan.id).all()
        for mc in merge_configs:
            merge_groups.setdefault(mc.primary_class_id, [])
            if mc.merged_class_id not in merge_groups[mc.primary_class_id]:
                merge_groups[mc.primary_class_id].append(mc.merged_class_id)
            merged_class_set.add(mc.merged_class_id)
    # 排课前一次性尝试从API更新该月节假日数据，失败就用本地
    _try_refresh_holidays_for_month(year, month)

    # 包含下个月的第一周作为溢出候选
    next_month_d = end_date
    while next_month_d.weekday() != 5:
        next_month_d += timedelta(days=1)
    
    pool_start = start_date
    pool_end = next_month_d + timedelta(days=7) # 包括溢出周

    candidate_saturdays = []
    d = pool_start
    while d.weekday() != 5:
        d += timedelta(days=1)
    while d < pool_end:
        sun = d + timedelta(days=1)
        # 移除了假期的强制跳过，使得算法即使在假期前后也能排期（由人工选择），如果不排算法会有冲突惩罚
        if not _is_blocked(d, constraints) and not _is_blocked(sun, constraints):
            candidate_saturdays.append(d)
        d += timedelta(days=7)

    ref_date = candidate_saturdays[0] if candidate_saturdays else start_date

    # ── 从当月已有排课记录构建 class_infos（课题固定，只优化日期和讲师） ──
    class_infos = []
    optimizing_class_ids = set()  # 正在被优化的班级ID，构建索引时排除

    # 每个班级只取一条 scheduled 记录（同一班级同月不会有两节课）
    seen_class_ids = set()
    for s in month_schedules:
        cid = s.class_id
        if cid in seen_class_ids or cid in skip_class_ids or cid in merged_class_set:
            continue
        if s.status != 'scheduled':
            continue
        seen_class_ids.add(cid)

        cls = s.class_
        if not cls or cls.status not in ('active', 'planning'):
            continue

        optimizing_class_ids.add(cid)
        urgency, last_date = _calculate_urgency(cid, ref_date)

        # 课题直接从已有记录读取
        next_topic = s.topic
        if not next_topic:
            next_topic = Topic.query.get(s.topic_id)
        if not next_topic:
            continue

        all_topics = list(cls.project.topics.order_by(Topic.sequence.asc()).all()) if cls.project else []

        # 科教组合
        all_combos = TeacherCourseCombo.query.filter_by(topic_id=next_topic.id)\
            .order_by(TeacherCourseCombo.priority.desc(), TeacherCourseCombo.id.desc()).all()

        cls_id_str = str(cid)
        user_co = combo_overrides.get(cls_id_str, {})
        skip_reason = None
        if not all_combos and not user_co.get('combo1'):
            skip_reason = f'课题「{next_topic.name}」缺少科教组合，请先在课程录入页配置讲师和课程'

        # 默认combo选择：用户覆盖 > 已有记录中的组合 > 自动选择
        combo1 = None
        combo2 = None
        user_locked_combo = False
        if user_co.get('combo1'):
            combo1 = TeacherCourseCombo.query.get(user_co['combo1'])
            user_locked_combo = True
        elif s.combo_id:
            combo1 = TeacherCourseCombo.query.get(s.combo_id)
        if user_co.get('combo2'):
            combo2 = TeacherCourseCombo.query.get(user_co['combo2'])
            user_locked_combo = True
        elif s.combo_id_2:
            combo2 = TeacherCourseCombo.query.get(s.combo_id_2)

        if not combo1 and all_combos:
            combo1 = all_combos[0]
        if not combo2:
            combo2 = combo1
            if all_combos:
                for c in all_combos:
                    if combo1 and c.teacher_id != combo1.teacher_id:
                        combo2 = c
                        break

        combos_list = [{'id': c.id, 'teacher_name': c.teacher.name if c.teacher else '?',
                        'course_name': c.course_name if c else '?'} for c in all_combos]

        class_infos.append({
            'cls': cls,
            'urgency': urgency,
            'last_date': last_date,
            'merged_ids': merge_groups.get(cid, []),
            'next_topic': next_topic,
            'all_topics': all_topics,
            'all_combos': all_combos,
            'combo1': combo1,
            'combo2': combo2,
            'user_locked_combo': user_locked_combo,
            'combos_list': combos_list,
            'skip_reason': skip_reason,
            'original_date': s.scheduled_date, # 记录原始档期亲和力用
            'original_schedule_id': s.id,  # 用于 UPDATE 时定位原记录
        })

    # ── 预计算日期维度的快查索引 ──
    # ★ 关键：排除正在被优化的班级，避免自我冲突
    #   算法会通过 assigned_map 跟踪本轮分配，不需要这些记录出现在索引中
    from collections import defaultdict
    homeroom_by_date = defaultdict(lambda: defaultdict(list))
    teacher_day1_by_date = defaultdict(lambda: defaultdict(list))
    teacher_day2_by_date = defaultdict(lambda: defaultdict(list))
    room_count_by_date_city = defaultdict(int)
    room_count_by_date = defaultdict(int)
    topic_by_date = defaultdict(list)
    class_name_cache = {}

    if candidate_saturdays:
        all_scheduled = ClassSchedule.query.filter(
            ClassSchedule.scheduled_date.in_(candidate_saturdays),
            ClassSchedule.status.in_(['scheduled', 'completed'])
        ).all()

        for s in all_scheduled:
            # ★ 排除正在被优化的班级记录
            if s.class_id in optimizing_class_ids and s.status == 'scheduled':
                # 但仍缓存 class name
                if s.class_id not in class_name_cache and s.class_:
                    class_name_cache[s.class_id] = s.class_.name
                continue

            s_date = s.scheduled_date
            s_cls = s.class_
            s_cls_name = s_cls.name if s_cls else '?'
            s_cls_id = s.class_id

            if s_cls_id not in class_name_cache and s_cls:
                class_name_cache[s_cls_id] = s_cls_name

            # H3: 班主任索引
            eff_hr = s.homeroom_override_id or (s_cls.homeroom_id if s_cls else None)
            if eff_hr:
                homeroom_by_date[s_date][eff_hr].append(s_cls_name)

            # H4: 周六讲师索引
            if s.combo_id:
                c1 = TeacherCourseCombo.query.get(s.combo_id)
                if c1 and c1.teacher_id:
                    teacher_day1_by_date[s_date][c1.teacher_id].append(s_cls_name)

            # H5: 周日讲师索引
            if s.combo_id_2:
                c2 = TeacherCourseCombo.query.get(s.combo_id_2)
                if c2 and c2.teacher_id:
                    teacher_day2_by_date[s_date][c2.teacher_id].append(s_cls_name)

            # S4: 教室容量索引 (只计非合班次记录)
            if s.merged_with is None:
                room_count_by_date[s_date] += 1
                city_id = s_cls.city_id if s_cls else None
                if city_id:
                    room_count_by_date_city[(s_date, city_id)] += 1

            # 合班建议索引
            if s.topic_id:
                topic_by_date[(s_date, s.topic_id)].append({
                    'schedule_id': s.id,
                    'class_id': s_cls_id,
                    'class_name': s_cls_name,
                    'topic_name': s.topic.name if s.topic else '未知',
                })

    # 构建原档期字典，用于计算排课亲和力
    original_dates = {info['cls'].id: info['original_date'] for info in class_infos if info.get('original_date')}
    # 构建"其他"课题标记，用于 affinity_score 的额外保护
    topic_is_other = {
        info['cls'].id: (info['next_topic'].is_other if info.get('next_topic') else False)
        for info in class_infos
    }

    return {
        'class_infos': class_infos,
        'candidate_saturdays': candidate_saturdays,
        'merge_groups': merge_groups,
        'merged_class_set': merged_class_set,
        'start_date': start_date,
        'end_date': end_date,
        'original_dates': original_dates,
        'topic_is_other': topic_is_other,
        # 性能优化索引
        'homeroom_by_date': dict(homeroom_by_date),
        'teacher_day1_by_date': dict(teacher_day1_by_date),
        'teacher_day2_by_date': dict(teacher_day2_by_date),
        'room_count_by_date_city': dict(room_count_by_date_city),
        'room_count_by_date': dict(room_count_by_date),
        'topic_by_date': dict(topic_by_date),
        'class_name_cache': class_name_cache,
    }











def _run_scheduling_algorithm(year, month, constraints, conflict_mode='smart', overrides=None, skip_class_ids=None, homeroom_overrides=None, combo_overrides=None, shuffle_seed=None, precomputed=None, **kwargs):
    """
    核心排课算法（共享逻辑：preview 和 generate 都调此函数）。
    
    参数:
        year, month: 目标月份
        constraints: 约束条件 dict
        conflict_mode: 'smart'(评分驱动) / 'postpone'(顺延) / 'mark'(标记冲突)
        overrides: {class_id_str: new_date_str} 用户在预览中的手动日期调整
        skip_class_ids: set of class IDs to skip (e.g. merged target classes)
        homeroom_overrides: {class_id: homeroom_id} 班主任覆盖
        combo_overrides: {class_id_str: {combo1: id, combo2: id}} 科教组合覆盖
        precomputed: 预计算数据(来自_precompute_class_data)，避免30轮重复查询

    返回: (assignments, quality_report)
        assignments: 排课分配详情列表
        quality_report: 质量报告 dict
    """
    if overrides is None:
        overrides = {}
    if skip_class_ids is None:
        skip_class_ids = set()
    if homeroom_overrides is None:
        homeroom_overrides = {}
    if combo_overrides is None:
        combo_overrides = {}

    # 如果有预计算数据（来自 _run_best_of_n 的30轮优化），直接使用
    if precomputed:
        class_infos = [dict(info) for info in precomputed['class_infos']]  # shallow copy
        candidate_saturdays = precomputed['candidate_saturdays']
        start_date = precomputed['start_date']
        end_date = precomputed['end_date']

        # 使用预计算的DB约束（30轮复用，不重复查询）
        db_constraints = precomputed.get('db_constraints')
        if db_constraints:
            constraints = _merge_constraints(constraints, db_constraints)
    else:
        # 无预计算时（直接调用），走原始查询
        db_constraints = _load_db_constraints(year, month)
        if db_constraints:
            constraints = _merge_constraints(constraints, db_constraints)

        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, month + 1, 1)

        # 加载合班配置
        merge_groups = {}
        merged_class_set = set()
        plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
        if plan:
            merge_configs = MergeConfig.query.filter_by(monthly_plan_id=plan.id).all()
            for mc in merge_configs:
                merge_groups.setdefault(mc.primary_class_id, [])
                if mc.merged_class_id not in merge_groups[mc.primary_class_id]:
                    merge_groups[mc.primary_class_id].append(mc.merged_class_id)
                merged_class_set.add(mc.merged_class_id)

        pool_start = start_date
        pool_end = end_date + timedelta(days=7)
        candidate_saturdays = []
        d = pool_start
        while d.weekday() != 5:
            d += timedelta(days=1)
        while d < pool_end:
            sun = d + timedelta(days=1)
            if not is_holiday(d) and not is_holiday(sun):
                if not _is_blocked(d, constraints) and not _is_blocked(sun, constraints):
                    candidate_saturdays.append(d)
            d += timedelta(days=7)

        if not candidate_saturdays:
            return [], _build_quality_report([])

        target_class_ids = kwargs.get('target_class_ids')
        if target_class_ids is not None:
            active_classes = Class.query.filter(Class.status.in_(['active', 'planning']), Class.id.in_(target_class_ids)).all()
        else:
            active_classes = Class.query.filter(Class.status.in_(['active', 'planning'])).all()
        
        ref_date = candidate_saturdays[0]
        from sqlalchemy import or_
        class_infos = []
        for cls in active_classes:
            if cls.id in skip_class_ids or cls.id in merged_class_set:
                continue
            urgency, last_date = _calculate_urgency(cls.id, ref_date)
            all_topics = list(cls.project.topics.order_by(Topic.sequence.asc()).all()) if cls.project else []
            scheduled_topic_ids = {
                s.topic_id for s in ClassSchedule.query.filter(
                    ClassSchedule.class_id == cls.id,
                    db.or_(
                        ClassSchedule.status == 'completed',
                        db.and_(
                            ClassSchedule.status == 'scheduled',
                            ClassSchedule.scheduled_date < start_date
                        )
                    )
                ).all()
            }
            next_topic = None
            for t in all_topics:
                if t.id not in scheduled_topic_ids:
                    next_topic = t
                    break
            if not next_topic:
                continue
            all_combos = TeacherCourseCombo.query.filter_by(topic_id=next_topic.id)\
                .order_by(TeacherCourseCombo.priority.desc(), TeacherCourseCombo.id.desc()).all()
            cls_id_str = str(cls.id)
            user_co = combo_overrides.get(cls_id_str, {})
            skip_reason = None
            if not all_combos and not user_co.get('combo1'):
                skip_reason = f'课题「{next_topic.name}」缺少科教组合'
            combo1 = None
            combo2 = None
            user_locked_combo = False
            if user_co.get('combo1'):
                combo1 = TeacherCourseCombo.query.get(user_co['combo1'])
                user_locked_combo = True
            if user_co.get('combo2'):
                combo2 = TeacherCourseCombo.query.get(user_co['combo2'])
                user_locked_combo = True
            if not combo1 and all_combos:
                combo1 = all_combos[0]
            if not combo2:
                combo2 = combo1
                if all_combos:
                    for c in all_combos:
                        if combo1 and c.teacher_id != combo1.teacher_id:
                            combo2 = c
                            break
            combos_list = [{'id': c.id, 'teacher_name': c.teacher.name if c.teacher else '?',
                            'course_name': c.course_name if c else '?'} for c in all_combos]
            class_infos.append({
                'cls': cls, 'urgency': urgency, 'last_date': last_date,
                'merged_ids': merge_groups.get(cls.id, []),
                'next_topic': next_topic, 'all_topics': all_topics,
                'all_combos': all_combos, 'combo1': combo1, 'combo2': combo2,
                'user_locked_combo': user_locked_combo, 'combos_list': combos_list,
                'skip_reason': skip_reason,
            })

    if not candidate_saturdays:
        return [], _build_quality_report([])

    # 按紧迫度降序 — 最急的先排，获得最好的档期
    class_infos.sort(key=lambda x: x['urgency'], reverse=True)

    # 多轮随机贪心：为保证蒙特卡洛随机性，紧急班级内部也必须打乱
    # 如果指定了 shuffle_seed，随机打乱班级顺序
    # 紧急班级（urgency > 1.5）依然在前面，但它们之间也会被打乱，防止死锁
    if shuffle_seed is not None:
        import random
        rng = random.Random(shuffle_seed)
        urgent = [c for c in class_infos if c['urgency'] > 1.5]
        normal = [c for c in class_infos if c['urgency'] <= 1.5]
        rng.shuffle(urgent)
        rng.shuffle(normal)
        class_infos = urgent + normal

    # Step 3: 为每个班级评分选日期
    assigned_map = {}
    assignments = []

    for info in class_infos:
        cls = info['cls']
        last_date = info['last_date']
        next_topic = info['next_topic']
        all_topics = info['all_topics']
        all_combos = info['all_combos']
        combo1 = info['combo1']
        combo2 = info['combo2']
        user_locked_combo = info['user_locked_combo']
        combos_list = info['combos_list']

        # 如果课题缺少科教组合 → 跳过
        if info.get('skip_reason'):
            assignments.append({
                'class_id': cls.id,
                'class_name': cls.name,
                'topic_id': next_topic.id,
                'topic_name': next_topic.name,
                'assigned_date': None,
                'skip_reason': info['skip_reason'],
                'urgency': info['urgency'],
            })
            continue

        # 记录默认组合（用于标记是否发生了自动切换）
        default_combo1 = combo1
        default_combo2 = combo2


        # 检查用户是否在预览中手动指定了日期
        override_date = overrides.get(str(cls.id))
        if override_date:
            try:
                forced_sat = date.fromisoformat(override_date)
                interval_days = (forced_sat - last_date).days if last_date else None
                assignments.append({
                    'class_id': cls.id,
                    'class_name': cls.name,
                    'project_name': cls.project.name if cls.project else None,
                    'topic_id': next_topic.id,
                    'topic_name': next_topic.name,
                    'display_topic_name': next_topic.name,
                    'combo_id': combo1.id if combo1 else None,
                    'combo_id_2': combo2.id if combo2 else None,
                    'combo1_teacher_name': combo1.teacher.name if combo1 and combo1.teacher else None,
                    'combo2_teacher_name': combo2.teacher.name if combo2 and combo2.teacher else None,
                    'combo1_course_name': combo1.course_name if combo1 else None,
                    'combo2_course_name': combo2.course_name if combo2 else None,
                    'assigned_date': forced_sat.isoformat(),
                    'last_date': last_date.isoformat() if last_date else None,
                    'interval_days': interval_days,
                    'score': 0.8,
                    'conflicts': [],
                    'conflict_type': None,
                    'merge_suggestions': [],
                    'homeroom_name': (Homeroom.query.get(homeroom_overrides[cls.id]).name if homeroom_overrides.get(cls.id) else (cls.homeroom.name if cls.homeroom else '未分配')),
                    'homeroom_id': homeroom_overrides.get(cls.id, cls.homeroom_id),
                    'city_name': cls.city.name if cls.city else None,
                    'default_city_id': cls.city_id,
                    'urgency': info['urgency'],
                    'status': 'scheduled',
                    'is_override': True,
                    'is_overflow': forced_sat >= end_date,
                    'available_combos': combos_list,
                })
                assigned_map[cls.id] = {
                    'date': forced_sat,
                    'combo1_teacher_id': combo1.teacher_id if combo1 else None,
                    'combo2_teacher_id': combo2.teacher_id if combo2 else None,
                    'topic_id': next_topic.id,
                    'city_id': cls.city_id,
                    'homeroom_id': homeroom_overrides.get(cls.id, cls.homeroom_id),
                }
                continue
            except Exception:
                pass  # 无效的 override date，按正常流程走

        # 评分选日期（所有模式共用 _find_best_combo_for_saturday 来自动换组合）
        best_sat = None
        best_score = -1
        best_conflicts = []
        best_merge_suggestions = []
        best_has_conflict = True
        best_combo1 = combo1
        best_combo2 = combo2
        combo_switched = False

        if conflict_mode == 'smart':
            # 优先选无冲突的候选，只有全冲突时选冲突最少的
            for sat in candidate_saturdays:
                score, is_hard, reasons, merges, c1, c2, switched = _find_best_combo_for_saturday(
                    cls, sat, last_date, all_combos, combo1, combo2,
                    assigned_map, constraints, homeroom_overrides,
                    user_locked=user_locked_combo, precomputed=precomputed
                )
                # 无冲突候选始终优于有冲突候选
                if best_has_conflict and not is_hard:
                    best_score = score
                    best_sat = sat
                    best_conflicts = []
                    best_merge_suggestions = merges
                    best_has_conflict = False
                    best_combo1, best_combo2, combo_switched = c1, c2, switched
                elif not best_has_conflict and is_hard:
                    continue
                elif score > best_score:
                    best_score = score
                    best_sat = sat
                    best_conflicts = reasons if is_hard else []
                    best_merge_suggestions = merges
                    best_combo1, best_combo2, combo_switched = c1, c2, switched

        elif conflict_mode == 'mark':
            # 兼容旧逻辑：找第一个无冲突位置（带自动换组合）
            for sat in candidate_saturdays:
                score, is_hard, reasons, merges, c1, c2, switched = _find_best_combo_for_saturday(
                    cls, sat, last_date, all_combos, combo1, combo2,
                    assigned_map, constraints, homeroom_overrides,
                    user_locked=user_locked_combo, precomputed=precomputed
                )
                if not is_hard:
                    best_sat = sat
                    best_score = score
                    best_conflicts = []
                    best_merge_suggestions = []
                    best_combo1, best_combo2, combo_switched = c1, c2, switched
                    break
                elif best_sat is None:
                    best_sat = sat
                    best_score = score
                    best_conflicts = reasons
                    best_merge_suggestions = merges
                    best_combo1, best_combo2, combo_switched = c1, c2, switched

        elif conflict_mode == 'postpone':
            for sat in candidate_saturdays:
                score, is_hard, reasons, merges, c1, c2, switched = _find_best_combo_for_saturday(
                    cls, sat, last_date, all_combos, combo1, combo2,
                    assigned_map, constraints, homeroom_overrides,
                    user_locked=user_locked_combo, precomputed=precomputed
                )
                if not is_hard:
                    best_sat = sat
                    best_score = score
                    best_conflicts = []
                    best_merge_suggestions = []
                    best_combo1, best_combo2, combo_switched = c1, c2, switched
                    break
        
        # 终极保底：如果真的全都有硬冲突（哪怕连溢出的那周也有硬冲突），千万不要跳过导致“隐身”
        # 强行把它塞到最后一周（通常就是下个月的溢出周），并满脸通红地展示所有冲突！
        if best_sat is None and candidate_saturdays:
            fallback_sat = candidate_saturdays[-1]  # 最后一条退路
            _, _, reasons, merges, c1, c2, switched = _find_best_combo_for_saturday(
                cls, fallback_sat, last_date, all_combos, combo1, combo2,
                assigned_map, constraints, homeroom_overrides,
                user_locked=user_locked_combo, precomputed=precomputed
            )
            best_sat = fallback_sat
            best_score = -9999
            best_conflicts = reasons if reasons else ["所有可用周末（含顺延周）均有硬冲突，必须手动处理"]
            best_merge_suggestions = merges
            best_combo1, best_combo2, combo_switched = c1, c2, switched
            best_has_conflict = True

        if best_sat:
            interval_days = (best_sat - last_date).days if last_date else None
            # 关键：用实际选中的组合写 assigned_map，确保后续班级检测到正确的讲师
            assigned_map[cls.id] = {
                'date': best_sat,
                'combo1_teacher_id': best_combo1.teacher_id if best_combo1 else None,
                'combo2_teacher_id': best_combo2.teacher_id if best_combo2 else None,
                'topic_id': next_topic.id,
                'city_id': cls.city_id,
                'homeroom_id': homeroom_overrides.get(cls.id, cls.homeroom_id),
            }
            # status 只管生命周期，conflict_type 独立标记冲突
            final_status = 'scheduled'
            conflict_type_val = None
            if best_conflicts:
                conflict_text = '；'.join(best_conflicts)
                if '班主任' in conflict_text:
                    conflict_type_val = 'homeroom'
                elif '讲师' in conflict_text:
                    conflict_type_val = 'teacher'
                elif '节假日' in conflict_text:
                    conflict_type_val = 'holiday'
                else:
                    conflict_type_val = 'other'

            # 构建自动换组合的标记信息
            combo_switch_info = None
            if combo_switched:
                combo_switch_info = {
                    'original_combo1_teacher': default_combo1.teacher.name if default_combo1 and default_combo1.teacher else None,
                    'original_combo2_teacher': default_combo2.teacher.name if default_combo2 and default_combo2.teacher else None,
                }

            assignments.append({
                'class_id': cls.id,
                'class_name': cls.name,
                'project_name': cls.project.name if cls.project else None,
                'topic_id': next_topic.id,
                'topic_name': next_topic.name,
                'display_topic_name': next_topic.name,
                'combo_id': best_combo1.id if best_combo1 else None,
                'combo_id_2': best_combo2.id if best_combo2 else None,
                'combo1_teacher_name': best_combo1.teacher.name if best_combo1 and best_combo1.teacher else None,
                'combo2_teacher_name': best_combo2.teacher.name if best_combo2 and best_combo2.teacher else None,
                'combo1_course_name': best_combo1.course_name if best_combo1 else None,
                'combo2_course_name': best_combo2.course_name if best_combo2 else None,
                'assigned_date': best_sat.isoformat(),
                'last_date': last_date.isoformat() if last_date else None,
                'interval_days': interval_days,
                'score': best_score,
                'conflicts': best_conflicts,
                'conflict_type': conflict_type_val,
                'merge_suggestions': best_merge_suggestions,
                'homeroom_name': (Homeroom.query.get(homeroom_overrides[cls.id]).name if homeroom_overrides.get(cls.id) else (cls.homeroom.name if cls.homeroom else '未分配')),
                'homeroom_id': homeroom_overrides.get(cls.id, cls.homeroom_id),
                'city_name': cls.city_ref.name if cls.city_ref else None,
                'default_city_id': cls.city_id,
                'urgency': info['urgency'],
                'status': final_status,
                'is_overflow': best_sat >= end_date,
                'available_combos': combos_list,
                'combo_switched': combo_switched,
                'combo_switch_info': combo_switch_info,
            })
        else:
            assignments.append({
                'class_id': cls.id,
                'class_name': cls.name,
                'project_name': cls.project.name if cls.project else None,
                'topic_id': next_topic.id,
                'topic_name': next_topic.name,
                'display_topic_name': next_topic.name,
                'assigned_date': None,
                'last_date': last_date.isoformat() if last_date else None,
                'interval_days': None,
                'score': 0,
                'conflicts': [],
                'conflict_type': None,
                'merge_suggestions': [],
                'homeroom_name': (Homeroom.query.get(homeroom_overrides[cls.id]).name if homeroom_overrides.get(cls.id) else (cls.homeroom.name if cls.homeroom else '未分配')),
                'homeroom_id': homeroom_overrides.get(cls.id, cls.homeroom_id),
                'city_name': cls.city.name if cls.city else None,
                'default_city_id': cls.city_id,
                'urgency': info['urgency'],
                'skip_reason': '所有候选日期均不可用',
                'available_combos': combos_list,
            })

    # 阶段二在 _run_best_of_n 中执行（只跑一次，不在30轮循环里）
    quality_report = _build_quality_report(assignments)
    return assignments, quality_report


@schedule_bp.route('/generate-preview', methods=['POST'])
def generate_schedule_preview():
    """预览排课结果（异步模式）- 立即返回 task_id，前端轮询 task-status"""
    data = request.get_json()
    year = data.get('year')
    month = data.get('month')
    constraints = data.get('constraints', {})
    conflict_mode = data.get('conflict_mode', 'smart')
    merges = data.get('merges', [])
    combo_overrides_raw = data.get('combo_overrides', {})
    raw_hr_overrides = data.get('homeroom_overrides', {})
    homeroom_overrides = {int(k): v for k, v in raw_hr_overrides.items()} if raw_hr_overrides else {}

    if not year or not month:
        return jsonify({'error': 'Missing year/month'}), 400

    # 发布保护
    existing_plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
    if existing_plan and existing_plan.status == 'published':
        return jsonify({'error': '本月计划已发布，请先取消发布再重新排课'}), 409

    task_id = str(uuid.uuid4())
    _task_store[task_id] = {'status': 'running', 'progress': '正在初始化排课算法...', 'result': None, 'error': None}

    # 在后台线程中执行排课
    app = current_app._get_current_object()
    request_data = {
        'year': year, 'month': month, 'constraints': constraints,
        'conflict_mode': conflict_mode, 'merges': merges,
        'combo_overrides_raw': combo_overrides_raw,
        'homeroom_overrides': homeroom_overrides,
    }

    def _run_async(app, task_id, req_data):
        with app.app_context():
            try:
                _task_store[task_id]['progress'] = '正在预计算班级数据...'
                yr, mo = req_data['year'], req_data['month']
                constr = req_data['constraints']
                cm = req_data['conflict_mode']
                merges_list = req_data['merges']
                co_raw = req_data['combo_overrides_raw']
                hr_overrides = req_data['homeroom_overrides']

                start_d = date(yr, mo, 1)
                end_d = date(yr, mo + 1, 1) if mo < 12 else date(yr + 1, 1, 1)

                # 读取当月已有的排课记录（智能排课只优化这些记录的日期和讲师）
                month_schedules = ClassSchedule.query.filter(
                    ClassSchedule.scheduled_date >= start_d,
                    ClassSchedule.scheduled_date < end_d,
                    ClassSchedule.status == 'scheduled'
                ).all()

                merged_target_ids = set()
                for m in merges_list:
                    merged_target_ids.add(m.get('target_class_id'))

                _task_store[task_id]['progress'] = '正在运行30轮优化算法...'
                assignments, quality_report = _run_best_of_n(
                    yr, mo, constr, n_rounds=30,
                    task_id=task_id,
                    conflict_mode=cm,
                    skip_class_ids=merged_target_ids,
                    homeroom_overrides=hr_overrides,
                    combo_overrides=co_raw,
                    month_schedules=month_schedules
                )

                # 检查是否被取消
                if _task_store[task_id].get('cancelled'):
                    _task_store[task_id] = {
                        'status': 'error', 'progress': '已取消',
                        'result': None, 'error': '排课任务已被用户取消'
                    }
                    return

                _task_store[task_id]['progress'] = '正在处理合班信息...'
                for m in merges_list:
                    for a in assignments:
                        if a.get('class_id') == m.get('class_id'):
                            a['is_merged'] = True
                            a['merged_with_class_id'] = m.get('target_class_id')
                            a['merged_with_class_name'] = m.get('target_class_name', '')
                            if m.get('combo_id'):
                                a['combo_id'] = m['combo_id']
                            if m.get('combo_id_2'):
                                a['combo_id_2'] = m['combo_id_2']
                            break

                for m in merges_list:
                    target_cls = Class.query.get(m.get('target_class_id'))
                    if target_cls:
                        source_a = next((a for a in assignments if a.get('class_id') == m.get('class_id')), None)
                        assignments.append({
                            'class_id': m['target_class_id'],
                            'class_name': target_cls.name,
                            'project_name': target_cls.project.name if target_cls.project else None,
                            'topic_name': source_a.get('topic_name', '') if source_a else '',
                            'assigned_date': m.get('date') or (source_a.get('assigned_date') if source_a else None),
                            'is_merged_target': True,
                            'merged_into_class_id': m.get('class_id'),
                            'merged_into_class_name': source_a.get('class_name', '') if source_a else '',
                            'interval_days': None,
                            'conflicts': [],
                            'conflict_type': None,
                            'skip_reason': None,
                        })

                _task_store[task_id]['progress'] = '正在计算班主任可用性...'
                all_hr = Homeroom.query.all()
                for a in assignments:
                    if a.get('is_merged_target'):
                        continue
                    assigned_date_str = a.get('assigned_date')
                    if not assigned_date_str:
                        a['homeroom_availability'] = [{'id': h.id, 'name': h.name, 'busy': False, 'busy_class': None} for h in all_hr]
                        continue
                    busy_homerooms = {}
                    for other_a in assignments:
                        if other_a.get('class_id') == a.get('class_id') or other_a.get('is_merged_target'):
                            continue
                        if other_a.get('assigned_date') == assigned_date_str:
                            other_hr_id = other_a.get('homeroom_id')
                            if other_hr_id:
                                busy_homerooms[other_hr_id] = other_a.get('class_name', '?')
                    a['homeroom_availability'] = [
                        {
                            'id': h.id,
                            'name': h.name,
                            'busy': h.id in busy_homerooms,
                            'busy_class': busy_homerooms.get(h.id)
                        }
                        for h in all_hr
                    ]

                quality_report = _build_quality_report(
                    [a for a in assignments if not a.get('is_merged_target')]
                )

                # 预览模式不修改数据库，无需 rollback

                all_saturdays = _get_all_saturdays_with_reasons(yr, mo, constr)

                _task_store[task_id] = {
                    'status': 'done',
                    'progress': '排课完成！',
                    'result': {
                        'success': True,
                        'preview': assignments,
                        'quality_report': quality_report,
                        'candidate_saturdays': [
                            sat.isoformat() for sat in _get_candidate_saturdays(yr, mo, constr)
                        ],
                        'all_saturdays': all_saturdays,
                        'merges_applied': len(merges_list),
                        'config': {
                            'target_interval_days': _cfg.TARGET_INTERVAL_DAYS,
                        },
                    },
                    'error': None
                }
            except Exception as e:
                db.session.rollback()
                import traceback
                traceback.print_exc()
                _task_store[task_id] = {
                    'status': 'error',
                    'progress': '排课失败',
                    'result': None,
                    'error': str(e)
                }

    thread = threading.Thread(target=_run_async, args=(app, task_id, request_data))
    thread.daemon = True
    thread.start()

    return jsonify({'task_id': task_id})


@schedule_bp.route('/evaluate-preview', methods=['POST'])
def evaluate_preview():
    """
    轻量评估预览：不重新排课，只重新检测冲突和计算评分。
    前端在调整日期/组合/班主任后调用此接口，避免完整重排。
    """
    data = request.get_json()
    assignments_raw = data.get('assignments', [])
    constraints = data.get('constraints', {})
    raw_hr_overrides = data.get('homeroom_overrides', {})
    combo_overrides_raw = data.get('combo_overrides', {})
    merges = data.get('merges', [])
    homeroom_overrides = {int(k): v for k, v in raw_hr_overrides.items()} if raw_hr_overrides else {}

    # 自动加载数据库中当月的约束条件
    all_dates_for_month = [a['assigned_date'] for a in assignments_raw if a.get('assigned_date')]
    if all_dates_for_month:
        first_date = date.fromisoformat(min(all_dates_for_month))
        db_constraints = _load_db_constraints(first_date.year, first_date.month)
        if db_constraints:
            constraints = _merge_constraints(constraints, db_constraints)
    if not assignments_raw:
        return jsonify({'assignments': [], 'quality_report': {'overall_score': 100, 'summary': {'good_interval': 0, 'long_interval': 0, 'conflicts': 0, 'skipped': 0}}})

    try:
        # 关键修复：和 generate-preview 一样，先在 session 中删除当月草稿记录再评估
        # 否则 _score_candidate 查 DB 会把预览自身的记录当成冲突
        # 收集所有涉及的日期来确定月份范围
        all_dates = [a['assigned_date'] for a in assignments_raw if a.get('assigned_date')]
        if all_dates:
            min_date = date.fromisoformat(min(all_dates))
            max_date = date.fromisoformat(max(all_dates))
            # 删除范围覆盖所有涉及的月份
            start_d = date(min_date.year, min_date.month, 1)
            end_month = max_date.month + 1 if max_date.month < 12 else 1
            end_year = max_date.year if max_date.month < 12 else max_date.year + 1
            end_d = date(end_year, end_month, 1)

            ClassSchedule.query.filter(
                ClassSchedule.scheduled_date >= start_d,
                ClassSchedule.scheduled_date < end_d,
                ClassSchedule.status.in_(['scheduled'])
            ).delete(synchronize_session='fetch')
            db.session.flush()

        # 构建 assigned_map 用于冲突检测（排除合班目标）
        assigned_map = {}
        for a in assignments_raw:
            if a.get('is_merged_target'):
                continue
            if not a.get('assigned_date'):
                continue
            cls_id = a['class_id']
            combo1_teacher_id = None
            combo2_teacher_id = None
            if a.get('combo_id'):
                c1 = TeacherCourseCombo.query.get(a['combo_id'])
                if c1:
                    combo1_teacher_id = c1.teacher_id
            if a.get('combo_id_2'):
                c2 = TeacherCourseCombo.query.get(a['combo_id_2'])
                if c2:
                    combo2_teacher_id = c2.teacher_id
            assigned_map[cls_id] = {
                'date': date.fromisoformat(a['assigned_date']),
                'combo1_teacher_id': combo1_teacher_id,
                'combo2_teacher_id': combo2_teacher_id,
                'topic_id': a.get('topic_id'),
            }

        # 收集合班 target IDs
        merged_target_ids = set(m.get('target_class_id') for m in merges)

        # 重新评估每个 assignment 的冲突
        evaluated = []
        all_hr = Homeroom.query.all()
        for a in assignments_raw:
            if a.get('is_merged_target'):
                evaluated.append(a)
                continue
            if not a.get('assigned_date'):
                evaluated.append(a)
                continue

            cls_id = a['class_id']
            cls = Class.query.get(cls_id)
            if not cls:
                evaluated.append(a)
                continue

            sat = date.fromisoformat(a['assigned_date'])
            last_date = date.fromisoformat(a['last_date']) if a.get('last_date') else None

            # 获取 combo 对象
            combo1 = TeacherCourseCombo.query.get(a['combo_id']) if a.get('combo_id') else None
            combo2 = TeacherCourseCombo.query.get(a['combo_id_2']) if a.get('combo_id_2') else None

            # 构建不含自身的 assigned_map
            other_map = {k: v for k, v in assigned_map.items() if k != cls_id}

            # 调用现有的评分函数
            score, is_hard, conflict_reasons, merge_suggestions = _score_candidate(
                cls, sat, last_date, combo1, combo2,
                other_map, constraints, homeroom_overrides
            )

            # 更新 interval_days
            interval_days = (sat - last_date).days if last_date else None

            # 确定 conflict_type
            conflict_type_val = None
            if conflict_reasons:
                conflict_text = '；'.join(conflict_reasons)
                if '班主任' in conflict_text:
                    conflict_type_val = 'homeroom'
                elif '讲师' in conflict_text:
                    conflict_type_val = 'teacher'
                elif '节假日' in conflict_text:
                    conflict_type_val = 'holiday'
                else:
                    conflict_type_val = 'other'

            # 更新 homeroom 信息
            effective_hr_id = homeroom_overrides.get(cls_id, cls.homeroom_id)
            effective_hr = Homeroom.query.get(effective_hr_id) if effective_hr_id else None

            a['conflicts'] = conflict_reasons
            a['conflict_type'] = conflict_type_val
            a['interval_days'] = interval_days
            a['score'] = score
            a['merge_suggestions'] = merge_suggestions
            a['homeroom_name'] = effective_hr.name if effective_hr else (cls.homeroom.name if cls.homeroom else '未分配')
            a['homeroom_id'] = effective_hr_id

            # 班主任可用性
            assigned_date_str = a['assigned_date']
            busy_homerooms = {}
            for other_a in assignments_raw:
                if other_a.get('class_id') == cls_id or other_a.get('is_merged_target'):
                    continue
                if other_a.get('assigned_date') == assigned_date_str:
                    other_hr_id = other_a.get('homeroom_id')
                    if other_hr_id:
                        busy_homerooms[other_hr_id] = other_a.get('class_name', '?')
            a['homeroom_availability'] = [
                {
                    'id': h.id,
                    'name': h.name,
                    'busy': h.id in busy_homerooms,
                    'busy_class': busy_homerooms.get(h.id)
                }
                for h in all_hr
            ]

            evaluated.append(a)

        # 回滚！评估不持久化任何更改
        db.session.rollback()

        # 构建质量报告
        non_merged = [a for a in evaluated if not a.get('is_merged_target')]
        quality_report = _build_quality_report(non_merged)

        return jsonify({
            'success': True,
            'preview': evaluated,
            'quality_report': quality_report,
        })
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def _get_candidate_saturdays(year, month, constraints):
    """获取候选周六列表（供前端预览调整时使用）"""
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)
    pool_start = start_date - timedelta(days=7)
    pool_end = end_date + timedelta(days=7)
    result = []
    d = pool_start
    while d.weekday() != 5:
        d += timedelta(days=1)
    while d < pool_end:
        sun = d + timedelta(days=1)
        if not is_holiday(d) and not is_holiday(sun):
            if not _is_blocked(d, constraints) and not _is_blocked(sun, constraints):
                result.append(d)
        d += timedelta(days=7)
    return result


def _get_all_saturdays_with_reasons(year, month, constraints):
    """获取月份范围内所有周六及其可用状态和原因，供前端显示"""
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)
    pool_start = start_date - timedelta(days=7)
    pool_end = end_date + timedelta(days=7)

    result = []
    d = pool_start
    while d.weekday() != 5:
        d += timedelta(days=1)
    while d < pool_end:
        sun = d + timedelta(days=1)
        item = {'date': d.isoformat(), 'sun_date': sun.isoformat(), 'available': True, 'reason': None}

        if is_holiday(d):
            item['available'] = False
            item['reason'] = f'{d.strftime("%m/%d")}(六) 节假日'
        elif is_holiday(sun):
            item['available'] = False
            item['reason'] = f'{sun.strftime("%m/%d")}(日) 节假日'
        elif _is_blocked(d, constraints):
            item['available'] = False
            item['reason'] = '约束条件排除'
        elif _is_blocked(sun, constraints):
            item['available'] = False
            item['reason'] = '约束条件排除(周日)'

        result.append(item)
        d += timedelta(days=7)
    return result


@schedule_bp.route('/generate', methods=['POST'])
def generate_schedule():
    """根据约束条件生成/重新生成月度排课（智能评分驱动）"""
    data = request.get_json()
    year = data.get('year')
    month = data.get('month')
    constraints = data.get('constraints', {})
    conflict_mode = data.get('conflict_mode', 'smart')
    overrides = data.get('overrides', {})
    combo_overrides = data.get('combo_overrides', {})  # {class_id_str: {combo1: id, combo2: id}}
    merges = data.get('merges', [])
    raw_hr_overrides = data.get('homeroom_overrides', {})
    homeroom_overrides = {int(k): v for k, v in raw_hr_overrides.items()} if raw_hr_overrides else {}
    raw_loc_overrides = data.get('location_overrides', {})
    location_overrides = {int(k): v for k, v in raw_loc_overrides.items()} if raw_loc_overrides else {}

    if not year or not month:
        return jsonify({'error': 'Missing year/month'}), 400

    # 验证智能排课中的手动地点调整，是否违反了历史地点限制
    for cid, loc_id in location_overrides.items():
        if loc_id:
            loc_id = int(loc_id)
            cls = Class.query.get(cid)
            if cls and cls.city_id and loc_id != cls.city_id:
                # _get_used_non_default_locations 需要排除当前月份该班级原来的排课记录？
                # 由于智能排课会覆盖原有 scheduled，这里不排除 id 也没有很大关系，但严格点最好不被自己卡住
                used_locs = _get_used_non_default_locations(cid, cls.city_id)
                if loc_id in used_locs:
                    # 检查是否被该班级在我们要覆写的月份以外的记录用过，逻辑较复杂，这里简化提示
                    loc = City.query.get(loc_id)
                    return jsonify({'error': f'班级「{cls.name}」已去过地点「{loc.name if loc else loc_id}」，不允许手动调整为此地点'}), 400

    # 发布保护：已发布的月度计划不允许重新排课
    existing_plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
    if existing_plan and existing_plan.status == 'published':
        return jsonify({'error': '本月计划已发布，请先取消发布再重新排课'}), 409

    try:
        # 1. 读取当月排课记录（不删除！智能排课只优化日期和讲师）
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, month + 1, 1)

        # 读取当月 scheduled 记录
        month_schedules = ClassSchedule.query.filter(
            ClassSchedule.scheduled_date >= start_date,
            ClassSchedule.scheduled_date < end_date,
            ClassSchedule.status == 'scheduled'
        ).all()

        # 清除旧的合班关联（合班会重新建立）
        month_schedule_ids = [s.id for s in month_schedules]
        if month_schedule_ids:
            ClassSchedule.query.filter(
                ClassSchedule.merged_with.in_(month_schedule_ids)
            ).update({ClassSchedule.merged_with: None, ClassSchedule.merge_snapshot: None}, synchronize_session=False)
            for s in month_schedules:
                if s.notes and '合班主记录' in (s.notes or ''):
                    s.notes = None
                    s.merge_snapshot = None

        # 2. 合班：收集被合并的 target 班级 ID
        merged_target_ids = set()
        for m in merges:
            merged_target_ids.add(m.get('target_class_id'))

        # 3. 运行核心排课算法（纯内存优化，不修改数据库）
        assignments, quality_report = _run_best_of_n(
            year, month, constraints, n_rounds=30,
            conflict_mode=conflict_mode,
            overrides=overrides,
            skip_class_ids=merged_target_ids,
            homeroom_overrides=homeroom_overrides,
            combo_overrides=combo_overrides,
            month_schedules=month_schedules
        )

        # 4. 将优化结果写回数据库（UPDATE 已有记录，不删除不重建）
        # 建立 class_id → original schedule 的映射
        schedule_by_class = {}
        for s in month_schedules:
            if s.status == 'scheduled' and s.class_id not in schedule_by_class:
                schedule_by_class[s.class_id] = s

        generated_count = 0
        skipped_classes_info = []

        for a in assignments:
            if a.get('assigned_date') is None:
                skipped_classes_info.append({
                    'class_name': a.get('class_name', '未知'),
                    'reason': a.get('skip_reason', '无可用档期')
                })
                continue

            assigned_sat = date.fromisoformat(a['assigned_date'])

            # 应用用户的组合调整
            cls_id_str = str(a['class_id'])
            final_combo_id = a.get('combo_id')
            final_combo_id_2 = a.get('combo_id_2')
            if cls_id_str in combo_overrides:
                co = combo_overrides[cls_id_str]
                if 'combo1' in co:
                    final_combo_id = co['combo1']
                if 'combo2' in co:
                    final_combo_id_2 = co['combo2']

            # 确定冲突类型
            conflict_type_val = a.get('conflict_type')
            final_notes = 'AI智能排课'
            if a.get('conflicts'):
                final_notes = '；'.join(a['conflicts'])

            # 班主任临时覆盖
            hr_override = homeroom_overrides.get(a['class_id'])
            # 地点临时覆盖
            loc_override = location_overrides.get(a['class_id'])

            # ★ UPDATE 已有记录（不删除不重建）
            existing = schedule_by_class.get(a['class_id'])
            if existing:
                existing.scheduled_date = assigned_sat
                existing.combo_id = final_combo_id
                existing.combo_id_2 = final_combo_id_2
                existing.conflict_type = conflict_type_val
                existing.notes = final_notes
                existing.homeroom_override_id = hr_override
                if loc_override:
                    existing.location_id = loc_override
            else:
                # 极端情况兜底：原记录不存在时才创建新记录
                _cls_for_loc = Class.query.get(a['class_id'])
                new_schedule = ClassSchedule(
                    class_id=a['class_id'],
                    topic_id=a['topic_id'],
                    combo_id=final_combo_id,
                    combo_id_2=final_combo_id_2,
                    scheduled_date=assigned_sat,
                    week_number=0,
                    status='scheduled',
                    conflict_type=conflict_type_val,
                    notes=final_notes,
                    homeroom_override_id=hr_override,
                    location_id=loc_override or (_cls_for_loc.city_id if _cls_for_loc else None)
                )
                db.session.add(new_schedule)
            # IMPORTANT: Store a reference inside `a` so we can access its location_id during merge syncing
            a['_generated_location_id'] = loc_override or (_cls_for_loc.city_id if 'a' in locals() and _cls_for_loc else None)
            if existing: a['_generated_location_id'] = existing.location_id
            else: a['_generated_location_id'] = new_schedule.location_id
            generated_count += 1

        db.session.flush()  # 让 new_schedule 获得 ID

        # 4. 合班记录写入
        import json as _json
        # 先统计每个 source 班级的所有 target 名称（处理多合班）
        source_targets = {}  # {source_class_id: [target_name, ...]}
        for m in merges:
            sid = m.get('class_id')
            tid = m.get('target_class_id')
            target_cls = Class.query.get(tid)
            tname = target_cls.name if target_cls else '?'
            source_targets.setdefault(sid, []).append(tname)

        for m in merges:
            source_class_id = m.get('class_id')
            target_class_id = m.get('target_class_id')
            merge_combo_id = m.get('combo_id')
            merge_combo_id_2 = m.get('combo_id_2')

            # 找到 source 班级的排课记录
            source_schedule = ClassSchedule.query.filter(
                ClassSchedule.class_id == source_class_id,
                ClassSchedule.scheduled_date >= start_date,
                ClassSchedule.scheduled_date < end_date,
                ClassSchedule.status.in_(['scheduled'])
            ).first()

            if source_schedule:
                # 为 target 班级创建合班记录
                target_cls = Class.query.get(target_class_id)
                target_topic_id = source_schedule.topic_id

                # 找到被合并的主课表数据以同步 location_id
                source_a_dict = next((a for a in assignments if a.get('class_id') == source_class_id), None)
                merged_loc_id = source_a_dict.get('_generated_location_id') if source_a_dict else None

                # 保存 source 的原始状态快照（拆分时恢复用）
                source_snapshot = _json.dumps({
                    'scheduled_date': source_schedule.scheduled_date.isoformat(),
                    'combo_id': source_schedule.combo_id,
                    'combo_id_2': source_schedule.combo_id_2,
                    'status': source_schedule.status,
                    'notes': source_schedule.notes or 'AI智能排课',
                    'homeroom_override_id': source_schedule.homeroom_override_id,
                    'location_id': source_schedule.location_id
                })

                merged_schedule = ClassSchedule(
                    class_id=target_class_id,
                    topic_id=target_topic_id,
                    combo_id=merge_combo_id or source_schedule.combo_id,
                    combo_id_2=merge_combo_id_2 or source_schedule.combo_id_2,
                    scheduled_date=source_schedule.scheduled_date,
                    week_number=0,
                    status='scheduled',
                    merged_with=source_schedule.id,
                    homeroom_override_id=source_schedule.homeroom_override_id,
                    location_id=merged_loc_id,
                    merge_snapshot=_json.dumps({
                        'scheduled_date': source_schedule.scheduled_date.isoformat(),
                        'combo_id': merge_combo_id or source_schedule.combo_id,
                        'combo_id_2': merge_combo_id_2 or source_schedule.combo_id_2,
                        'status': 'scheduled',
                        'notes': None,
                        'homeroom_override_id': None,
                        'location_id': None
                    }),
                    notes=f'合班至 {source_schedule.class_.name if source_schedule.class_ else "?"}'
                )
                db.session.add(merged_schedule)
                generated_count += 1

                # 主记录写入合班信息（使用统计好的所有 target 名称）
                source_schedule.merge_snapshot = source_snapshot
                all_targets = source_targets.get(source_class_id, [])
                all_names = '、'.join(all_targets)
                source_schedule.notes = f'合班主记录（含 {all_names}）'

        db.session.flush()

        # 对本月所有日期重检冲突（新排课可能与已有排课冲突）
        all_dates = set()
        month_schedules = ClassSchedule.query.filter(
            ClassSchedule.scheduled_date >= start_date,
            ClassSchedule.scheduled_date < end_date
        ).all()
        for ms in month_schedules:
            if ms.scheduled_date:
                all_dates.add(ms.scheduled_date)
        _recheck_conflicts_for_dates(all_dates)
        db.session.commit()

        # 4. 自动创建/更新月度计划为草稿
        plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
        if not plan:
            plan = MonthlyPlan(year=year, month=month, status='draft')
            db.session.add(plan)
        plan.updated_at = datetime.now()
        db.session.commit()

        # 清理全库残留记录
        stale_count = _cleanup_stale_scheduled_records()

        # 排课日期变动后重排所有相关班级的课题序号（含合班目标班级）
        affected_class_ids = set(a['class_id'] for a in assignments if a.get('assigned_date'))
        for m in merges:
            tid = m.get('target_class_id')
            if tid:
                affected_class_ids.add(tid)
        for cid in affected_class_ids:
            _resequence_topics_by_date(cid)

        result_msg = f'已生成 {generated_count} 节课程安排'
        if skipped_classes_info:
            result_msg += f'。有 {len(skipped_classes_info)} 个班级未排课'

        return jsonify({
            'success': True,
            'message': result_msg,
            'skipped': skipped_classes_info,
            'plan': plan.to_dict(),
            'quality_report': quality_report,
        })

    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ==================== 月度计划发布管理 ====================

@schedule_bp.route('/monthly-plan/publish', methods=['POST'])
def publish_monthly_plan():
    """发布月度计划"""
    data = request.get_json()
    year = data.get('year')
    month = data.get('month')
    if not year or not month:
        return jsonify({'error': '缺少 year/month'}), 400

    plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
    if not plan:
        return jsonify({'error': '该月度计划不存在'}), 404
    if plan.status == 'published':
        return jsonify({'error': '该月度计划已经是发布状态'}), 400

    plan.status = 'published'
    plan.published_at = datetime.now()
    db.session.commit()
    return jsonify({'success': True, 'message': f'{year}年{month}月计划已发布', 'plan': plan.to_dict()})


@schedule_bp.route('/monthly-plan/unpublish', methods=['POST'])
def unpublish_monthly_plan():
    """取消发布月度计划（回到草稿状态）"""
    data = request.get_json()
    year = data.get('year')
    month = data.get('month')
    if not year or not month:
        return jsonify({'error': '缺少 year/month'}), 400

    plan = MonthlyPlan.query.filter_by(year=year, month=month).first()
    if not plan:
        return jsonify({'error': '该月度计划不存在'}), 404
    if plan.status != 'published':
        return jsonify({'error': '该月度计划未发布，无需取消'}), 400

    plan.status = 'draft'
    plan.published_at = None
    db.session.commit()
    return jsonify({'success': True, 'message': f'{year}年{month}月计划已取消发布', 'plan': plan.to_dict()})
